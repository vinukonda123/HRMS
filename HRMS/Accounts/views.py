# Create your views here.

from django.core.paginator import Paginator
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from openpyxl import load_workbook
from .models import Rbi_Data_Import

from django.utils import timezone
from datetime import timedelta


from django.shortcuts import render, get_object_or_404
from .models import Projects


from django.core.paginator import Paginator
from django.http import JsonResponse

from datetime import time
from datetime import time, datetime, timedelta
from datetime import date, timedelta
from collections import OrderedDict
from django.core.mail import send_mail
from django.db.models import Count, Case, When, Value
from django.db.models.functions import Concat
from django.db.models import Q, Count
from datetime import date
from django.db import connection
from datetime import date
import random
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count
from datetime import date
from django.db.models import F
from django.db.models import Q
from django.db.models.functions import ExtractMonth, ExtractDay
import base64
import datetime
from django.contrib.auth import get_user_model
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404
from .forms import LeaveRequestForm
from .forms import CustomSignupForm,EmployeeMasterForm,CustomUpdateForm
from django.shortcuts import render,redirect
from .forms import RegisterForm
from . models import EmployeeMaster,Projects,Tasks,Shift,Company_Info,Holidays,Tickets,Expenses,Source,Priority,Status, Leavecategory,Attendance,Permission,Email_inbox,EmailReply,EmailForword,PermissionApproval,Immegration,Emergency_Contacts,Social_Profile,Documents_All,Qualifications,Work_Experience,Bank_Account,Basic_Salary,Allowances,Commissions,Loan,Statutory_Deduction,Other_Payments,Over_Time,Company,Company_Policy,Announcements,Designation,Location,Department,Promotions,Awards,Travel,Transfer,Resignation,Complains,Warnings,Terminations,Trainer,Training_Type,Training_List,Category,Assets,Job_Post,Job_Candidate,Job_Interview,Client,Events,Financial_year,Meetings,My_Tasks
from django.contrib import messages
from django.contrib.auth import login as auth_login , authenticate, logout
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from Accounts.models import CustomUser,LeaveRequest,Notification,Public_holidays#Permission,PermissionApproval,H
from django.contrib import messages
from .models import Event,Pro
from django.urls import reverse
import json
from django.contrib.auth.tokens import default_token_generator
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.shortcuts import render, redirect
from django.contrib.auth.views import PasswordResetView

def index(request,id):
    data_id=id # Replace this with the actual value you want to pas1

    data1=CustomUser.objects.get(id=data_id)
    emp_obj=EmployeeMaster.objects.get(user_id=data1.id)
    data2=Immegration.objects.filter(user_id=data1.id).all()
    data3=Emergency_Contacts.objects.filter(user_id=data1.id).all()
    data4=Documents_All.objects.filter(user_id=data1.id).all()
    data5=Qualifications.objects.filter(user_id=data1.id).all()
    data6=Work_Experience.objects.filter(user_id=data1.id).all()
    data7=Bank_Account.objects.filter(user_id=data1.id).all()

    data8=Basic_Salary.objects.filter(user_id=data1.id).all()
    data9=Allowances.objects.filter(user_id=data1.id).all()
    data10=Commissions.objects.filter(user_id=data1.id).all()
    data11=Loan.objects.filter(user_id=data1.id).all()
    data12=Statutory_Deduction.objects.filter(user_id=data1.id).all()
    data13=Other_Payments.objects.filter(user_id=data1.id).all()
    data14=Over_Time.objects.filter(user_id=data1.id).all()
    data00=Social_Profile.objects.filter(user_id=data1.id).all()
    data15 = Company_Policy.objects.all()
    data16 = Announcements.objects.all()
    data17 = Designation.objects.all()
    data18 = Location.objects.all()
    data19 = Department.objects.all()
    data20 = Company.objects.all()

    data21 = Promotions.objects.all()
    data23 = Awards.objects.all()
    data24 = Travel.objects.all()
    data25 = Transfer.objects.all()
    data26 = Resignation.objects.all()
    data27 = Complains.objects.all()
    data28 = Warnings.objects.all()
    data29 = Terminations.objects.all()

    data30 = Trainer.objects.all()
    data31 = Training_Type.objects.all()
    data32 = Training_List.objects.all()

    data35=Category.objects.all()
    data36=Assets.objects.all()


    data37 = Job_Post.objects.all()
    data38 = Job_Candidate.objects.all()
    data39 = Job_Interview.objects.all()
    data40 = Client.objects.all()
    data41 = Events.objects.all()

    data42 = Meetings.objects.all()
    data44 = Company_Info.objects.all()

    data48 = My_Tasks.objects.all()

    data49 = Source.objects.all()
    data51 = Priority.objects.all()
    data50 = Status.objects.all()

    data45 = Holidays.objects.all()

    data60 = Tickets.objects.all()
    data61 = Expenses.objects.all()
    data62 = Financial_year.objects.all()

    k=Pro.objects.all()
    context = {'data62':data62,'data61':data61,'data60':data60,'data49':data49,'data50':data50,'data51':data51,'data00':data00,'data_id':data_id,'emp_obj':emp_obj,'k':k,'data44':data44,'data45':data45,'data1':data1,'data2':data2,'data3':data3,'data4':data4,'data5':data5,'data6':data6,'data7':data7,'data8':data8,'data9':data9,'data10':data10,'data11':data11,'data12':data12,'data13':data13,'data14':data14,'data15':data15,'data16':data16,'data17':data17,'data18':data18,'data19':data19,'data20':data20,'data21':data21,'data23':data23,'data24':data24,'data25':data25,'data26':data26,'data27':data27,'data28':data28,'data29':data29,'data30':data30,'data31':data31,'data32':data32,'data35':data35,'data36':data36,'data37':data37,'data38':data38,'data39':data39,'data40':data40,'data41':data41,'data42':data42,'data48':data48}
    return render(request,'hr/emp.edit.html',context)



def emp_update_view(request, id):
    if request.method == 'POST':
        user = CustomUser.objects.get(id=id)
        emp_u=EmployeeMaster.objects.get(user_id=user.id)

        email = request.POST.get('email')
        username = request.POST.get('username')
        emp_id = request.POST.get('emp_id')
        phone_no = request.POST.get('phone_no')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state= request.POST.get('state')
        zip_code = request.POST.get('zip_code')
        country = request.POST.get('country')
        date_of_birth = request.POST.get('date_of_birth')
        gender = request.POST.get('gender')
        marital_status = request.POST.get('marital_status')
        #company = request.POST.get('company')
        department = request.POST.get('department')
        role = request.POST.get('role')
        shift = request.POST.get('shift')
        print(shift)
        joining_date= request.POST.get('joining_date')


        #print(first_name,phone_no,last_name, emp_id, role, department,joining_date,email,last_name ,address,city,state,zip_code,country, date_of_birth,  gender, marital_status,shift)
        # Update the item's fields based on the new data
        user.first_name=first_name
        user.last_name=last_name
        user.username=username
        user.email=email
        user.save()
        emp_u.EmployeeID =emp_id
        emp_u. phone_number=phone_no
        emp_u.Address =address
        emp_u.JoiningDate="-".join(joining_date.split("-"))
        emp_u.Shift_time=shift
        emp_u.City=city
        emp_u.state=state
        emp_u.PinCode= zip_code
        emp_u.country=country
        emp_u.BirthDate="-".join(date_of_birth.split("-"))
        emp_u.Gender=gender
        emp_u.MaratialStatus=marital_status

        emp_u.Department_id=department
        emp_u.Role=role
        emp_u.save()




        response_data = {'first_name': first_name}
        return JsonResponse(response_data)

    response_data = {'error': 'Invalid request'}
    return JsonResponse(response_data, status=400)



def Add_immegration_data(request):
    if request.method == 'POST' and request.FILES.get('immigration_document_file'):
        getdata=request.POST.get("data_id")


        user = CustomUser.objects.get(id=getdata)
        immigration_document_file= request.FILES['immigration_document_file']
        document_type = request.POST.get('document_type')
     #   immigration_document_file = request.POST.get('immigration_document_file')
        immigration_document_number = request.POST.get('immigration_document_number')
        immigration_issue_date = request.POST.get('immigration_issue_date')
        immigration_expiry_date = request.POST.get('immigration_expiry_date')
        country = request.POST.get('country')




        #print(document_type,immigration_document_file,immigration_document_number,immigration_issue_date,country,immigration_expiry_date,document_type)

        obj=Immegration.objects.create(user_id=user.id, document_name=document_type,document_number=immigration_document_number,issue_date="-".join( immigration_issue_date.split("-")),expired_date="-".join(immigration_expiry_date.split("-")),country=country,document_file= immigration_document_file)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)




def update_immegration_data(request):
    if request.method == 'POST' and request.FILES.get('immigration_document_file_update'):
        #data_id = request.POST.get('data_id')
        immigration_id = request.POST.get('immigration_id')

        countrys_update = request.POST.get('countrys_update')
        immigration_expiry_date_update = request.POST.get('immigration_expiry_date_update')
        immigration_issue_date_update = request.POST.get('immigration_issue_date_update')
        immigration_document_file_update = request.FILES['immigration_document_file_update']
        immigration_document_number_update = request.POST.get('immigration_document_number_update')
        document_type_update = request.POST.get('document_type_update')

        #print(immigration_id,countrys_update, immigration_expiry_date_update, immigration_issue_date_update,immigration_document_file_update,immigration_document_number_update,document_type_update)


        obj = Immegration.objects.get(id=immigration_id)


        obj.country = countrys_update
        obj.expired_date = immigration_expiry_date_update
        obj.issue_date = immigration_issue_date_update
        obj.document_file = immigration_document_file_update
        obj.document_number = immigration_document_number_update
        obj.document_name = document_type_update


        obj.save()

        response_data = {'message': "Training List record updated successfully."}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)






def delete_immegration_data(request):
    if request.method == 'POST':

        immigration_id = request.POST.get('immigration_id')
       # data_id = request.POST.get('data_id')
        #print(immigration_id)

        #user = CustomUser.objects.get(id=data_id)
        obj = Immegration.objects.get(id=immigration_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)



def get_immegration_data(request, id):
    if request.method == 'GET':
        obj = Immegration.objects.get(id=id)

        data = {
            'countrys_update': obj.country,
            'immigration_expiry_date_update': obj.expired_date,
            'immigration_issue_date_update': obj.issue_date,
            'immigration_document_number_update': obj.document_number,
            'document_type_update': obj.document_name,



        }
        #print(data)
        return JsonResponse(data)













def Add_emergency_contacts_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        relation = request.POST.get('relation')
        contact_work_email = request.POST.get('contact_work_email')
        contact_name = request.POST.get('contact_name')
        contact_address = request.POST.get('contact_address')
        mobile = request.POST.get('mobile')
        contact_city = request.POST.get('contact_city')
        contact_state = request.POST.get('contact_state')
        contact_zip = request.POST.get('contact_zip')
        country1 = request.POST.get('country1')



        #print(relation,contact_work_email,contact_name,contact_address,mobile,contact_city,contact_state,contact_zip,country1)

        obj=Emergency_Contacts.objects.create(user_id=user.id, relation=relation,email=contact_work_email,name=contact_name,address=contact_address, mobile_no=mobile,city=contact_city,state=contact_state,zip_code=contact_zip,country=country1)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)






def Add_profile_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        fb_id = request.POST.get('fb_id')
        skype_id = request.POST.get('skype_id')
        linkedIn_id = request.POST.get('linkedIn_id')
        twitter_id = request.POST.get('twitter_id')
        whatsapp_id = request.POST.get('whatsapp_id')




        #print(fb_id,skype_id,linkedIn_id,twitter_id,whatsapp_id)

        obj=Social_Profile.objects.create(user_id=user.id, facebook_profile=fb_id,linkedin_profile=linkedIn_id,skype_profile=skype_id,twitter_profile=twitter_id, whatsapp_profile=whatsapp_id)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)







def Add_documents_data(request):
    if request.method == 'POST' and request.FILES.get('document_document_file'):
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)
        document_document_file= request.FILES['document_document_file']
        document_title = request.POST.get('document_title')
        document_expiry_date = request.POST.get('document_expiry_date')
        document_end_date = request.POST.get('document_end_date')
        document_description = request.POST.get('document_description')
        document_is_notify = request.POST.get('document_is_notify')




        obj=Documents_All.objects.create(user_id=user.id,title=document_title,expired_date="-".join( document_expiry_date.split("-")),end_date="-".join( document_end_date.split("-")),description=document_description,document_file=document_document_file)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def Add_qualifications_data(request):
    if request.method == 'POST' and request.FILES.get('qualification_file'):
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        qualification_file= request.FILES['qualification_file']


        institution_name = request.POST.get('institution_name')
        education_level = request.POST.get('education_level')
        qualification_from_date = request.POST.get('qualification_from_date')
        qualification_to_date = request.POST.get('qualification_to_date')
        language = request.POST.get('language')
        professional_skills = request.POST.get('professional_skills')
        qualification_description = request.POST.get('qualification_description')





        #print(qualification_file,institution_name,education_level,qualification_from_date,qualification_to_date,language,professional_skills,qualification_description)

        obj=Qualifications.objects.create(user_id=user.id, name_of_university=institution_name,branch=education_level,from_date=qualification_from_date,to_date=qualification_to_date, languages=language,skills=professional_skills,qualification_file=qualification_file,description=qualification_description)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)






def Add_work_experience_data(request):
    if request.method == 'POST':
        getdata = request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        work_company_name = request.POST.get('work_company_name')
        work_experience_from_date = request.POST.get('work_experience_from_date')
        work_experience_to_date = request.POST.get('work_experience_to_date')
        designations = request.POST.get('designations')
        work_experience_description = request.POST.get('work_experience_description')






        #print(work_company_name,work_experience_from_date,work_experience_to_date,designations,work_experience_description)

        obj=Work_Experience.objects.create(user_id=user.id,company=work_company_name,from_date=work_experience_from_date,to_date=work_experience_to_date,description=work_experience_description,designation=designations)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)







def Add_bank_account_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        bank_account_number = request.POST.get('bank_account_number')
        bank_bank_name = request.POST.get('bank_bank_name')
        bank_bank_code = request.POST.get('bank_bank_code')
        bank_bank_branch = request.POST.get('bank_bank_branch')
        status = request.POST.get('status')






        #bank_account_title,bank_account_number,bank_bank_name,bank_bank_code,bank_bank_branch)

        obj=Bank_Account.objects.create(user_id=user.id,account_number=bank_account_number,bankname=bank_bank_name,ifsc_code=bank_bank_code,branch=bank_bank_branch,is_active=status)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)







def Add_basic_salary_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)


        basic_Salary_mon = request.POST.get('basic_Salary_mon')

        payslip_type = request.POST.get('payslip_type')
        basic_salary_amount = request.POST.get('basic_salary_amount')


        # Convert 'basic_salary_mon' from 'YYYY-MM' to 'Month, Year' format
        month_year_obj = datetime.datetime.strptime(basic_Salary_mon, '%Y-%m')
        formatted_month_year = month_year_obj.strftime('%B, %Y')

        #print(basic_Salary_mon,payslip_type,basic_salary_amount)

        obj=Basic_Salary.objects.create(user_id=user.id,month_year=formatted_month_year,payslip_type=payslip_type,basic_salary=basic_salary_amount)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def update_basic_salary_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        basic_salary_id=request.POST.get('basic_salary_id')

        basic_Salary_mon_upd = request.POST.get('basic_Salary_mon_upd')
        payslip_type_upd = request.POST.get('payslip_type_upd')
        basic_salary_amount_upd = request.POST.get('basic_salary_amount_upd')

        month_year_obj = datetime.datetime.strptime(basic_Salary_mon_upd, '%Y-%m')
        formatted_month_year = month_year_obj.strftime('%B, %Y')

        obj = Basic_Salary.objects.get(id=basic_salary_id)

        obj.month_year=formatted_month_year
        obj.payslip_type=payslip_type_upd
        obj.basic_salary=basic_salary_amount_upd




        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def get_basic_salary_data(request, id):
    if request.method == 'GET':
        obj = Basic_Salary.objects.get(id=id)

        data = {
            'basic_Salary_mon_upd': obj.month_year,
            'payslip_type_upd': obj.payslip_type,
            'basic_salary_amount_upd': obj.basic_salary,



        }
        #print(data)
        return JsonResponse(data)



def delete_basic_salary_data(request):
    if request.method == 'POST':

        basic_salary_id = request.POST.get('basic_salary_id')
       # data_id = request.POST.get('data_id')
       # print(basic_salary_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Basic_Salary.objects.get(id=basic_salary_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)









def Add_allowances_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        mon_year = request.POST.get('mon_year')
        allowance_type = request.POST.get('allowance_type')
        allowance_title = request.POST.get('allowance_title')
        allowance_amount = request.POST.get('allowance_amount')







       # print(mon_year,allowance_type,allowance_title,allowance_amount)

        obj=Allowances.objects.create(user_id=user.id,month_year=mon_year,allowance_type=allowance_type,allowance_title=allowance_title,allowance_amount=allowance_amount)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def update_allowances_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        allowances_id=request.POST.get('allowances_id')

        mon_year_upd = request.POST.get('mon_year_upd')
        allowance_type_upd = request.POST.get('allowance_type_upd')
        allowance_title_upd = request.POST.get('allowance_title_upd')
        allowance_amount_upd = request.POST.get('allowance_amount_upd')




        #print(allowances_id,data_id,mon_year_upd,allowance_type_upd,allowance_title_upd,allowance_amount_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Allowances.objects.get(user_id=data_id,id=allowances_id)




        obj.month_year=mon_year_upd
        obj.allowance_type=allowance_type_upd
        obj.allowance_title=allowance_title_upd
        obj.allowance_amount=allowance_amount_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)






def get_allowances_data(request, id):
    if request.method == 'GET':
        obj = Allowances.objects.get(id=id)

        data = {
            'mon_year_upd': obj.month_year,
            'allowance_type_upd': obj.allowance_type,
            'allowance_title': obj.allowance_title_upd,
            'allowance_amount': obj.allowance_amount_upd,

        }
        #print(data)
        return JsonResponse(data)



def delete_allowances_data(request):
    if request.method == 'POST':
        allowances_id = request.POST.get('allowances_id')
       # data_id = request.POST.get('data_id')
        #print(allowances_id)

        obj = Allowances.objects.get(id=allowances_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)













def Add_commissions_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")
        user = CustomUser.objects.get(id=getdata)

        mon_year_c = request.POST.get('mon_year_c')
        commission_title = request.POST.get('commission_title')
        commission_amount = request.POST.get('commission_amount')




        #print(mon_year_c,commission_title,commission_amount)

        obj=Commissions.objects.create(user_id=user.id,month_year=mon_year_c,commission_title=commission_title,commission_amount=commission_amount)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def update_commissions_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        commissions_id=request.POST.get('commissions_id')

        mon_year_c_upd = request.POST.get('mon_year_c_upd')
        commission_title_upd = request.POST.get('commission_title_upd')
        commission_amount_upd = request.POST.get('commission_amount_upd')




        #print(commissions_id,data_id,mon_year_c_upd,commission_title_upd,commission_amount_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Commissions.objects.get(user_id=data_id,id=commissions_id)




        obj.month_year=mon_year_c_upd
        obj.commission_title=commission_title_upd
        obj.commission_amount=commission_amount_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def get_commissions_data(request, id):
    if request.method == 'GET':
        obj = Commissions.objects.get(id=id)

        data = {
            'mon_year_c_upd': obj.month_year,
            'commission_title_upd': obj.commission_title,
            'commission_amount_upd': obj.commission_amount,

        }
        #print(data)
        return JsonResponse(data)




def delete_commissions_data(request):
    if request.method == 'POST':

        commissions_id = request.POST.get('commissions_id')
       # data_id = request.POST.get('data_id')
        #print(commissions_id)

        #user = CustomUser.objects.get(id=data_id)
        obj = Commissions.objects.get(id=commissions_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)








def Add_loan_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        mon_year_l = request.POST.get('mon_year_l')
        title_l = request.POST.get('title_l')
        number_of_installments = request.POST.get('number_of_installments')
        loan_option = request.POST.get('loan_option')
        amount_l = request.POST.get('amount_l')
        reason_l = request.POST.get('reason_l')




       # print(mon_year_l,title_l,number_of_installments,loan_option,amount_l,reason_l)

        obj=Loan.objects.create(user_id=user.id,month_year=mon_year_l,title=title_l,number_of_installments=number_of_installments,loan_option=loan_option,amount=amount_l,reason=reason_l)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def update_loan_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        loan_id=request.POST.get('loan_id')

        mon_year_l_upd = request.POST.get('mon_year_l_upd')
        title_l_upd = request.POST.get('title_l_upd')
        number_of_installments_upd = request.POST.get('number_of_installments_upd')
        loan_option_upd = request.POST.get('loan_option_upd')
        amount_l_upd = request.POST.get('amount_l_upd')
        reason_l_upd = request.POST.get('reason_l_upd')



       # print(loan_id,data_id,mon_year_l_upd,title_l_upd,number_of_installments_upd,loan_option_upd,amount_l_upd,reason_l_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Loan.objects.get(user_id=data_id,id=loan_id)




        obj.month_year=mon_year_l_upd
        obj.title=title_l_upd
        obj.number_of_installments=number_of_installments_upd
        obj.loan_option = loan_option_upd
        obj.amount = amount_l_upd
        obj.reason = reason_l_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def get_loan_data(request, id):
    if request.method == 'GET':
        obj = Loan.objects.get(id=id)

        data = {
            'mon_year_l_upd': obj.month_year,
            'title_l_upd': obj.title,
            'number_of_installments_upd': obj.number_of_installments,
            'loan_option_upd': obj.loan_option,
            'amount_l_upd': obj.amount,
            'reason_l_upd': obj.reason,

        }
      #  print(data)
        return JsonResponse(data)




def delete_loan_data(request):
    if request.method == 'POST':

        loan_id = request.POST.get('loan_id')
       # data_id = request.POST.get('data_id')
      #  print(loan_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Loan.objects.get(id=loan_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)








def Add_statutory_deduction_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        mon_year_s = request.POST.get('mon_year_s')
        title_s = request.POST.get('title_s')
        deduction_option = request.POST.get('deduction_option')
        deduction_amount = request.POST.get('deduction_amount')
        reason_s = request.POST.get('reason_s')





        #print(mon_year_s,title_s,deduction_option,deduction_amount,reason_s)

        obj=Statutory_Deduction.objects.create(user_id=user.id,month_year=mon_year_s,deduction_title=title_s,deduction_option=deduction_option,deduction_amount=deduction_amount,reason=reason_s)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)









def update_statutory_deduction_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        statutory_deduction_id=request.POST.get('statutory_deduction_id')

        mon_year_s_upd = request.POST.get('mon_year_s_upd')
        title_s_upd = request.POST.get('title_s_upd')
        deduction_option_upd = request.POST.get('deduction_option_upd')
        deduction_amount_upd = request.POST.get('deduction_amount_upd')
        reason_s_upd = request.POST.get('reason_s_upd')



      #  print(statutory_deduction_id,data_id,mon_year_s_upd,title_s_upd,deduction_option_upd,deduction_amount_upd,reason_s_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Statutory_Deduction.objects.get(user_id=data_id,id=statutory_deduction_id)




        obj.month_year=mon_year_s_upd
        obj.deduction_title=title_s_upd
        obj.deduction_option=deduction_option_upd
        obj.deduction_amount = deduction_amount_upd
        obj.reason = reason_s_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def get_statutory_deduction_data(request, id):
    if request.method == 'GET':
        obj = Statutory_Deduction.objects.get(id=id)

        data = {
            'mon_year_s_upd': obj.month_year,
            'title_s_upd': obj.deduction_title,
            'deduction_option_upd': obj.deduction_option,
            'deduction_amount_upd': obj.deduction_amount,
            'reason_s_upd': obj.reason,

        }
       # print(data)
        return JsonResponse(data)



def delete_statutory_deduction_data(request):
    if request.method == 'POST':

        statutory_deduction_id = request.POST.get('statutory_deduction_id')
       # data_id = request.POST.get('data_id')
       # print(statutory_deduction_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Statutory_Deduction.objects.get(id=statutory_deduction_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)









def Add_other_payments_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        mon_year_p = request.POST.get('mon_year_p')
        title_p = request.POST.get('title_p')
        amount_p = request.POST.get('amount_p')






       # print(mon_year_p,title_p,amount_p)

        obj=Other_Payments.objects.create(user_id=user.id,month_year=mon_year_p,title=title_p,amount=amount_p)

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def update_other_payments_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        other_payments_id=request.POST.get('other_payments_id')

        mon_year_p_upd = request.POST.get('mon_year_p_upd')
        title_p_upd = request.POST.get('title_p_upd')
        amount_p_upd = request.POST.get('amount_p_upd')



     #   print(other_payments_id,data_id,mon_year_p_upd,title_p_upd,amount_p_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Other_Payments.objects.get(user_id=data_id,id=other_payments_id)




        obj.month_year=mon_year_p_upd
        obj.title=title_p_upd
        obj.amount=amount_p_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)




def get_other_payments_data(request, id):
    if request.method == 'GET':
        obj = Other_Payments.objects.get(id=id)

        data = {
            'mon_year_p_upd': obj.month_year,
            'title_p_upd': obj.title,
            'amount_p_upd': obj.amount,


        }
       # print(data)
        return JsonResponse(data)



def delete_other_payments_data(request):
    if request.method == 'POST':

        other_payments_id = request.POST.get('other_payments_id')
       # data_id = request.POST.get('data_id')
       # print(other_payments_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Other_Payments.objects.get(id=other_payments_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)









def Add_over_time_data(request):
    if request.method == 'POST':
        getdata=request.POST.get("data_id")

        user = CustomUser.objects.get(id=getdata)

        mon_year_o = request.POST.get('mon_year_o')
        title_o = request.POST.get('title_o')
        number_of_days = request.POST.get('number_of_days')
        total_hours = request.POST.get('total_hours')
        rate = request.POST.get('rate')






       # print(mon_year_o,title_o,number_of_days,total_hours,rate)

        obj=Over_Time.objects.create(user_id=user.id,month_year=mon_year_o,title=title_o,number_of_days=number_of_days,total_hours=total_hours,rate=rate    )

        obj.save()
        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)









def update_over_time_data(request):
    if request.method == 'POST':
        data_id=request.POST.get('data_id')
        over_time_id=request.POST.get('over_time_id')

        mon_year_o_upd = request.POST.get('mon_year_o_upd')
        title_o_upd = request.POST.get('title_o_upd')
        number_of_days_upd = request.POST.get('number_of_days_upd')
        total_hours_upd = request.POST.get('total_hours_upd')
        rate_upd = request.POST.get('rate_upd')


       # print(over_time_id,data_id,mon_year_o_upd,title_o_upd,number_of_days_upd,total_hours_upd,rate_upd)








        user = CustomUser.objects.get(id=data_id)
        obj = Over_Time.objects.get(user_id=data_id,id=over_time_id)






        obj.month_year=mon_year_o_upd
        obj.title=title_o_upd
        obj.number_of_days=number_of_days_upd
        obj.total_hours=total_hours_upd
        obj.rate=rate_upd


        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def get_over_time_data(request, id):
    if request.method == 'GET':
        obj = Over_Time.objects.get(id=id)

        data = {
            'mon_year_o_upd': obj.month_year,
            'title_o_upd': obj.title,
            'number_of_days_upd': obj.number_of_days,
            'total_hours_upd': obj.total_hours,
            'rate_upd': obj.rate,

        }
      #  print(data)
        return JsonResponse(data)


def delete_over_time_data(request):
    if request.method == 'POST':

        over_time_id = request.POST.get('over_time_id')
       # data_id = request.POST.get('data_id')
     #   print(over_time_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Over_Time.objects.get(id=over_time_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)



def List_policy_data(request):
    data15 = Company_Policy.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Company_Policy/list_policy.html', {'data15':data15,'data22':data22})



def Adding_policy_data(request):
    if request.method == 'POST':

        title_company_policy = request.POST.get('title')
        description_company_policy = request.POST.get('description')
        company_company_policy = request.POST.get('company')

        obj = Company_Policy.objects.create(
            name_id=company_company_policy,
            title=title_company_policy,
            description=description_company_policy
        )



        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_policy')
    else:
        data22 = Company.objects.all()

        return render(request, 'corehr/Company_Policy/add_policy.html',{'data22':data22} )





def update_policy_data(request,id):
    if request.method == 'POST':

        title_company_policy_upd = request.POST.get('title')
        description_company_policy_upd = request.POST.get('description')
        company_company_policy_upd = request.POST.get('company')


        obj = Company_Policy.objects.get(id=id)
        obj.title = title_company_policy_upd
        obj.description = description_company_policy_upd
        obj.name_id = company_company_policy_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_policy')
    else:

        policy = Company_Policy.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'corehr/Company_Policy/update_policy.html',{"policy": policy,'data22':data22})




def delete_policy_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Company_Policy.objects.get(id=id)
        except Company_Policy.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_policy')
    else:
        return render(request, 'corehr/Company_Policy/delete_policy.html', )



def List_announcements_data(request):
    today = timezone.now().date()

    data16 = Announcements.objects.filter(start_date__gte=today)
    data22 = Company.objects.all()
    return render(request, 'corehr/Announcements/list_announcements.html', {'data16':data16,'data22':data22})



def Adding_announcements_data(request):
    if request.method == 'POST':

        title_announcements = request.POST.get('title')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        description_announcements = request.POST.get('description')
        company_announcements = request.POST.get('company')

        obj = Announcements.objects.create(

            title=title_announcements,
            start_date=start_date,
            end_date=end_date,
            description=description_announcements,
            name_id=company_announcements,
        )



        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_announcements')
    else:
        data22 = Company.objects.all()

        return render(request, 'corehr/Announcements/add_announcements.html',{'data22':data22} )





def update_announcements_data(request,id):
    if request.method == 'POST':

        title_announcements_upd = request.POST.get('title')
        start_date_upd = request.POST.get('start_date')
        end_date_upd = request.POST.get('end_date')
        description_announcements_upd = request.POST.get('description')
        company_announcements_upd = request.POST.get('company')


        obj = Announcements.objects.get(id=id)
        obj.title = title_announcements_upd
        obj.start_date = start_date_upd
        obj.end_date = end_date_upd
        obj.description = description_announcements_upd
        obj.name_id = company_announcements_upd
        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_announcements')
    else:

        announcements = Announcements.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'corehr/Announcements/update_announcements.html',{"announcements": announcements,'data22':data22})




def delete_announcements_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Announcements.objects.get(id=id)
        except Announcements.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_announcements')
    else:
        return render(request, 'corehr/Announcements/delete_announcements.html', )










def List_designation_data(request):
    data17 = Designation.objects.all()
    data19 = Department.objects.all()
    data22 = Company.objects.all()
    data1 = CustomUser.objects.all()

    return render(request, 'corehr/Designation/list_designation.html', {'data17':data17,'data22':data22,'data19':data19,'data1':data1})



def Adding_designation_data(request):
    if request.method == 'POST':

        designation_designation = request.POST.get('designation')
        company_designation = request.POST.get('company')
        department_dep = request.POST.get('department')
        employee = request.POST.get('employee')

        obj = Designation.objects.create(

            designation=designation_designation,
            name_id=company_designation,
            department_id=department_dep,
            employee_id=employee
        )




        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_designation')
    else:
        data22 = Company.objects.all()
        data19 = Department.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Designation/add_designation.html',{'data22':data22,'data19':data19,'data1':data1} )





def update_designation_data(request,id):
    if request.method == 'POST':
        designation_designation_upd = request.POST.get('designation')
        company_designation_upd = request.POST.get('company')
        department_dep_upd = request.POST.get('department')
        employee = request.POST.get('employee')


        obj = Designation.objects.get(id=id)
        obj.designation = designation_designation_upd
        obj.name_id = company_designation_upd
        obj.department_id = department_dep_upd
        obj.employee_id = employee

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_designation')
    else:

        designation = Designation.objects.get(id=id)
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request,'corehr/Designation/update_designation.html',{"designation": designation,'data22':data22,'data19':data19,'data1':data1})




def delete_designation_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Designation.objects.get(id=id)
        except Designation.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_designation')
    else:
        return render(request, 'corehr/Designation/delete_designation.html', )












def list_location_data(request):
    if request.method == 'GET':
       location_data = Location.objects.all()
       users_data = CustomUser.objects.all()
       return render(request,'hr/corehr/location/list_location.html',{'location_data':location_data,'users_data':users_data})












def Add_location_data(request):
    if request.method == 'POST':
        location = request.POST.get('location')
        location_head = request.POST.get('location_head')
        address_1 = request.POST.get('address_1')
        city = request.POST.get('city')
        state= request.POST.get('state')
        country = request.POST.get('country')
        zip = request.POST.get('zip')





        obj = Location.objects.create(location_head_id=location_head,location=location,address_line_1=address_1,city=city,state=state,country=country,zip=zip)

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_location')


    else:
        users_data = CustomUser.objects.all()

        return render(request, 'hr/corehr/location/add_location.html', {'users_data': users_data,})







def update_location_data(request,id):
    if request.method == 'POST':

        location_location_upd = request.POST.get('location_upd')
        location_head_location_upd = request.POST.get('location_head_upd')
        address_1_location_upd = request.POST.get('address_1_upd')
        city_location_upd = request.POST.get('city_upd')
        state_location_upd = request.POST.get('state_upd')
        country_location_upd = request.POST.get('country_upd')
        zip_location_upd = request.POST.get('zip_upd')



        obj = Location.objects.get(id=id)

        obj.location = location_location_upd
        obj.address_line_1 = address_1_location_upd
        obj.city = city_location_upd
        obj.state = state_location_upd
        obj.country = country_location_upd
        obj.zip = zip_location_upd

        obj.location_head_id = location_head_location_upd

        obj.save()

        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_location')
    else:

        location_data = Location.objects.get(id=id)
        users_data = CustomUser.objects.all()
        return render(request, 'hr/corehr/location/update_location.html', {'location_data': location_data,'users_data':users_data})




def delete_location_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Location.objects.get(id=id)
        except Location.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.success(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_location')
    else:
        return render(request, 'hr/corehr/location/delete_location.html')








def List_department_data(request):
    data19 = Department.objects.all()
    data22 = Company.objects.all()

    return render(request, 'hr/Department/list_department.html', {'data19':data19,'data22':data22})



def Adding_department_data(request):
    if request.method == 'POST':

        department_department = request.POST.get('department')
        company_department = request.POST.get('company')


       # print(department_department,company_department)


        obj = Department.objects.create(
                    name_id=company_department,
                    department=department_department,

            )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_department')
    else:
        data22 = Company.objects.all()

        return render(request, 'hr/Department/add_department.html',{'data22':data22} )




def update_department_data(request,id):
    if request.method == 'POST':

        department_department_upd = request.POST.get('department')
        company_department_upd = request.POST.get('company')



        obj = Department.objects.get(id=id)


        obj.department = department_department_upd
        obj.name_id = company_department_upd
        # obj.department_head = department_head_department_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_department')
    else:

        department_data = Department.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'hr/Department/update_department.html',{"department_data": department_data,'data22':data22})




def delete_department_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Department.objects.get(id=id)
        except Department.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.success(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_department')
    else:
        return render(request, 'hr/Department/delete_department.html', )






def List_company_data(request):
    if request.method == 'GET':
        company_data = Company.objects.all()
        return render(request,'hr/corehr/company/list_company.html',{'company_data':company_data})






def Add_company_data(request):

    if request.method == 'POST':


        company = request.POST.get('company')
        trading_name = request.POST.get('trading_name')
        registration_number = request.POST.get('registration_number')
        phone = request.POST.get('phone')
        email = request.POST.get('email')

       # print(company,trading_name,registration_number,phone,email)

        obj=Company.objects.create(name=company,trading_name=trading_name,registration_number=registration_number,phone=phone,email=email)

        obj.save()
        messages.success(request, "Data saved Successfully",
                         extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_company')

    else:
        return render(request,'hr/corehr/company/add_company.html')







def update_company_data(request,id):
    if request.method == 'POST':

        company_upd=request.POST.get('company_upd')


        trading_name_upd = request.POST.get('trading_name_upd')
        registration_number_upd = request.POST.get('registration_number_upd')
        phone_upd = request.POST.get('phone_upd')
        email_upd = request.POST.get('email_upd')

       # print(company_upd,trading_name_upd,registration_number_upd,phone_upd,email_upd)

        obj = Company.objects.get(id=id)

        obj.name=company_upd
        obj.trading_name=trading_name_upd
        obj.registration_number=registration_number_upd
        obj.phone = phone_upd
        obj.email = email_upd
        obj.save()
        messages.success(request, "Data updated Successfully",
                         extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_company')
    else:
        company_data = Company.objects.get(id=id)
        return render(request, 'hr/corehr/company/update_company.html',{'company_data':company_data})



def delete_company_data(request,id):
    if request.method == 'POST':


        obj = Company.objects.get(id=id)

        if obj:

            obj.delete()

            messages.success(request, "Data Deleted Successfully",
                             extra_tags='alert alert-success alert-dismissible fade show')
            return redirect('list_company')
        else:
            messages.error(request, "object not found",
                             extra_tags='alert alert-error alert-dismissible fade show')

    else:
        return render(request, 'hr/corehr/company/delete_company.html', )

def company_info(request):

    if request.method == 'POST' and request.FILES.get('logo_add'):


        company_name = request.POST.get('company_name')
        email = request.POST.get('email')
        country1 = request.POST.get('country1')
        state = request.POST.get('state')
        city = request.POST.get('city')
        mobile = request.POST.get('mobile')
        toll_free_no = request.POST.get('toll_free_no')
        website = request.POST.get('website')
        address = request.POST.get('address')
        logo = request.FILES['logo_add']



       # print(company_company,trading_name,registration_number,phone,email_company)

        obj=Company_Info.objects.create(company_info=company_name,email=email,country=country1,state=state,city=city,mobile=mobile,toll_free_no=toll_free_no,website=website,address=address,logo=logo)

        obj.save()

        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)

    else:
        data44 = Company_Info.objects.all()
        return render(request, 'hr/company_info.html', {'data44': data44})








def update_company_info_update_data(request):
    if request.method == 'POST' and request.FILES.get('logo'):
        data_id=request.POST.get('data_id')
        company_info_id=request.POST.get('company_info_id')

        company_name_upd = request.POST.get('company_name_upd')
        email_upd = request.POST.get('email_upd')
        country1_upd = request.POST.get('country1_upd')
        state_upd = request.POST.get('state_upd')
        city_upd = request.POST.get('city_upd')
        mobile_upd = request.POST.get('mobile_upd')
        toll_free_no_upd = request.POST.get('toll_free_no_upd')
        website_upd = request.POST.get('website_upd')
        address_upd = request.POST.get('address_upd')
        logo_upd = request.FILES['logo']



       # print(company_info_id,data_id,company_name_upd,email_upd,country1_upd,state_upd,city_upd,mobile_upd,toll_free_no_upd,website_upd,address_upd)



        obj = Company_Info.objects.get(id=company_info_id)




        obj.company_info=company_name_upd
        obj.email=email_upd
        obj.country=country1_upd
        obj.state = state_upd
        obj.city = city_upd
        obj.mobile = mobile_upd
        obj.toll_free_no = toll_free_no_upd
        obj.website = website_upd
        obj.address = address_upd
        obj.logo = logo_upd



        obj.save()


        response_data = {'first_name': "working fine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def get_company_info_data(request, id):
    if request.method == 'GET':
        obj = Company_Info.objects.get(id=id)

        data = {
            'company_name_upd': obj.company_info,
            'email_upd': obj.email,
            'country1_upd': obj.country,
            'state_upd': obj.state,
            'city_upd': obj.city,
            'mobile_upd': obj.mobile,
            'toll_free_no_upd': obj.toll_free_no,
            'website_upd': obj.website,
            'address_upd': obj.address,




        }
       # print(data)
        return JsonResponse(data)


def List_promotions_data(request):
    data1 = CustomUser.objects.all()
    data21 = Promotions.objects.all()
    data22 = Company.objects.all()

    return render(request, 'corehr/Promotions/list.html', {'data1':data1,'data21':data21,'data22':data22})



def Adding_promotions_data(request):
    if request.method == 'POST':

        company_promotions = request.POST.get('company')
        employee_promotions = request.POST.get('employee')
        title_promotions = request.POST.get('title')
        promotion_date = request.POST.get('promotion_date')
        description_promotions = request.POST.get('description')



        obj = Promotions.objects.create(title=title_promotions,
                                        promotion_date=promotion_date,
                                        description=description_promotions,
                                        company_id=company_promotions,
                                        employee_id=employee_promotions
                                        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_promotions')
    else:
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Promotions/add.html',{'data22':data22,'data1':data1} )




def update_promotions_data(request,id):
    if request.method == 'POST':

        company_promotions_upd = request.POST.get('company')
        employee_promotions_upd = request.POST.get('employee')
        title_promotions_upd = request.POST.get('title')
        promotion_date_upd = request.POST.get('promotion_date')
        description_promotions_upd = request.POST.get('description')



        obj = Promotions.objects.get(id=id)

        obj.title = title_promotions_upd
        obj.promotion_date = promotion_date_upd
        obj.description = description_promotions_upd
        obj.company_id = company_promotions_upd
        obj.employee_id = employee_promotions_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_promotions')
    else:

        promotions_data = Promotions.objects.get(id=id)
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Promotions/update.html',{"promotions_data": promotions_data,'data22':data22,'data1':data1})




def delete_promotions_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Promotions.objects.get(id=id)
        except Promotions.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_promotions')
    else:
        return render(request, 'corehr/Promotions/delete.html', )






def List_awards_data(request):
    data1 = CustomUser.objects.all()
    data23 = Awards.objects.all()
    data22 = Company.objects.all()
    data19 = Department.objects.all()


    return render(request, 'corehr/Awards/list_awards.html', {'data1':data1,'data23':data23,'data22':data22,'data19':data19})



def Adding_awards_data(request):
    if request.method == 'POST':

        company_awards = request.POST.get('company')
        department_awards = request.POST.get('department')
        employee_awards = request.POST.get('employee')
        award_type_awards = request.POST.get('award_type')
        gift_awards = request.POST.get('gift')
        award_date = request.POST.get('award_date')

        obj = Awards.objects.create(
            award_type=award_type_awards,
            gift=gift_awards,
            award_date=award_date,
            company_id=company_awards,
            department_id=department_awards,
            employee_id=employee_awards
        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_awards')
    else:
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        data19 = Department.objects.all()

        return render(request, 'corehr/Awards/add_awards.html',{'data22':data22,'data1':data1,'data19':data19} )





def update_awards_data(request,id):
    if request.method == 'POST':

        company_awards_upd = request.POST.get('company')
        department_awards_upd = request.POST.get('department')
        employee_awards_upd = request.POST.get('employee')
        award_type_awards_upd = request.POST.get('award_type')
        gift_awards_upd = request.POST.get('gift')
        award_date_upd = request.POST.get('award_date')



        obj = Awards.objects.get(id=id)

        obj.award_type = award_type_awards_upd
        obj.gift = gift_awards_upd
        obj.award_date = award_date_upd
        obj.department_id = department_awards_upd
        obj.company_id = company_awards_upd
        obj.employee_id = employee_awards_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_awards')
    else:

        awards_data = Awards.objects.get(id=id)
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Awards/update_awards.html',{"awards_data": awards_data,'data22':data22,'data1':data1,'data19':data19})




def delete_awards_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Awards.objects.get(id=id)
        except Awards.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_awards')
    else:
        return render(request, 'corehr/Awards/delete_awards.html', )





def List_travel_data(request):
    data1 = CustomUser.objects.all()
    data24 = Travel.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Travel/list_travel.html', {'data1':data1,'data24':data24,'data22':data22})



def Adding_travel_data(request):
    if request.method == 'POST':

        company_travel = request.POST.get('company')
        employee_travel = request.POST.get('employee')
        arrangement_type = request.POST.get('arrangement_type')
        purpose_of_visit = request.POST.get('purpose_of_visit')
        place_of_visit = request.POST.get('place_of_visit')
        description_travel = request.POST.get('description')
        start_date_travel = request.POST.get('start_date')
        end_date_travel = request.POST.get('end_date')
        expected_budget = request.POST.get('expected_budget')
        actual_budget = request.POST.get('actual_budget')
        travel_mode = request.POST.get('travel_mode')



        obj = Travel.objects.create(
                arrangement_type=arrangement_type,
                purpose_of_visit=purpose_of_visit,
                place_of_visit=place_of_visit,
                description=description_travel,
                start_date=start_date_travel,
                end_date=end_date_travel,
                expected_budget=expected_budget,
                actual_budget=actual_budget,
                travel_mode=travel_mode,
                company_id=company_travel,
                employee_id=employee_travel
            )
        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_travel')
    else:
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Travel/add_travel.html',{'data22':data22,'data1':data1} )





def update_travel_data(request,id):
    if request.method == 'POST':

        company_travel_upd = request.POST.get('company')
        employee_travel_upd = request.POST.get('employee')
        arrangement_type_upd = request.POST.get('arrangement_type')
        purpose_of_visit_upd = request.POST.get('purpose_of_visit')
        place_of_visit_upd = request.POST.get('place_of_visit')
        description_travel_upd = request.POST.get('description')
        start_date_travel_upd = request.POST.get('start_date')
        end_date_travel_upd = request.POST.get('end_date')
        expected_budget_upd = request.POST.get('expected_budget')
        actual_budget_upd = request.POST.get('actual_budget')
        travel_mode_upd = request.POST.get('travel_mode')



        obj = Travel.objects.get(id=id)

        obj.arrangement_type = arrangement_type_upd
        obj.purpose_of_visit = purpose_of_visit_upd
        obj.place_of_visit = place_of_visit_upd
        obj.description = description_travel_upd
        obj.start_date = start_date_travel_upd
        obj.end_date = end_date_travel_upd
        obj.expected_budget = expected_budget_upd
        obj.actual_budget = actual_budget_upd
        obj.travel_mode = travel_mode_upd
        obj.company_id = company_travel_upd
        obj.employee_id = employee_travel_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_travel')
    else:

        travel_data = Travel.objects.get(id=id)
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Travel/update_travel.html',{"travel_data": travel_data,'data22':data22,'data1':data1})




def delete_travel_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Travel.objects.get(id=id)
        except Travel.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_travel')
    else:
        return render(request, 'corehr/Travel/delete_travel.html', )








def List_transfer_data(request):
    data1 = CustomUser.objects.all()
    data25 = Transfer.objects.all()
    data19 = Department.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Transfer/list_transfer.html', {'data1':data1,'data25':data25,'data19':data19,'data22':data22})



def Adding_transfer_data(request):
    if request.method == 'POST':

        company_transfer = request.POST.get('company')
        employee_transfer = request.POST.get('employee')
        to_department = request.POST.get('to_department')
        transfer_date = request.POST.get('transfer_date')
        description_transfer = request.POST.get('description')

        obj = Transfer.objects.create(
            department_id=to_department,
            transfer_date=transfer_date,
            description=description_transfer,
            company_id=company_transfer,
            employee_id=employee_transfer
        )
        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_transfer')
    else:
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Transfer/add_transfer.html',{'data19':data19,'data22':data22,'data1':data1} )





def update_transfer_data(request,id):
    if request.method == 'POST':

        company_transfer_upd = request.POST.get('company')
        employee_transfer_upd = request.POST.get('employee')
        to_department_upd = request.POST.get('to_department')
        transfer_date_upd = request.POST.get('transfer_date')
        description_transfer_upd = request.POST.get('description')

        obj = Transfer.objects.get(id=id)

        obj.department_id = to_department_upd
        obj.transfer_date = transfer_date_upd
        obj.description = description_transfer_upd
        obj.company_id = company_transfer_upd
        obj.employee_id = employee_transfer_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_transfer')
    else:

        transfer_data = Transfer.objects.get(id=id)
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Transfer/update_transfer.html',{"transfer_data": transfer_data,'data19':data19,'data22':data22,'data1':data1})




def delete_transfer_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Transfer.objects.get(id=id)
        except Transfer.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_transfer')
    else:
        return render(request, 'corehr/Transfer/delete_transfer.html', )









def List_resignation_data(request):
    data1 = CustomUser.objects.all()
    data26 = Resignation.objects.all()
    data19 = Department.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Resignation/list_resignation.html', {'data1':data1,'data26':data26,'data19':data19,'data22':data22})



def Adding_resignation_data(request):
    if request.method == 'POST':

        company_resignation = request.POST.get('company')
        department_resignation = request.POST.get('department')
        employee_resignation = request.POST.get('employee')
        notice_date = request.POST.get('notice_date')
        resignation_date = request.POST.get('resignation_date')
        description_resignation = request.POST.get('description')

        obj = Resignation.objects.create(
            notice_date=notice_date,
            resignation_date=resignation_date,
            description=description_resignation,
            company_id=company_resignation,
            department_id=department_resignation,
            employee_id=employee_resignation
        )
        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_resignation')
    else:
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Resignation/add_resignation.html',{'data19':data19,'data22':data22,'data1':data1} )





def update_resignation_data(request,id):
    if request.method == 'POST':

        company_resignation_upd = request.POST.get('company')
        department_resignation_upd = request.POST.get('department')
        employee_resignation_upd = request.POST.get('employee')
        notice_date_upd = request.POST.get('notice_date')
        resignation_date_upd = request.POST.get('resignation_date')
        description_resignation_upd = request.POST.get('description')



        obj = Resignation.objects.get(id=id)

        obj.notice_date = notice_date_upd
        obj.resignation_date = resignation_date_upd
        obj.description = description_resignation_upd
        obj.company_id = company_resignation_upd
        obj.department_id = department_resignation_upd
        obj.employee_id = employee_resignation_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_resignation')
    else:

        resignation_data = Resignation.objects.get(id=id)
        data19 = Department.objects.all()
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Resignation/update_resignation.html',{"resignation_data": resignation_data,'data19':data19,'data22':data22,'data1':data1})




def delete_resignation_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Resignation.objects.get(id=id)
        except Resignation.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_resignation')
    else:
        return render(request, 'corehr/Resignation/delete_resignation.html', )








def List_complains_data(request):
    data1 = CustomUser.objects.all()
    data27 = Complains.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Complains/list_complians.html', {'data1':data1,'data27':data27,'data22':data22})



def Adding_complains_data(request):
    if request.method == 'POST':

        company_complains = request.POST.get('company')
        employee_complains = request.POST.get('employee')
        complain_title = request.POST.get('complain_title')
        description_complain = request.POST.get('description')
        complain_date = request.POST.get('complain_date')

        obj = Complains.objects.create(
            complain_title=complain_title,
            description=description_complain,
            complain_date=complain_date,
            company_id=company_complains,
            employee_id=employee_complains
        )
        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_complains')
    else:
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()

        return render(request, 'corehr/Complains/add_complians.html',{'data22':data22,'data1':data1} )





def update_complains_data(request,id):
    if request.method == 'POST':

        company_complains_upd = request.POST.get('company')
        employee_complains_upd = request.POST.get('employee')
        complain_title_upd = request.POST.get('complain_title')
        description_complain_upd = request.POST.get('description')
        complain_date_upd = request.POST.get('complain_date')




        obj = Complains.objects.get(id=id)

        obj.complain_title = complain_title_upd
        obj.description = description_complain_upd
        obj.complain_date = complain_date_upd
        obj.company_id = company_complains_upd
        obj.employee_id = employee_complains_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_complains')
    else:

        complains_data = Complains.objects.get(id=id)
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        return render(request,'corehr/Complains/update_complains.html',{"complains_data": complains_data,'data22':data22,'data1':data1})




def delete_complains_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Complains.objects.get(id=id)
        except Complains.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_complains')
    else:
        return render(request, 'corehr/Complains/delete_complians.html', )










def List_warnings_data(request):
    data1 = CustomUser.objects.all()
    data19 = Department.objects.all()
    data28 = Warnings.objects.all()
    data22 = Company.objects.all()
    return render(request, 'corehr/Warnings/list_warnings.html', {'data1':data1,'data28':data28,'data22':data22,'data19':data19})



def Adding_warnings_data(request):
    if request.method == 'POST':

        title = request.POST.get('title')
        company_warnings = request.POST.get('company')
        employee_warnings = request.POST.get('employee')
        department = request.POST.get('department')
        description_warnings = request.POST.get('description')
        warning_date = request.POST.get('warning_date')

        obj = Warnings.objects.create(
            title=title,
            description=description_warnings,
            warning_date=warning_date,
            company_id=company_warnings,
            employee_id=employee_warnings,
            department_id=department
        )




        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_warnings')
    else:
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        data19 = Department.objects.all()

        return render(request, 'corehr/Warnings/add_warnings.html',{'data22':data22,'data1':data1,'data19':data19} )





def update_warnings_data(request,id):
    if request.method == 'POST':
        title = request.POST.get('title')
        company_warnings_upd = request.POST.get('company')
        employee_warnings_upd = request.POST.get('employee')
        department = request.POST.get('department')
        description_warnings_upd = request.POST.get('description')
        warning_date_upd = request.POST.get('warning_date')








        obj = Warnings.objects.get(id=id)
        title = title
        obj.description = description_warnings_upd
        obj.warning_date = warning_date_upd
        obj.company_id = company_warnings_upd
        obj.employee_id = employee_warnings_upd
        obj.department_id = department

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_warnings')
    else:

        warnings_data = Warnings.objects.get(id=id)
        data22 = Company.objects.all()
        data1 = CustomUser.objects.all()
        data19 = Department.objects.all()
        return render(request,'corehr/Warnings/update_warnings.html',{"warnings_data": warnings_data,'data22':data22,'data1':data1,'data19':data19})




def delete_warnings_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Warnings.objects.get(id=id)
        except Warnings.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_warnings')
    else:
        return render(request, 'corehr/Warnings/delete_warnings.html', )






def Add_terminations_data(request):
    if request.method == 'POST':



        company_terminations = request.POST.get('company_terminations')
        employee_terminations = request.POST.get('employee_terminations')
        termination_type = request.POST.get('termination_type')
        description_terminations = request.POST.get('description_terminations')
        termination_date = request.POST.get('termination_date')
        notice_date = request.POST.get('notice_date')




        #print("Company:", company_terminations)
        #print("Employee:", employee_terminations)
        #print("Termination Type:", termination_type)
        #print("Description:", description_terminations)
        #print("Termination Date:", termination_date)
        #print("Notice Date:", notice_date)


        obj = Terminations.objects.create(
                termination_type=termination_type,
                description=description_terminations,
                termination_date=termination_date,
                notice_date=notice_date,
                company_id=company_terminations,
                employee_id=employee_terminations
            )
        response_data = {'message': "Termination created successfully"}


        return JsonResponse(response_data)
    else:
        data1 = CustomUser.objects.all()
        data29 = Terminations.objects.all()
        data22 = Company.objects.all()
        return render(request, 'hr/corehr/add_terminations.html', {'data29': data29,'data1':data1,'data22':data22})
    return JsonResponse({'error': 'Invalid form data.'}, status=400)


def update_terminations_data(request):
    if request.method == 'POST':
        #data_id = request.POST.get('data_id')
        terminations_id = request.POST.get('terminations_id')

        company_terminations_upd = request.POST.get('company_terminations_upd')
        employee_terminations_upd = request.POST.get('employee_terminations_upd')
        termination_type_upd = request.POST.get('termination_type_upd')
        description_terminations_upd = request.POST.get('description_terminations_upd')
        termination_date_upd = request.POST.get('termination_date_upd')
        notice_date_upd = request.POST.get('notice_date_upd')

       # print(terminations_id,company_terminations_upd, employee_terminations_upd, termination_type_upd,description_terminations_upd, termination_date_upd, notice_date_upd)


        obj = Terminations.objects.get(id=terminations_id)

        obj.termination_type = termination_type_upd
        obj.description = description_terminations_upd
        obj.termination_date = termination_date_upd
        obj.notice_date = notice_date_upd
        obj.company_id = company_terminations_upd
        obj.employee_id = employee_terminations_upd

        obj.save()

        response_data = {'message': "Termination record updated successfully."}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def get_terminations_data(request, id):
    if request.method == 'GET':
        obj = Terminations.objects.get(id=id)

        data = {
            'termination_type_upd': obj.termination_type,
            'description_terminations_upd': obj.description,
            'termination_date_upd': obj.termination_date,
            'notice_date_upd': obj.notice_date,
            'company_terminations_upd': obj.company_id,
            'employee_terminations_upd': obj.employee_id



        }
        return JsonResponse(data)





def delete_terminations_data(request):
    if request.method == 'POST':

        terminations_id = request.POST.get('terminations_id')
       # data_id = request.POST.get('data_id')
       # print(terminations_id)

     #   user = CustomUser.objects.get(id=data_id)
        obj = Terminations.objects.get(id=terminations_id)
        if obj:
            obj.delete()
            response_data = {'first_name': "working fine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)






def List_trainer_data(request):
    data30 = Trainer.objects.all()
    data22 = Company.objects.all()


    return render(request, 'Training/Trainer/list_trainer.html', {'data30':data30,'data22':data22})



def Adding_trainer_data(request):
    if request.method == 'POST':

        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        company_trainer = request.POST.get('company')
        address = request.POST.get('address')

        obj = Trainer.objects.create(
            full_name=full_name,
            email=email,
            phone=phone,
            address=address,
            company_id=company_trainer,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_trainer')
    else:
        data22 = Company.objects.all()

        return render(request, 'Training/Trainer/add_trainer.html',{'data22':data22} )





def update_trainer_data(request,id):
    if request.method == 'POST':

        full_name_upd = request.POST.get('full_name')
        email_upd = request.POST.get('email')
        phone_upd = request.POST.get('phone')
        company_trainer_upd = request.POST.get('company')
        expertise_upd = request.POST.get('expertise')
        address_upd = request.POST.get('address')



        obj = Trainer.objects.get(id=id)

        obj.full_name = full_name_upd
        obj.email = email_upd
        obj.phone = phone_upd
        obj.expertise = expertise_upd
        obj.address = address_upd
        obj.company_id = company_trainer_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_trainer')
    else:

        trainer = Trainer.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'Training/Trainer/update_trainer.html',{"trainer": trainer,'data22':data22})




def delete_trainer_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Trainer.objects.get(id=id)
        except Trainer.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_trainer')
    else:
        return render(request, 'Training/Trainer/delete_trainer.html', )








def List_training_type_data(request):
    data31 = Training_Type.objects.all()

    return render(request, 'Training/Training_Type/list_training_type.html', {'data31':data31})



def Adding_training_type_data(request):
    if request.method == 'POST':

        training_type_trainer = request.POST.get('training_type')

        obj = Training_Type.objects.create(
            training_type=training_type_trainer,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_training_type')
    else:
        return render(request, 'Training/Training_Type/add_training_type.html')





def update_training_type_data(request,id):
    if request.method == 'POST':

        training_type_trainer_upd = request.POST.get('training_type')

        obj = Training_Type.objects.get(id=id)

        obj.training_type = training_type_trainer_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_training_type')
    else:

        training_type = Training_Type.objects.get(id=id)
        return render(request,'Training/Training_Type/update_training_type.html',{"training_type": training_type})




def delete_training_type_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Training_Type.objects.get(id=id)
        except Training_Type.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_training_type')
    else:
        return render(request, 'Training/Training_Type/delete_training_type.html', )








def List_training_list_data(request):
    data22 = Company.objects.all()
    data30 = Trainer.objects.all()
    data1 = CustomUser.objects.all()
    data32 = Training_List.objects.all()
    data31 = Training_Type.objects.all()

    return render(request, 'Training/Training_List/list_training_list.html', {'data22':data22,'data30':data30,'data1':data1,'data32':data32,'data31':data31})



def Adding_training_list_data(request):
    if request.method == 'POST':

        company_training_list = request.POST.get('company')
        training_type_list = request.POST.get('training_type')
        full_name_list = request.POST.get('full_name')
        employee_list = request.POST.get('employee')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        description = request.POST.get('description')

        obj = Training_List.objects.create(
            start_date=start_date,
            end_date=end_date,
            description=description,
            company_id=company_training_list,
            employee_id=employee_list,
            full_name_id=full_name_list,
            training_type_id=training_type_list,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_training_list')
    else:
        data22 = Company.objects.all()
        data30 = Trainer.objects.all()
        data1 = CustomUser.objects.all()
        data31 = Training_Type.objects.all()

        return render(request, 'Training/Training_List/add_training_list.html',{'data22':data22,'data30':data30,'data1':data1,'data31':data31})





def update_training_list_data(request,id):
    if request.method == 'POST':

        company_training_list_upd = request.POST.get('company')
        training_type_list_upd = request.POST.get('training_type')
        full_name_list_upd = request.POST.get('full_name')
        employee_list_upd = request.POST.get('employee')
        start_date_upd = request.POST.get('start_date')
        end_date_upd = request.POST.get('end_date')
        description_upd = request.POST.get('description')

        obj = Training_List.objects.get(id=id)

        obj.start_date = start_date_upd
        obj.end_date = end_date_upd
        obj.description = description_upd
        obj.company_id = company_training_list_upd
        obj.employee_id = employee_list_upd
        obj.full_name_id = full_name_list_upd
        obj.training_type_id = training_type_list_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_training_list')
    else:
        data22 = Company.objects.all()
        data30 = Trainer.objects.all()
        data1 = CustomUser.objects.all()
        data31 = Training_Type.objects.all()

        training_list = Training_List.objects.get(id=id)
        return render(request,'Training/Training_List/update_training_list.html',{"training_list": training_list,'data22':data22,'data30':data30,'data1':data1,'data31':data31})




def delete_training_list_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Training_List.objects.get(id=id)
        except Training_List.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_training_list')
    else:
        return render(request, 'Training/Training_List/delete_training_listr.html', )











def List_category_data(request):
    data35 = Category.objects.all()
    return render(request, 'Assets/Category/list_category.html', {'data35':data35})



def Adding_category_data(request):
    if request.method == 'POST':
        category_name = request.POST.get('category')

        obj = Category.objects.create(
            category=category_name,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_category')
    else:

        return render(request, 'Assets/Category/add_category.html')





def update_category_data(request,id):
    if request.method == 'POST':

        category_name_upd = request.POST.get('category')

        obj = Category.objects.get(id=id)

        obj.category = category_name_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_category')
    else:

        category = Category.objects.get(id=id)


        return render(request,'Assets/Category/update_category.html',{"category": category})




def delete_category_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Category.objects.get(id=id)
        except Category.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_category')
    else:
        return render(request, 'Assets/Category/delete_category.html', )












def List_assets_data(request):
    data1 = CustomUser.objects.all()
    data35 = Category.objects.all()
    data36 = Assets.objects.all()
    return render(request, 'Assets/assets/list_assets.html', {'data1':data1,'data35':data35,'data36':data36})



def Adding_assets_data(request):
    if request.method == 'POST':
        asset_id = request.POST.get('asset_id')
        category = request.POST.get('category')
        is_working = request.POST.get('is_working')
        employee_assets = request.POST.get('employee')
        purchase_date = request.POST.get('purchase_date')
        invoice_number = request.POST.get('invoice_number')
        status = request.POST.get('status')


        obj = Assets.objects.create(
            asset_id=asset_id,
            is_working=is_working,
            purchase_date=purchase_date,
            invoice_number=invoice_number,
            status=status,
            category_id=category,
            employee_id=employee_assets,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_assets')
    else:
        data1 = CustomUser.objects.all()
        data35 = Category.objects.all()

        return render(request, 'Assets/assets/add_assets.html',{'data1':data1,'data35':data35} )





def update_assets_data(request,id):
    if request.method == 'POST':

        asset_id = request.POST.get('asset_id')
        category = request.POST.get('category')
        employee_assets = request.POST.get('employee')
        purchase_date = request.POST.get('purchase_date')
        is_working = request.POST.get('is_working')
        invoice_number = request.POST.get('invoice_number')
        status = request.POST.get('status')


        obj = Assets.objects.get(id=id)


        obj.asset_id = asset_id
        obj.is_working = is_working
        obj.purchase_date = purchase_date
        obj.invoice_number = invoice_number
        obj.status = status
        obj.category_id = category
        obj.employee_id = employee_assets
        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_assets')
    else:

        assets = Assets.objects.get(id=id)
        data1 = CustomUser.objects.all()
        data35 = Category.objects.all()

        return render(request,'Assets/assets/update_assets.html',{"assets": assets,'data1':data1,'data35':data35})




def delete_assets_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Assets.objects.get(id=id)
        except Assets.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_assets')
    else:
        return render(request, 'Assets/assets/delete_assets.html', )








def List_job_post_data(request):
    data37 = Job_Post.objects.all()
    data22 = Company.objects.all()
    return render(request, 'recruitment/Job_Post/list_job_post.html', {'data37':data37,'data22':data22})



def Adding_job_post_data(request):
    if request.method == 'POST':
        company_job_post = request.POST.get('company')
        job_title = request.POST.get('job_title')
        job_type = request.POST.get('job_type')
        job_category = request.POST.get('job_category')
        no_of_vacancy = request.POST.get('no_of_vacancy')
        date_of_closing = request.POST.get('date_of_closing')
        experience = request.POST.get('experience')
        status = request.POST.get('status')

        obj = Job_Post.objects.create(
            job_title=job_title,
            job_type=job_type,
            job_category=job_category,
            no_of_vacancy=no_of_vacancy,
            date_of_closing=date_of_closing,
            minimum_experience=experience,
            status=status,
            company_id=company_job_post,

        )




        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_job_post')
    else:
        data22 = Company.objects.all()

        return render(request, 'recruitment/Job_Post/add_job_post.html',{'data22':data22} )





def update_job_post_data(request,id):
    if request.method == 'POST':
        company_job_post_upd = request.POST.get('company')
        job_title_upd = request.POST.get('job_title')
        job_type_upd = request.POST.get('job_type')
        job_category_upd = request.POST.get('job_category')
        no_of_vacancy_upd = request.POST.get('no_of_vacancy')
        date_of_closing_upd = request.POST.get('date_of_closing')
        experience_upd = request.POST.get('experience')
        status_upd = request.POST.get('status')


        obj = Job_Post.objects.get(id=id)
        obj.job_title = job_title_upd
        obj.job_type = job_type_upd
        obj.job_category = job_category_upd
        obj.no_of_vacancy = no_of_vacancy_upd
        obj.date_of_closing = date_of_closing_upd
        obj.minimum_experience = experience_upd
        obj.status = status_upd
        obj.company_id = company_job_post_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_job_post')
    else:
        job_post = Job_Post.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'recruitment/Job_Post/update_job_post.html',{"job_post": job_post,'data22':data22})




def delete_job_post_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Job_Post.objects.get(id=id)
        except Job_Post.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_job_post')
    else:
        return render(request, 'recruitment/Job_Post/delete_job_post.html', )






def List_job_candidate_data(request):
    data37 = Job_Post.objects.all()
    data38 = Job_Candidate.objects.all()
    return render(request, 'recruitment/Job_Candidate/list_job_candidate.html', {'data37':data37,'data38':data38})



def Adding_job_candidate_data(request):
    if request.method == 'POST':
        job_title = request.POST.get('job_title')
        candidate_name = request.POST.get('candidate_name')
        candidate_email = request.POST.get('candidate_email')
        resume = request.POST.get('resume')
        experience = request.POST.get('experience')
        applied_date = request.POST.get('applied_date')

        obj = Job_Candidate.objects.create(
            candidate_name=candidate_name,
            candidate_email=candidate_email,
            resume=resume,
            experience=experience,
            apply_date=applied_date,
            job_title_id=job_title,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_job_candidate')
    else:
        data37 = Job_Post.objects.all()

        return render(request, 'recruitment/Job_Candidate/add_job_candidate.html',{'data37':data37} )



def update_job_candidate_data(request,id):
    job_candidate = Job_Candidate.objects.get(id=id)

    if request.method == 'POST':
        job_title_upd = request.POST.get('job_title')
        candidate_name_upd = request.POST.get('candidate_name')
        candidate_email_upd = request.POST.get('candidate_email')
        resume_upd = request.FILES.get('resume')
        print(resume_upd)
        if resume_upd:
            job_candidate.resume=resume_upd
        experience = request.POST.get('experience')
        applied_date_upd = request.POST.get('applied_date')


        job_candidate.candidate_name = candidate_name_upd
        job_candidate.candidate_email = candidate_email_upd
        job_candidate.experience = experience
        job_candidate.apply_date = applied_date_upd
        job_candidate.job_title_id = job_title_upd

        job_candidate.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_job_candidate')
    else:

        data37 = Job_Post.objects.all()
        return render(request,'recruitment/Job_Candidate/update_job_candidate.html',{"job_candidate": job_candidate,'data37':data37})




def delete_job_candidate_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Job_Candidate.objects.get(id=id)
        except Job_Candidate.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_job_candidate')
    else:
        return render(request, 'recruitment/Job_Candidate/delete_job_candidate.html', )




def List_job_interview_data(request):
    today = timezone.now().date()

    data37 = Job_Post.objects.all()
    data38 = Job_Candidate.objects.all()
    data39 = Job_Interview.objects.filter(interview_date__gte=today).all()

    return render(request, 'recruitment/Job_Interview/list_job_interview.html', {'data37':data37,'data38':data38,'data39':data39})



def Adding_job_interview_data(request):
    if request.method == 'POST':
        job_title = request.POST.get('job_title')
        candidate_name = request.POST.get('candidate_name')
        interview_place = request.POST.get('interview_place')
        interview_date = request.POST.get('interview_date')
        interview_time = request.POST.get('interview_time')

        obj = Job_Interview.objects.create(
            interview_place=interview_place,
            interview_date=interview_date,
            interview_time=interview_time,
            candidate_name_id=candidate_name,
            job_title_id=job_title

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_job_interview')
    else:
        data37 = Job_Post.objects.all()
        data38 = Job_Candidate.objects.all()

        return render(request, 'recruitment/Job_Interview/add_job_interview.html',{'data37':data37,'data38':data38} )





def update_job_interview_data(request,id):
    if request.method == 'POST':
        job_title_upd = request.POST.get('job_title')
        candidate_name_upd = request.POST.get('candidate_name')
        interview_place_upd = request.POST.get('interview_place')
        interview_date_upd = request.POST.get('interview_date')
        interview_time_upd = request.POST.get('interview_time')

        obj = Job_Interview.objects.get(id=id)
        obj.interview_place = interview_place_upd
        obj.interview_date = interview_date_upd
        obj.interview_time = interview_time_upd
        obj.candidate_name_id = candidate_name_upd
        obj.job_title_id = job_title_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_job_interview')
    else:

        job_interview = Job_Interview.objects.get(id=id)
        data37 = Job_Post.objects.all()
        data38 = Job_Candidate.objects.all()
        return render(request,'recruitment/Job_Interview/update_job_interview.html',{"job_interview": job_interview,'data37':data37,'data38':data38})




def delete_job_interview_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Job_Interview.objects.get(id=id)
        except Job_Interview.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_job_interview')
    else:
        return render(request, 'recruitment/Job_Interview/delete_job_interview.html', )







def List_clients_data(request):
    data40 = Client.objects.all()


    return render(request, 'corehr/Project_Management/Clients/list_clients.html', {'data40':data40})



def Adding_clients_data(request):
    if request.method == 'POST':

        name_clients = request.POST.get('name')
        client_id = request.POST.get('client_ID')
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')

        obj = Client.objects.create(
            name=name_clients,
            client_id=client_id,
            mobile_no=mobile_no,
            email=email,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_clients')
    else:

        return render(request, 'corehr/Project_Management/Clients/add_clients.html')


def update_clients_data(request,id):
    if request.method == 'POST':

        name_clients_upd = request.POST.get('name')
        client_id_upd = request.POST.get('client_ID')
        mobile_no_upd = request.POST.get('mobile_no')
        email_upd = request.POST.get('email')



        obj = Client.objects.get(id=id)

        obj.name = name_clients_upd
        obj.client_id = client_id_upd
        obj.mobile_no = mobile_no_upd
        obj.email = email_upd

        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_clients')
    else:

        clients = Client.objects.get(id=id)

        return render(request, 'corehr/Project_Management/Clients/update_clients.html',{'clients':clients})




def delete_clients_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Client.objects.get(id=id)
        except Client.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_clients')
    else:
        return render(request, 'corehr/Project_Management/Clients/delete_clients.html')





def List_events_data(request):
    today = timezone.now().date()

    data41 = Events.objects.filter(event_date__gte=today)
    data22 = Company.objects.all()
    return render(request, 'Events&Meetings/Events/list_events.html', {'data41':data41,'data22':data22})



def Adding_events_data(request):
    if request.method == 'POST':
        company = request.POST.get('company')
        event_title = request.POST.get('event_title')
        event_date = request.POST.get('event_date')
        event_time = request.POST.get('end_time')
        description = request.POST.get('description')
        start_time = request.POST.get('start_time')

        obj = Events.objects.create(
            event_title=event_title,
            event_date=event_date,
            event_time=event_time,
            description=description,
            start_time=start_time,
            company_id=company,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_events')
    else:
        data22 = Company.objects.all()

        return render(request, 'Events&Meetings/Events/add_events.html',{'data22':data22} )





def update_events_data(request,id):
    if request.method == 'POST':

        company_upd = request.POST.get('company')
        event_title_upd = request.POST.get('event_title')
        event_date_upd = request.POST.get('event_date')
        event_time_upd = request.POST.get('end_time')
        start_time_upd = request.POST.get('start_time')
        description_upd = request.POST.get('description')


        obj = Events.objects.get(id=id)
        obj.event_title = event_title_upd
        obj.event_date = event_date_upd
        obj.event_time = event_time_upd
        obj.start_time = start_time_upd
        obj.description = description_upd
        obj.company_id = company_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_events')
    else:

        events = Events.objects.get(id=id)
        data22 = Company.objects.all()
        return render(request,'Events&Meetings/Events/update_events.html',{"events": events,'data22':data22})




def delete_events_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Events.objects.get(id=id)
        except Events.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_events')
    else:
        return render(request, 'Events&Meetings/Events/delete_events.html')





def List_meetings_data(request):
    today = timezone.now().date()

    data42 = Meetings.objects.filter(meeting_date__gte=today)
    data19 = Department.objects.all()
    return render(request, 'Events&Meetings/Meetings/list_meetings.html', {'data42':data42,'data19':data19})


def Adding_meetings_data(request):
    if request.method == 'POST':
        department = request.POST.get('department')
        meeting_title = request.POST.get('meeting_title')
        meeting_date = request.POST.get('meeting_date')
        meeting_time_str = request.POST.get('start_time')
        end_time_str = request.POST.get('end_time')

        # Convert string representations to datetime.time objects
        meeting_time = datetime.datetime.strptime(meeting_time_str, '%H:%M').time()
        end_time = datetime.datetime.strptime(end_time_str, '%H:%M').time()

        # Get the existing meetings for the specified date and department
        existing_meetings = Meetings.objects.filter(
            meeting_date=meeting_date
        )

        for existing_meeting in existing_meetings:
            if existing_meeting.meeting_time < meeting_time < existing_meeting.end_time or existing_meeting.meeting_time < end_time < existing_meeting.end_time:
                messages.error(request, 'New meeting time must not overlap with existing meetings',
                               extra_tags='alert alert-danger alert-dismissible fade show')
                return redirect('add_meetings')
        # Create and save the new meeting
        obj = Meetings.objects.create(
            meeting_title=meeting_title,
            meeting_date=meeting_date,
            meeting_time=meeting_time,
            end_time=end_time,
            department_id=department,
        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_meetings')
    else:
        data19 = Department.objects.all()
        return render(request, 'Events&Meetings/Meetings/add_meetings.html', {'data19': data19})






def update_meetings_data(request,id):
    if request.method == 'POST':
        department_upd = request.POST.get('department')
        meeting_title_upd = request.POST.get('meeting_title')
        meeting_date_upd = request.POST.get('meeting_date')
        meeting_time_upd = request.POST.get('start_time')
        end_time_upd = request.POST.get('end_time')

        obj = Meetings.objects.get(id=id)
        obj.meeting_title = meeting_title_upd
        obj.meeting_date = meeting_date_upd
        obj.meeting_time = meeting_time_upd
        obj.end_time = end_time_upd
        obj.department_id = department_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_meetings')
    else:

        meetings = Meetings.objects.get(id=id)
        data19 = Department.objects.all()
        return render(request,'Events&Meetings/Meetings/update_meetings.html',{"meetings": meetings,'data19':data19})




def delete_meetings_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Meetings.objects.get(id=id)
        except Meetings.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_meetings')
    else:
        return render(request, 'Events&Meetings/Meetings/delete_meetings.html', )


















def create(request):
    if request.is_ajax and request.method == "POST":
        name=request.POST.get('name')
        email=request.POST.get('email')
       # print(name,email)
        profile= Pro.objects.create(name=name,email=email)
        profile.save()
        success= 'User'+name+'created successfully'
        response_data = {'message': f"Form submitted successfully. Name: {name}, Email: {email}"}
        return HttpResponse(success)


def update_p(request,id):
    profile= get_object_or_404(Pro, pk=id)
    if request.method == 'POST':
        profile.name = request.POST['name']
        profile.description = request.POST['description']
        profile.save()
        return JsonResponse({'message': 'Item updated successfully.'})
    return JsonResponse({'message': 'Invalid request method.'}, status=400)


def get_item(request, id):
    item = get_object_or_404(Pro, pk=id)
    data = {'name': item.name, 'description': item.email}
    return JsonResponse(data)













def my_get_view(request):
    data = list(Pro.objects.values())
    return JsonResponse(data, safe=False)








def get_events(request):
    events = Event.objects.all()
    events_data = list(events.values('title', 'description', 'date'))


    return JsonResponse(events_data, safe=False)

   # events = Event.objects.all()

  #  events_data = list(events.values('title', 'description', 'date'))

   # return JsonResponse(events_data, safe=False)
import pandas as pd


# d=datetime.date()
#  print(d)
#   date_series = pd.date_range('08/10/2019', periods=12, freq='D')
#  print(date_series)
def test(request):



   return render(request,'hr/employee_details.html')

@login_required(login_url='/login/')
def home(request):
    if request.user.is_authenticated :

        user = CustomUser.objects.get(id=request.user.id)
        #print(user)
        if user.is_employee == True:
            # Retrieve events and festivals data from your models
            events = Events.objects.all()
            # festivals = Festival.objects.all()

            # Prepare data as a JSON object
            data = {
                'events': [
                    {
                        'title': event.event_title ,
                        'date': event.event_date.isoformat(),  # Assuming start_date is a DateTimeField
                         # Assuming end_date is a DateTimeField
                        'time':event.event_time.isoformat(),

                    }
                    for event in events
                ],

            }

            # Convert data to JSON and pass it to the template
            data_json = json.dumps(data)


            today = timezone.now().date()



            emp_data = EmployeeMaster.objects.get(user_id=user.id)
            public_holidays= Public_holidays.objects.all()
            birthdays = EmployeeMaster.objects.annotate(
                birthdate_month=ExtractMonth('BirthDate'),
                birthdate_day=ExtractDay('BirthDate')
            ).filter(
                Q(birthdate_month=today.month, birthdate_day__gte=today.day) |
                Q(birthdate_month__gt=today.month)
            ).order_by('birthdate_month', 'birthdate_day')[:3]
            holidays = Holidays.objects.filter(date__gte=today).order_by('date')[:5]

            events = Events.objects.filter(event_date__gte=today).order_by('event_date')[:5]
            meetings = Meetings.objects.filter(meeting_date=today).order_by('meeting_date')[:5]

            announcements = Announcements.objects.filter(start_date__lte=today,end_date__gte=today).order_by('start_date')
           # print(announcements)

            warnings = Warnings.objects.filter(warning_date=today)

            total_leves = LeaveRequest.objects.filter(employee_id=request.user.id).count()

            if total_leves > 0:
                total_leaves_count = \
                    LeaveRequest.objects.filter(employee_id=request.user.id).order_by('-id')[0]
                total_leaves_asper_emp = 24.0
                usedleaves_count = total_leaves_asper_emp - (
                        float(total_leaves_count.casual_leave_count) + float(
                    total_leaves_count.sick_leave_count) + float(total_leaves_count.emergency_leave_count))
                original_casual_leaves = 12.0
                used_casual_leaves = original_casual_leaves - float(total_leaves_count.casual_leave_count) - float(
                    total_leaves_count.emergency_leave_count)
                original_medical_leave = 12.0
                used_medical_leave = original_medical_leave - float(total_leaves_count.sick_leave_count)
            else:
                total_leaves_count = {'total_leaves_count': 0.0, 'sick_leave_count': 0.0, 'casual_leave_count': 0.0,
                                      'emergency_leave_count': 0.0,
                                      'loss_of_pay': 0.0}
                usedleaves_count = 24.0
                used_casual_leaves = 12.0
                used_medical_leave = 12.0

           # print(total_leaves_count)





            start_date =datetime.date.today()+datetime.timedelta(days=1)
            end_date = datetime.datetime(2023, 8, 1) # Two months have a maximum of 60 days
            punch_records = Attendance.objects.filter(employee=request.user.id, attendance_date__range=(end_date,start_date )).all()
            return render(request, 'dashboards.html',
                          {'meetings':meetings,'announcements':announcements,'emp_data': emp_data, 'public_holidays': public_holidays, 'birthdays': birthdays,'events':events,'holidays':holidays,
                           'punch_records': punch_records,'data_json': data_json,'total_leaves_count':total_leaves_count,'usedleaves_count':usedleaves_count,'used_casual_leaves':used_casual_leaves,'used_medical_leave':used_medical_leave,'warnings':warnings})
        elif user.is_manager == True:
            birthdays = EmployeeMaster.objects.all().order_by('BirthDate')[:3:]
            return render(request,'hr/hr_dashboard.html',{'birthdays':birthdays})

        elif user.is_hr == True:
            birthdays = EmployeeMaster.objects.all().order_by('BirthDate')[:3:]
            return render(request, 'hr/hr_dashboard.html',{'birthdays':birthdays})
    else:
        return render(request,'login1.html')

    #return render(request,'dashboards.html',{'emp_data': emp_data,'public_holidays':public_holidays,'birthdays':birthdays,'punch_records': punch_records})




@csrf_exempt
def fetch_festivals(request):
    if request.method == 'GET':
        # Fetch festivals data from your database or any other source
        holidays=Holidays.objects.all()
        festival_data = [{'title': festival.title, 'date': festival.date} for festival in holidays]
        return JsonResponse({'festivals': festival_data})


# views.py


def emp(request):
    form= EmployeeMaster.objects.all()

    k=type(form.CreatedDate)
    return render(request,'view.html',{'form':form,"k":k})
"""
def Register(request):
    if request.method == "GET":
        form=RegisterForm()

    return render(request,'signup.html',{'form':form})

"""

def signup_view(request):
    if request.method == 'POST':
        form = CustomSignupForm(request.POST)
        form1=EmployeeMasterForm(request.POST,request.FILES)
        if form.is_valid() and form1.is_valid():


            u= form.save()

            u = form.save()
            k = form1.save(commit=False)
            k.user = u
            k.save()

            return redirect('login')  # Redirect to login page after successful signup
    else:
        form = CustomUpdateForm()
        form1 = EmployeeMasterForm()

    return render(request, 'signup2.html', {'form': form,'form1':form1})





def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
           # print(user)
            auth_login(request, user)

            if user.is_employee == True:
                return redirect('home')
            elif  user.is_manager == True:
                return redirect('hr_dashboard')
            elif user.is_hr == True:
                return redirect('hr_dashboard')
            else:
                return redirect('login')




        else:
            # Invalid credentials
            messages.error(request, 'invalid credentials.',extra_tags='alert alert-danger alert-dismissible fade show')
            return render(request, 'login1.html', {'error': 'Invalid username or password'})
    else:
        form = AuthenticationForm()
        return render(request, 'login1.html', {'form': form})





def logout_view(request):
   logout(request)
   messages.info(request, 'successfully logout',extra_tags='alert alert-danger alert-dismissible fade show')
   return redirect('login')



def employee_detail(request,id):
    if request.method == "GET":
        user1 = CustomUser.objects.get(id=id)
        emp_obj = EmployeeMaster.objects.get(user_id=user1.id)
        return render(request,'details.html',{"emp_obj":emp_obj,"user":user1})





def company_detail(request, id):
    if request.method == "GET":
        user1 = CustomUser.objects.get(id=id)
        company_obj = Company.objects.get(user_id=user1.id)
        return render(request,'company.html',{"company_obj":company_obj,"user":user1})






def Register(request):
    if request.method == "GET":
        form = RegisterForm()
        form1=EmployeeMasterForm()
        return render(request,"signup2.html",{"form":form,"form1":form1})
    elif request.method == "POST":
        form = RegisterForm(request.POST)
        form1 = EmployeeMasterForm(request.POST,request.FILES)
       # print(form.is_valid)

        if  form.is_valid() and  form1.is_valid():
           # print(form)
            u=form.save()
            k=form1.save(commit=False)
            k.user=u
            k.save()
            return redirect('login')
        else:
            return HttpResponse("requist is not valid")





def update_emp(request,id):
        user = CustomUser.objects.get(id=id)
        emp_edit = EmployeeMaster.objects.get(user_id=user.id)
        if request.method == "POST":
            username = request.POST['username']
            firstname = request.POST['firstname']
            lastname = request.POST['lastname']
            email = request.POST['email']

            user.username = username
            user.first_name = firstname
            user.last_name =  lastname
            user.email = email
            user.save()
            form1 = EmployeeMasterForm(request.POST ,instance=emp_edit)


            if form1.is_valid():

                form1.save()
                return redirect("home")
               # user = form.save()
               # profile = form1.save(commit=False)
               # profile.user_id = user.id
              #  profile.save()



        else:
            form = CustomUpdateForm( instance=user)
            form1 = EmployeeMasterForm( instance=emp_edit)

            return render(request,'emp_update.html',{"form": form,"form1":form1,"user":user,"emp_edit":emp_edit})







def delete_emp(request,id):
    user = CustomUser.objects.get(id=id)

    user.is_active=False
    user.save()
    return redirect("emp_list")




def details_view(request,id):
    if request.method == "GET":
        user1 = CustomUser.objects.get(id=id)
        emp_obj = EmployeeMaster.objects.get(user_id=user1.id)
        if user1.is_employee == True:
            return render(request,'profile.html',{"emp_obj":emp_obj,"user":user1})
        else :
            return render(request, 'hr/profile.html', {"emp_obj": emp_obj, "user": user1})



    elif request.method == "POST" and request.FILES.get('img'):
        user1 = CustomUser.objects.get(id=id)
        emp_obj1 = EmployeeMaster.objects.get(user_id=user1.id)
       # username = request.POST['username']
        fullname = request.POST['fullName']
        phonenumber=request.POST['phone']
        email=request.POST['email']
        city = request.POST['city']
        address=request.POST['address']
        shifttime=request.POST.get('shift')
        MaratialStatus=request.POST.get('marital_status')
        Date_of_birth=request.POST['dob']
        profile= request.FILES['img']
        educationtype=request.POST.get('education type')
        designation=request.POST.get('designation')
        branch=request.POST.get('branch')
        passingyear=request.POST['passingyear']
        university = request.POST['university']

        #(fullname,phonenumber,email,address,shifttime,MaratialStatus,Date_of_birth,profile,educationtype,designation,branch,passingyear)


        user1.email=email
        user1.save()

        emp_obj1.fullname= fullname
        emp_obj1.phone_number=phonenumber
        emp_obj1.City = city
        emp_obj1.Address=address
        emp_obj1.Shift_time=shifttime
        emp_obj1.MaratialStatus=MaratialStatus
        emp_obj1.BirthDate="-".join(Date_of_birth.split("-"))


        emp_obj1.Education_type=educationtype
        emp_obj1.university = university
        emp_obj1.Designation=designation
        emp_obj1.Branch=branch
        emp_obj1.passing_year= passingyear
        emp_obj1.image =profile
        emp_obj1.save()
        return redirect('http://127.0.0.1:8000/details/{}/'.format(emp_obj1.user_id))

    elif request.method== "POST":
        username = request.POST['username']
        newpassword = request.POST['newpassword']
        confirmpassword = request.POST['confirmpassword']
        if newpassword !=confirmpassword:
            return HttpResponse("password is not same")
        u = CustomUser.objects.get(username__exact=username)
        u.set_password(confirmpassword)
        u.save()
        messages.success(request,"Successfully Change The Password",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('login')






       # print(fullname)
       # lastname = request.POST['lastname']
       # email = request.POST['email']

      #  user.username = username
      #  user.first_name = firstname
      #  user.last_name = lastname
      #  user.email = email
      #  user.save()
       # form1 = EmployeeMasterForm(request.POST, instance=emp_obj)

        #if form1.is_valid():
         #   form1.save()
         #   return redirect("home")
        # user = form.save()
        # profile = form1.save(commit=False)
        # profile.user_id = user.id
        #  profile.save()





      #  return render(request, 'profile.html', {"form": form, "form1": form1, "user": user, "emp_edit": emp_edit})










@login_required(login_url='/login/')
def Register_view(request):
    if request.method == 'GET':
        companies = Company.objects.all()
        departments = Department.objects.all()
        return render(request, 'signup4.html', {'companies': companies, 'departments': departments})

    elif request.method == 'POST' and request.FILES.get('profile_photo'):

        username = request.POST['username']

        confirmpassword = request.POST['conformpassword']
        encryptedpassword = request.POST['encrypted_password']

        fullname = request.POST['fullname']
        email = request.POST['email']
        gender = request.POST.get('gender')
        department = request.POST.get('department')
        company = request.POST['company']
        worktype = request.POST.get('worktype')
        dob = request.POST['dob']
        employeeid = request.POST['employeeid']
        doj = request.POST['dateofjoining']
        role = request.POST.get('role')
        shift = request.POST.get('shift')
        phone = request.POST['phonenumber']
        city = request.POST['city']
        pincode = request.POST['pincode']
        designation = request.POST.get('designation')

        profile = request.FILES['profile_photo']

        #print(fullname, dob ,department,phone,doj,worktype, designation,pincode,city,profile)

        plain_password = base64.b64decode(encryptedpassword).decode('utf-8')
        user=CustomUser.objects.create_user(username=username,email=email)
        user.set_password(plain_password)
        if designation == "employee":
            user.is_employee=True
        elif designation  == "hr":
            user.is_hr = True
        elif designation  == "manager":
            user.is_manager = True

        user.save()
        emp = EmployeeMaster.objects.create(user=user, EmployeeID=employeeid, Gender=gender, Role=role,
                                            Shift_time=shift, fullname=fullname, Department_id=department, BirthDate=dob,
                                            City=city, PinCode=pincode,phone_number=phone,
                                            JoiningDate=doj, image=profile )
        emp.save()
        messages.success(request, 'You are successfully signup',extra_tags='alert alert-success alert-dismissible fade show')  #
        return redirect('login')

    else:
        messages.success(request, 'You have singed up not successfully.',extra_tags='alert alert-error alert-dismissible fade show')
        return render(request, 'signup4.html')




@login_required(login_url="/login/")
def Add_employee_view(request):
    if request.method == 'GET':
        return render(request, 'add_employee.html')
    if request.method == 'POST' and request.FILES.get('profile_photo'):

        username = request.POST['username']
        password = request.POST['password']
        confirmpassword = request.POST['conformpassword']

        fullname = request.POST['fullname']
        email = request.POST['email']
        gender = request.POST.get('gender')
        department = request.POST.get('department')
        company = request.POST['company']
        worktype = request.POST.get('worktype')
        dob = request.POST['dob']
        employeeid = request.POST['employeeid']
        doj = request.POST['dateofjoining']
        role = request.POST.get('role')
        shift = request.POST.get('shift')
        phone = request.POST['phonenumber']
        city = request.POST['city']
        pincode = request.POST['pincode']
        designation = request.POST.get('designation')

        profile = request.FILES['profile_photo']

       # print(fullname, dob ,department,phone,doj,worktype, designation,pincode,city,profile)


        user=CustomUser.objects.create_user(username=username,email=email)
        user.set_password(password)
        if designation == "employee":
            user.is_employee=True
        elif designation  == "hr":
            user.is_hr = True
        elif designation  == "manager":
            user.is_manager = True

        user.save()
        emp = EmployeeMaster.objects.create(user=user, EmployeeID=employeeid, Gender=gender, Role=role,
                                            Shift_time=shift, fullname=fullname, Department=department, BirthDate=dob,
                                             City=city, PinCode=pincode,
                                            JoiningDate=doj, image=profile )
        emp.save()
        messages.success(request, 'You are successfully signup',extra_tags='alert alert-success alert-dismissible fade show')  #
        return redirect('login')

    else:
        messages.success(request, 'You have singed up not successfully.',extra_tags='alert alert-error alert-dismissible fade show')




def birthdays(request):
    birthday=EmployeeMaster.objects.all()
    return render(request,'k5.html',{'data':birthday})

















"""


def Registr(request):
    if request.method == 'GET':
        return render(request, 'signup.html')

    if request.method == 'POST':
        #form = RegisterForm(request.POST)
      #  form1 = EmployeeMasterForm(request.POST)
        username = request.POST['username']
        firstname = request.POST['firstname']
        lastname = request.POST['lastname']
        password = request.POST['password']
        confirmpassword = request.POST['cpwd']
        Email = request.POST['email']
        dob = request.POST['dob']
        empid = request.POST['empid']
        course = request.POST.get('course')
        department = request.POST.get('department')
        phone = request.POST['phno']
        gender = request.POST.get('gender')
        role = request.POST.get('role')
        shift = request.POST.get('shift')
        category = request.POST.get('category')

        print(username,firstname,password,Email, dob,empid ,course,department,phone, gender,role,shift,category)
        if len(username) < 5 or len(username) > 8:
            raise ValidationError(f'Length of the name:{username} is not between 5 -7 characters')
        if (password != confirmpassword):
            raise ValidationError("pasword and confirm password's is  not same")

        user=CustomUser.objects.create(username=username,first_name=firstname,last_name=lastname,password=password,email=Email)
        if category== "Employee":

            user.is_employee=True
        elif category == "Hr":
            user.is_hr = True
        elif category == "Manager":
            user.is_manager = True
        user.save()
        emp=EmployeeMaster.objects.create(user=user,EmployeeID=empid,Gender=gender,Role=role,Shift_time=shift)
        emp.save()
        messages.success(request, 'You .')  #
        return redirect('login')

    else:
        messages.success(request, 'You have singed up not successfully.')





def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        print(username)
        print(password)
        user = authenticate(username=username, password=password)
        print(user)
        if user is not None:
            login(request, user)
            return redirect('home')



        else:
            messages.error(request, 'Invalid credentials')
            return redirect('login')

          #  return redirect('login')

    else:
        form = AuthenticationForm()
        return render(request, 'login.html', {'form': form})

"""





def logout_view(request):
    logout(request)
    return redirect('login')




def Emp_list(request):

    user=CustomUser.objects.filter(is_active=True,is_employee=True).all()
    data=EmployeeMaster.objects.filter(user_id__in=[i.id for i in user]).all()
    return render(request,'emp_data.html',{'data':data})




def Total_Employees(request):

    user=CustomUser.objects.filter(is_employee=True).all()
    data=EmployeeMaster.objects.filter(user_id__in=[i.id for i in user]).all()
    return render(request,'hr/Total_Employees.html',{'data':data})





def Total_Employees_manager(request):

    user=CustomUser.objects.all()
    data=EmployeeMaster.objects.filter(user_id__in=[i.id for i in user]).all()
    return render(request,'hr/Total_Employee_manager.html',{'data':data})





def Birthday(request):
    pass













#Leave related views
def leave_request_view(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES)

        category = request.POST.get('category')
        startdate = request.POST['startdate']
        enddate = request.POST['enddate']
        totaldays = request.POST['totaldays']
        reason = request.POST['reason']
        halfday_vlue=request.POST.getlist('check')
       # secondhalfday= request.POST.get('check')


        #print(startdate,enddate,totaldays,category,reason,halfday_vlue)
        k = LeaveRequest.objects.filter(employee_id=request.user.id).all()

        leavecategory = get_object_or_404(Leavecategory, name=category)
      #  print(leavecategory)
        if len(halfday_vlue)==2:
            messages.warning(request, 'Please Select  Half-Day Leave Only',
                             extra_tags='alert alert-warning alert-dismissible fade show')


        elif category == "Casual Leave" and len(halfday_vlue) == 1:
            employee_id = request.user.id


            overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                            EndDate__gte=startdate, isfirst_halfday=False,
                                                            islast_halfday=False).exists()
            if overlapping_leave == True:
                messages.warning(request, 'Leave Is Already Applied For Given Dates',
                               extra_tags='alert alert-warning alert-dismissible fade show')
                return redirect('leave_request')
            else:
                if "start_date_half" in halfday_vlue:
                    first_half = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                             EndDate__gte=startdate, isfirst_halfday=True,
                                                             islast_halfday=False).exists()
                    if first_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason)

                        for i in halfday_vlue:
                            if i == "start_date_half":
                                if startdate == enddate:

                                    obj.isfirst_halfday = True

                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')

                                    return redirect('leave_request')
                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')
                elif "end_date_half" in halfday_vlue:
                    second_half = LeaveRequest.objects.filter(Q(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate) and Q(islast_halfday=True,
                                                                                              isfirst_halfday=False) | Q(
                        islast_halfday=True, isfirst_halfday=True)).exists()
                    if second_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason, comments=comments, )
                        for i in halfday_vlue:
                            if i == "end_date_half":
                                if startdate == enddate:
                                    obj.islast_halfday = True

                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')

                                    return redirect('leave_request')

                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')








        elif category == "Casual Leave" and len(halfday_vlue)==0:
            if startdate:
                employee_id = request.user.id
                overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate, isfirst_halfday=False,
                                                                islast_halfday=False).exists()

                if overlapping_leave == True:
                    messages.error(request, 'Leave Is Already Applied For Given Dates',
                                   extra_tags='alert alert-danger alert-dismissible fade show')
                    return redirect('leave_request')
                else:
                    if int(totaldays) <=5:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason)


                        obj.isfirst_halfday = False
                        obj.islast_halfday = False
                        obj.save()
                        messages.success(request, 'Leave Is Submitted Successfully',
                                     extra_tags='alert alert-success alert-dismissible fade show')


                        return redirect('leave_request')
                    else:
                        messages.warning(request, 'Please Change the Dates. Casual Leaves not exceed 5 days.',
                                         extra_tags='alert alert-warning alert-dismissible fade show')
                        return redirect('leave_request')




        elif category == "Medical Leave" and len(halfday_vlue) == 1:
            employee_id = request.user.id


            overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                            EndDate__gte=startdate, isfirst_halfday=False,
                                                            islast_halfday=False).exists()
            if overlapping_leave == True:
                messages.warning(request, 'Leave Is Already Applied For Given Dates',
                               extra_tags='alert alert-warning alert-dismissible fade show')
                return redirect('leave_request')
            else:
                if "start_date_half" in halfday_vlue:
                    first_half = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                             EndDate__gte=startdate, isfirst_halfday=True,
                                                             islast_halfday=False).exists()
                    if first_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason, comments=comments, )

                        for i in halfday_vlue:
                            if i == "start_date_half":
                                if startdate == enddate:
                                    obj.isfirst_halfday = True
                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')
                                    return redirect('leave_request')
                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')
                elif "end_date_half" in halfday_vlue:
                    second_half = LeaveRequest.objects.filter(Q(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate) and Q(islast_halfday=True,
                                                                                              isfirst_halfday=False) | Q(
                        islast_halfday=True, isfirst_halfday=True)).exists()
                    if second_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason, comments=comments,)
                        for i in halfday_vlue:
                            if i == "end_date_half":
                                if startdate == enddate:
                                    obj.islast_halfday = True
                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')
                                    return redirect('leave_request')

                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')





        elif category == "Medical Leave" and len(halfday_vlue) == 0:
            if startdate:
                employee_id = request.user.id
                overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate, isfirst_halfday=False,
                                                                islast_halfday=False).exists()

                if overlapping_leave == True:
                    messages.error(request, 'Leave Is Already Applied For Given Dates',
                                   extra_tags='alert alert-danger alert-dismissible fade show')
                    return redirect('leave_request')
                else:
                    obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                       EndDate=enddate,
                                       totaldays=totaldays, reason=reason)
                    obj.isfirst_halfday = False
                    obj.islast_halfday = False
                    obj.save()
                    messages.success(request, 'Leave Is Submitted Successfully',
                                     extra_tags='alert alert-success alert-dismissible fade show')
                    return redirect('leave_request')
        elif category == "Emergency Leave" and len(halfday_vlue) == 1:
            employee_id = request.user.id

            overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                            EndDate__gte=startdate, isfirst_halfday=False,
                                                            islast_halfday=False).exists()
            if overlapping_leave == True:
                messages.warning(request, 'Leave Is Already Applied For Given Dates',
                               extra_tags='alert alert-warning alert-dismissible fade show')
                return redirect('leave_request')
            else:
                if "start_date_half" in halfday_vlue:
                    first_half = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                             EndDate__gte=startdate, isfirst_halfday=True,
                                                             islast_halfday=False).exists()
                    if first_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason, comments=comments, )

                        for i in halfday_vlue:
                            if i == "start_date_half":
                                if startdate == enddate:
                                    obj.isfirst_halfday = True
                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')
                                    return redirect('leave_request')
                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')
                elif "end_date_half" in halfday_vlue:
                    second_half = LeaveRequest.objects.filter(Q(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate) and Q(islast_halfday=True,
                                                                                              isfirst_halfday=False) | Q(
                        islast_halfday=True, isfirst_halfday=True)).exists()
                    if second_half == True:
                        messages.warning(request, 'Half-day Leave Is Already Applied',
                                         extra_tags='alert alert-waring alert-dismissible fade show')
                        return redirect('leave_request')
                    else:
                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                           EndDate=enddate,
                                           totaldays=totaldays, reason=reason, comments=comments, )
                        for i in halfday_vlue:
                            if i == "end_date_half":
                                if startdate == enddate:
                                    obj.islast_halfday = True
                                    obj.save()
                                    messages.success(request, 'Leave Is Submitted Successfully',
                                                     extra_tags='alert alert-success alert-dismissible fade show')
                                    return redirect('leave_request')

                                else:
                                    messages.warning(request, 'half day leave works  with in the day only.',
                                                     extra_tags='alert alert-warning alert-dismissible fade show')
                                    return redirect('leave_request')



        elif category == "Emergency Leave" and len(halfday_vlue) == 0:
            if startdate:
                employee_id = request.user.id
                overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,
                                                                EndDate__gte=startdate, isfirst_halfday=False,
                                                                islast_halfday=False).exists()

                if overlapping_leave == True:
                    messages.error(request, 'Leave Is Already Applied For Given Dates',
                                   extra_tags='alert alert-danger alert-dismissible fade show')
                    return redirect('leave_request')
                else:
                    obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,
                                       EndDate=enddate,
                                       totaldays=totaldays, reason=reason)
                    obj.isfirst_halfday = False
                    obj.islast_halfday = False
                    obj.save()
                    messages.success(request, 'Leave Is Submitted Successfully',
                                     extra_tags='alert alert-success alert-dismissible fade show')
                    return redirect('leave_request')




        elif category == "Loss Of Pay" and len(halfday_vlue) == 0:

            if startdate:

                employee_id = request.user.id

                overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,

                                                                EndDate__gte=startdate, isfirst_halfday=False,

                                                                islast_halfday=False).exists()

                if overlapping_leave == True:

                    messages.error(request, 'Leave Is Already Applied For Given Dates',

                                   extra_tags='alert alert-danger alert-dismissible fade show')

                    return redirect('leave_request')

                else:

                    obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,

                                       EndDate=enddate,

                                       totaldays=totaldays, reason=reason)

                    obj.isfirst_halfday = False

                    obj.islast_halfday = False

                    obj.save()

                    messages.success(request, 'Leave Is Submitted Successfully',

                                     extra_tags='alert alert-success alert-dismissible fade show')

                    return redirect('leave_request')

        elif category == "Loss Of Pay" and len(halfday_vlue) == 1:

            employee_id = request.user.id

            overlapping_leave = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,

                                                            EndDate__gte=startdate, isfirst_halfday=False,

                                                            islast_halfday=False).exists()

            if overlapping_leave == True:

                messages.warning(request, 'Leave Is Already Applied For Given Dates',

                                 extra_tags='alert alert-warning alert-dismissible fade show')

                return redirect('leave_request')

            else:

                if "start_date_half" in halfday_vlue:

                    first_half = LeaveRequest.objects.filter(employee_id=employee_id, StartDate__lte=startdate,

                                                             EndDate__gte=startdate, isfirst_halfday=True,

                                                             islast_halfday=False).exists()

                    if first_half == True:

                        messages.warning(request, 'Half-day Leave Is Already Applied',

                                         extra_tags='alert alert-waring alert-dismissible fade show')

                        return redirect('leave_request')

                    else:

                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,

                                           EndDate=enddate,

                                           totaldays=totaldays, reason=reason, comments=comments, )

                        for i in halfday_vlue:

                            if i == "start_date_half":

                                if startdate == enddate:

                                    obj.isfirst_halfday = True

                                    obj.save()

                                    messages.success(request, 'Leave Is Submitted Successfully',

                                                     extra_tags='alert alert-success alert-dismissible fade show')

                                    return redirect('leave_request')

                                else:

                                    messages.warning(request, 'half day leave works  with in the day only.',

                                                     extra_tags='alert alert-warning alert-dismissible fade show')

                                    return redirect('leave_request')

                elif "end_date_half" in halfday_vlue:

                    second_half = LeaveRequest.objects.filter(Q(employee_id=employee_id, StartDate__lte=startdate,

                                                                EndDate__gte=startdate) and Q(islast_halfday=True,

                                                                                              isfirst_halfday=False) | Q(

                        islast_halfday=True, isfirst_halfday=True)).exists()

                    if second_half == True:

                        messages.warning(request, 'Half-day Leave Is Already Applied',

                                         extra_tags='alert alert-waring alert-dismissible fade show')

                        return redirect('leave_request')

                    else:

                        obj = LeaveRequest(employee_id=employee_id, LeaveCategory=leavecategory, StartDate=startdate,

                                           EndDate=enddate,

                                           totaldays=totaldays, reason=reason, comments=comments, )

                        for i in halfday_vlue:

                            if i == "end_date_half":

                                if startdate == enddate:

                                    obj.islast_halfday = True

                                    obj.save()

                                    messages.success(request, 'Leave Is Submitted Successfully',

                                                     extra_tags='alert alert-success alert-dismissible fade show')

                                    return redirect('leave_request')


                                else:

                                    messages.warning(request, 'half day leave works  with in the day only.',

                                                     extra_tags='alert alert-warning alert-dismissible fade show')

                                    return redirect('leave_request')

    else:
        total_leves = LeaveRequest.objects.filter(employee_id=request.user.id,status="Approved").count()

        if total_leves > 0:
            total_leaves_count = \
            LeaveRequest.objects.filter(employee_id=request.user.id, status="Approved").order_by('-id')[0]
            total_leaves_asper_emp = 24.0
            usedleaves_count = total_leaves_asper_emp - (
                    float(total_leaves_count.casual_leave_count) + float(total_leaves_count.sick_leave_count) + float(
                total_leaves_count.emergency_leave_count))
            original_casual_leaves = 12.0
            used_casual_leaves = original_casual_leaves - float(total_leaves_count.casual_leave_count) - float(
                total_leaves_count.emergency_leave_count)
            original_medical_leave = 12.0
            used_medical_leave = original_medical_leave - float(total_leaves_count.sick_leave_count)
            # print(float(total_leaves_count.sick_leave_count))
            # usedleaves_count=int(total_leaves_count.total_leaves_count)-(int(total_leaves_count.casual_leaves_count)+int(total_leaves_count.sick_leaves_count))
        else:
            total_leaves_count = {'total_leaves_count': 0.0, 'sick_leave_count': 0.0, 'casual_leave_count': 0.0,'emergency_leave_count':0.0,
                                  'loss_of_pay': 0.0}
            usedleaves_count = 24.0
            used_casual_leaves = 12.0
            used_medical_leave = 12.0

       # print(total_leaves_count)
        casual_l_c=LeaveRequest.objects.filter(employee_id=request.user.id,casual_leave_count=12.0,status="Approved").exists()
        sick_l_c=LeaveRequest.objects.filter(employee_id=request.user.id,sick_leave_count=12.0,status="Approved").exists()

        context = {
            'total_leaves_count': total_leaves_count,
            'usedleaves_count': usedleaves_count,
            'used_casual_leaves': used_casual_leaves,
            'used_medical_leave': used_medical_leave,
            'casual_l_c':casual_l_c,
            'sick_l_c':sick_l_c,
        }


    return render(request, 'leave/leave_request.html', context)


def leave_request_success_view(request):
    return render(request, 'leave/leave_request_success.html')
def leave_approval_view(request, request_id):
    leave_request = get_object_or_404(LeaveRequest, id=request_id)


    if request.method == 'POST':
        # Handle the approval logic

        approval_status = request.POST.get('approval_status')
       # print(approval_status)

        comments = request.POST.get('comments')
        c= LeaveRequest.objects.get(id=request_id)
        k= LeaveRequest.objects.filter(employee_id=c.employee_id,status="Approved").all()
        if k:
            get_category = LeaveRequest.objects.filter(id=request_id,status="Pending").order_by('-total_leaves_count')[0]
            get_category_approve = LeaveRequest.objects.filter(employee_id=c.employee_id, status="Approved").order_by('-total_leaves_count')[0]
            confirm_leave_catagory = Leavecategory.objects.get(id=get_category.LeaveCategory_id)

            if get_category.totaldays and (
                    confirm_leave_catagory.name == "Casual Leave"):
               # print(approval_status)


                if (approval_status == "Approved"):

                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    #print(type(float(get_category_approve.casual_leave_count)))
                    #print(type(get_category.totaldays),get_category.totaldays)
                    leave_request.casual_leave_count = float(get_category_approve.casual_leave_count)+float(get_category.totaldays)
                    leave_request.sick_leave_count=float(get_category_approve.sick_leave_count)
                    leave_request.emergency_leave_count = float(get_category_approve.emergency_leave_count)

                    leave_request.total_leaves_count = float(get_category_approve.total_leaves_count) +float(get_category.totaldays)

                    leave_request.total_leave_days = float(get_category_approve.total_leave_days) - float(get_category.totaldays)
                    leave_request.emergency_leave_count = get_category_approve.emergency_leave_count
                    leave_request.loss_of_pay = float(get_category_approve.loss_of_pay)
                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.casual_leave_count = get_category_approve.casual_leave_count
                    leave_request.sick_leave_count=get_category_approve.sick_leave_count
                    leave_request.total_leaves_count = get_category_approve.total_leaves_count
                    leave_request.total_leave_days=get_category_approve.total_leave_days
                    leave_request.loss_of_pay = get_category_approve.loss_of_pay
                    leave_request.emergency_leave_count = get_category_approve.emergency_leave_count
                else:
                    leave_request.status = "Pending"
            elif get_category.totaldays and (
                    confirm_leave_catagory.name == "Emergency Leave"):
               # print(approval_status)

                if (approval_status == "Approved"):

                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    leave_request.emergency_leave_count = float(get_category_approve.emergency_leave_count)+float(get_category.totaldays)

                    leave_request.casual_leave_count = float(get_category_approve.casual_leave_count)
                    leave_request.sick_leave_count=float(get_category_approve.sick_leave_count)

                    leave_request.total_leaves_count = float(get_category_approve.total_leaves_count) +float(get_category.totaldays)

                    leave_request.total_leave_days = float(get_category_approve.total_leave_days) - float(get_category.totaldays)
                    leave_request.loss_of_pay = float(get_category_approve.loss_of_pay)
                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.casual_leave_count = get_category_approve.casual_leave_count
                    leave_request.sick_leave_count=get_category_approve.sick_leave_count
                    leave_request.emergency_leave_count= get_category_approve.emergency_leave_count
                    leave_request.total_leaves_count = get_category_approve.total_leaves_count
                    leave_request.total_leave_days=get_category_approve.total_leave_days
                    leave_request.loss_of_pay = get_category_approve.loss_of_pay

                else:
                    leave_request.status = "Pending"








            elif get_category.totaldays and confirm_leave_catagory.name == "Medical Leave":
                if (approval_status == "Approved"):

                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    leave_request.sick_leave_count = float(get_category_approve.sick_leave_count) +float(get_category.totaldays)
                    leave_request.total_leaves_count = float(get_category_approve.total_leaves_count) +float(get_category.totaldays)
                    leave_request.casual_leave_count = float(get_category_approve.casual_leave_count)
                    leave_request.loss_of_pay = float(get_category_approve.loss_of_pay)
                    leave_request.emergency_leave_count = float(get_category_approve.emergency_leave_count)


                    leave_request.total_leave_days = float(get_category_approve.total_leave_days) -float(get_category.totaldays)



                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.sick_leave_count = get_category_approve.sick_leave_count
                    leave_request.casual_leave_count=get_category_approve.casual_leave_count
                    leave_request.emergency_leave_count = get_category_approve.emergency_leave_count
                    leave_request.total_leaves_count = get_category_approve.total_leaves_count

                else:
                    leave_request.status = "Pending"

            elif get_category.totaldays and confirm_leave_catagory.name == "Loss Of Pay":
                if (approval_status == "Approved"):
                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    leave_request.loss_of_pay = float(get_category_approve.loss_of_pay) + float(
                        get_category.totaldays)
                    leave_request.total_leaves_count = float(get_category_approve.total_leaves_count) + float(
                        get_category.totaldays)
                    leave_request.casual_leave_count = get_category_approve.casual_leave_count
                    leave_request.sick_leave_count = get_category_approve.sick_leave_count
                    leave_request.emergency_leave_count = get_category_approve.emergency_leave_count

                    leave_request.total_leave_days = float(get_category_approve.total_leave_days) - float(
                        get_category.totaldays)
                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.loss_of_pay = get_category_approve.loss_of_pay
                    leave_request.sick_leave_count = get_category_approve.sick_leave_count
                    leave_request.emergency_leave_count = get_category_approve.emergency_leave_count
                    leave_request.casual_leave_count = get_category_approve.casual_leave_count
                    leave_request.total_leaves_count = get_category_approve.total_leaves_count

                else:
                    leave_request.status = "Pending"






        else:
            get_category = LeaveRequest.objects.filter(employee_id=c.employee_id, status="Pending").order_by('-total_leaves_count')[0]
            confirm_leave_catagory = Leavecategory.objects.get(id=get_category.LeaveCategory_id)

            if get_category.totaldays and (confirm_leave_catagory.name == "Casual Leave"):
                if (approval_status == "Approved"):
                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                  #  print( int(get_category.casual_leave_count)+ get_category.totaldays)
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)+float(get_category.totaldays)
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)

                    leave_request.total_leaves_count = float(get_category.total_leaves_count) + float(get_category.totaldays)

                    leave_request.total_leave_days = float(get_category.total_leave_days) - float(get_category.totaldays)


                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)
                    leave_request.total_leaves_count = float(get_category.total_leaves_count)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)
                    leave_request.total_leave_days = float(get_category.total_leave_days)

                else:
                   leave_request.status = "Pending"
            elif get_category.totaldays and ( confirm_leave_catagory.name == "Emergency Leave"):
                if (approval_status == "Approved"):
                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    #print( float(get_category.casual_leave_count)+ get_category.totaldays)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count) + float(get_category.totaldays)
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)

                    leave_request.total_leaves_count = float(get_category.total_leaves_count) + float(get_category.totaldays)

                    leave_request.total_leave_days = float(get_category.total_leave_days) - float(get_category.totaldays)


                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)
                    leave_request.total_leaves_count = float(get_category.total_leaves_count)
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)
                    leave_request.total_leave_days = float(get_category.total_leave_days)

                else:
                   leave_request.status = "Pending"
            elif get_category.totaldays and confirm_leave_catagory.name == "Medical Leave":
                if (approval_status == "Approved"):

                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)+ float(get_category.totaldays)
                    leave_request.casual_leave_count= float(get_category.casual_leave_count)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)
                    leave_request.total_leaves_count =float(get_category.total_leaves_count)+float(get_category.totaldays)

                    leave_request.total_leave_days = float(get_category.total_leave_days) - float(get_category.totaldays)



                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.sick_leave_count = get_category.sick_leave_count
                    leave_request.total_leaves_count = get_category.total_leaves_count
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)

                else:
                    leave_request.status = "Pending"
            elif get_category.totaldays and confirm_leave_catagory.name == "Loss Of Pay":
                if (approval_status == "Approved"):

                    leave_request.status = "Approved"
                    leave_request.IsActive = False
                    leave_request.loss_of_pay = float(get_category.loss_of_pay)+ float(get_category.totaldays)
                    leave_request.casual_leave_count= float(get_category.casual_leave_count)
                    leave_request.sick_leave_count = float(get_category.sick_leave_count)
                    leave_request.total_leaves_count =float(get_category.total_leaves_count)+float(get_category.totaldays)
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)

                    leave_request.total_leave_days = float(get_category.total_leave_days) - float(get_category.totaldays)



                elif (approval_status == "Rejected"):
                    leave_request.status = "Rejected"
                    leave_request.IsActive = False
                    leave_request.loss_of_pay=get_category.loss_of_pay
                    leave_request.sick_leave_count = get_category.sick_leave_count
                    leave_request.total_leaves_count = get_category.total_leaves_count
                    leave_request.emergency_leave_count = float(get_category.emergency_leave_count)
                    leave_request.casual_leave_count = float(get_category.casual_leave_count)

                else:
                    leave_request.status = "Pending"









        leave_request.comments = comments
        leave_request.approvedby = request.user.username

        # Create a notification
        recipient = leave_request.employee
       # print(recipient)
        message = "Your leave request with {},Id{} has been {}".format(str(leave_request.employee),leave_request.id, approval_status)
        notification = Notification(recipient=recipient, message=message,  notification_type="Leave")
        notification.save()
        leave_request.save()

        if request.user.is_employee == "True":
            return redirect('leave_request_list')  # Redirect to a leave request list page or success page
        else :
            return redirect('leave_approval_pending')
    else:
        context = {
            'leave_request':leave_request
        }
        return render(request,"leave/approval.html",context)



def leave_approvals_pending(request):
    #pending_requests = CustomUser.objects.filter(is_active=True, is_employee=True, leaverequest__status='Pending').all()

    pending_requests = LeaveRequest.objects.filter(
        IsActive=True,
        employee__is_active=True,
        employee__is_employee=True
    ).all()
    return render(request,'leave/pending_approvals.html',{'pendingapprovals':pending_requests})





#manager or Hr side

def leave_request_list_view(request):
    if request.user.is_authenticated:
        user = CustomUser.objects.get(id=request.user.id)
        emp = EmployeeMaster.objects.filter(user_id=user.id)
    leave_requests = LeaveRequest.objects.filter(employee_id=user.id).order_by("StartDate")

    context = {
        'leave_requests': leave_requests,
     }
    return render(request, 'leave/leave_request_list.html', context)




def leave_data_employee(request):
    if request.user.is_authenticated:
        user = CustomUser.objects.get(id=request.user.id)
        emp = EmployeeMaster.objects.filter(user_id=user.id)
    leave_requests = LeaveRequest.objects.filter(employee_id=user.id)
    context = {
        'leave_requests': leave_requests,
     }
    return render(request, 'hr/emp.edit.html', context)






#manager or Hr side
"""
def leave_request_list_view(request):
   #leave_requests = LeaveRequest.objects.filter(IsActive=True)
   leave_requests = LeaveRequest.objects.all()
   context = {
        'leave_requests': leave_requests,
    }
   return render(request, 'leave/leave_request_list.html', context)
"""



def notification_view(request):
   # notification =Notification.objects.all()
    notifications = Notification.objects.filter(status=False).count()

       # Notification.objects.filter(recipient=request.user).order_by('-timestamp')
    context = {
        'notifications': notifications
    }
    return render(request, 'leave/notification.html', context)



def Notifyed(request,id):

    obj=Notification.objects.get(id=id)
    obj.status=True
    obj.save()
    user =CustomUser.objects.get(id=obj.recipient.id)

    id_1=EmployeeMaster.objects.get(user_id=user.id)
    #print(id_1.user_id)

    return redirect('http://127.0.0.1:8000/details/{}/'.format(id_1.user_id))













def permission_request_view(request):
    if request.method== "POST":

        permission_category = request.POST.get('category')
        startdate = request.POST['startdate']
        enddate = request.POST['enddate']
        choose_time = request.POST.get('choose_time')
        location = request.POST['location']
        reason= request.POST['reason']

        user = CustomUser.objects.get(id=request.user.id)
        current_datetime = datetime.datetime.now()
        if permission_category =="WFH" :
           wfh = Permission.objects.filter(Q(employee_id=user.id),Q(permission_category=permission_category),
                                           (Q(EndDate_gte=startdate) & Q(StartDate_lte=startdate)) | (
                                                       Q(EndDate_gte=enddate) & Q(StartDate_lte=enddate))
                                           ).exists()
           if wfh:
                messages.warning(request,
                                 "Already Applied {} Request".format(permission_category),
                                 extra_tags='alert alert-warning alert-dismissible fade show')
           else:
                has_permission = Permission.objects.filter(Q(employee_id=user.id),
                                                           Q(permission_category="Early Login") | Q(
                                                               permission_category="Late Login"),
                                                           (Q(EndDate_lte=startdate) & Q(StartDate_gte=startdate))).exists()
                if has_permission:
                    messages.warning(request,
                                     "Already Applied Permission Request By Today. Please Change The Date Period.",
                                     extra_tags='alert alert-warning alert-dismissible fade show')
                else:
                    obj = Permission(employee_id=user.id, permission_category=permission_category,
                                       StartDate=startdate,
                                       EndDate=enddate,
                                       location=location,
                                       reason=reason)
                    obj.save()
                    obj1 = PermissionApproval(permission=obj)
                    obj1.save()
                    messages.success(request, "Successfully Saved Data",
                         extra_tags='alert alert-success alert-dismissible fade show')
                    return redirect('permission_request')





























        elif permission_category:

              print(current_datetime)

              has_permission = Permission.objects.filter(Q(employee_id=user.id),Q(permission_category=permission_category),Q(EndDate=current_datetime),Q(StartDate=current_datetime)).exists()

              wfh_permission = Permission.objects.filter(Q(employee_id=user.id),
                                                         Q(permission_category="WFH"),
                                                              Q(EndDate__gte=startdate) & Q(StartDate__lte=startdate)).exists()


              if has_permission  :

                  messages.warning(request,
                                   "Already Applied {} Request".format(permission_category),
                                   extra_tags='alert alert-warning alert-dismissible fade show')
              elif wfh_permission :
                          messages.warning(request,
                                           "Already Applied Work From Home Request ",
                                           extra_tags='alert alert-warning alert-dismissible fade show')
              else:
                  if permission_category == "Early Login":
                      choose_time = time(int(choose_time.split(':')[0]), int(choose_time.split(':')[1]))

                      comparison_times = [
                          time(6, 30),
                          time(9, 30),
                          time(14, 0)
                      ]

                      current_datetime = datetime.datetime.now()
                      current_date = current_datetime.date()
                      choose_datetime = datetime.datetime(current_date.year, current_date.month, current_date.day,
                                                          choose_time.hour, choose_time.minute)

                      for comparison_time in comparison_times:
                          comparison_datetime = datetime.datetime(current_date.year, current_date.month,
                                                                  current_date.day,
                                                                  comparison_time.hour, comparison_time.minute)

                          time_diff = choose_datetime - comparison_datetime
                          time_diff_hours = time_diff.total_seconds() / 3600  # Convert time difference to hours

                          if choose_time < comparison_time and -2 <= time_diff_hours <= 0:

                              has_permission = Permission.objects.filter(Q(employee_id=user.id),
                                                                         Q(permission_category=permission_category) | Q(
                                                                             permission_category="Late Login"),
                                                                         Q(EndDate=current_datetime),
                                                                         Q(StartDate=current_datetime)).exists()

                              if has_permission:
                                  messages.warning(request,
                                                   f"Already You Applied {permission_category} Request.",
                                                   extra_tags='alert alert-warning alert-dismissible fade show')
                                  return render(request, 'permission/permission_request.html')
                              else:
                                  obj = Permission(employee_id=user.id, permission_category=permission_category,
                                                   StartDate=startdate, EndDate=enddate, choose_time=choose_time,
                                                   reason=reason)
                                  obj.save()
                                  obj1 = PermissionApproval(permission=obj)
                                  obj1.save()
                                  messages.success(request, "Successfully Saved Data",
                                                   extra_tags='alert alert-success alert-dismissible fade show')
                                  return redirect('permission_request')

                      messages.warning(request, "Time should not exceed more than two and apply before your shift time start's",
                                       extra_tags='alert alert-warning alert-dismissible fade show')


                  elif permission_category == "Late Login":

                      choose_time = time(int(choose_time.split(':')[0]), int(choose_time.split(':')[1]))

                      # Assuming you have a list of comparison times

                      comparison_times = [

                          time(6, 30),

                          time(9, 30),  # Add your specific time, for example, 2:00 PM

                          time(14, 0)  # Add another specific time, for example, 6:00 PM

                      ]

                      current_datetime = datetime.datetime.now()

                      current_date = current_datetime.date()

                      choose_datetime = datetime.datetime(current_date.year, current_date.month, current_date.day,

                                                          choose_time.hour, choose_time.minute)

                      for comparison_time in comparison_times:

                          comparison_datetime = datetime.datetime(current_date.year, current_date.month,
                                                                  current_date.day,

                                                                  comparison_time.hour, comparison_time.minute)

                          time_diff = choose_datetime - comparison_datetime

                          time_diff_hours = time_diff.total_seconds() / 3600  # Convert time difference to hours

                          if choose_time > comparison_time and 0 <= time_diff_hours <= 2:

                              has_permission = Permission.objects.filter(Q(employee_id=user.id),

                                                                         Q(permission_category=permission_category) | Q(

                                                                             permission_category="Early Login"),

                                                                         Q(EndDate=current_datetime),

                                                                         Q(StartDate=current_datetime)).exists()


                              if has_permission:

                                  messages.warning(request,
                                                   "Already You Applied Early Login Request.",
                                                   extra_tags='alert alert-warning alert-dismissible fade show')

                                  return render(request, 'permission/permission_request.html')

                              else:

                                  obj = Permission(employee_id=user.id, permission_category=permission_category,

                                                   StartDate=startdate, EndDate=enddate, choose_time=choose_time,
                                                   reason=reason)

                                  obj.save()

                                  obj1 = PermissionApproval(permission=obj)

                                  obj1.save()

                                  messages.success(request, "Successfully Saved Data",

                                                   extra_tags='alert alert-success alert-dismissible fade show')

                                  return redirect('permission_request')


                      messages.warning(request, "Time should not exceed more than two and apply after your shift time start's",extra_tags='alert alert-warning alert-dismissible fade show')



                  elif permission_category == "Early Logout":
                      choose_time = time(int(choose_time.split(':')[0]), int(choose_time.split(':')[1]))

                      # Assuming you have three specified comparison times
                      comparison_times = [
                          time(14, 0),  # 5:00 PM
                          time(19, 0),  # 6:00 PM
                          time(22, 0),  # 7:00 PM
                          # Add other specific times as needed
                      ]

                      current_datetime = datetime.datetime.now()
                      current_date = current_datetime.date()
                      choose_datetime = datetime.datetime(current_date.year, current_date.month, current_date.day,
                                                          choose_time.hour, choose_time.minute)


                      for comparison_time in comparison_times:
                          comparison_datetime = datetime.datetime(current_date.year, current_date.month,
                                                                  current_date.day,
                                                                  comparison_time.hour, comparison_time.minute)

                          time_diff = choose_datetime - comparison_datetime
                          time_diff_hours = time_diff.total_seconds() / 3600  # Convert time difference to hours

                          if choose_time < comparison_time and time_diff_hours >= -2:
                              has_permission = Permission.objects.filter(Q(employee_id=user.id),
                                                                         Q(permission_category=permission_category),
                                                                         Q(EndDate__lte=current_datetime),
                                                                         Q(StartDate__gte=current_datetime)).exists()

                              if has_permission:
                                  messages.warning(request,
                                                   f"Already you applied {permission_category} request.",
                                                   extra_tags='alert alert-success alert-dismissible fade show')
                                  return render(request, 'permission/permission_request.html')
                              else:
                                  obj = Permission(employee_id=user.id, permission_category=permission_category,
                                                   StartDate=startdate, EndDate=enddate, choose_time=choose_time,
                                                   reason=reason)
                                  obj.save()
                                  obj1 = PermissionApproval(permission=obj)
                                  obj1.save()
                                  messages.success(request, "Successfully Saved Data",
                                                   extra_tags='alert alert-success alert-dismissible fade show')
                                  return redirect('permission_request')

                      messages.warning(request, "Time should not exceed more than two and apply before your shift time end's",
                                       extra_tags='alert alert-warning alert-dismissible fade show')







                  elif permission_category == "WFH":

                    has_permission = Permission.objects.filter(Q(employee_id=user.id),
                                                               Q(permission_category=permission_category),
                                                               Q(EndDate__lte=current_datetime),
                                                               Q(StartDate__gte=current_datetime)).exists()
                    if has_permission:
                        messages.warning(request,
                                         "Already you applied {} request .".format(permission_category),5,extra_tags='alert alert-success alert-dismissible fade show')

                        return render(request, 'permission/permission_request.html')



                    else:
                        obj = Permission(employee_id=user.id, permission_category=permission_category,
                                         StartDate=startdate,
                                         EndDate=enddate,choose_time=choose_time,
                                         Time_duration=duration,
                                         reason=reason)

                        obj.save()
                        obj1 = PermissionApproval(permission=obj)
                        obj1.save()
                        messages.success(request, "Successfully Saved Data",
                                         extra_tags='alert alert-success alert-dismissible fade show')
                        return redirect('permission_request')



        else:
            messages.warning(request,
                             "please enter permission category",
                             extra_tags='alert alert-success alert-dismissible fade show')

    return render(request,'permission/permission_request.html')

def permission_request_success_view(request):
    return render(request, 'permission/permission_request_success.html')










def permission_approval_view(request, request_id):
    permission = get_object_or_404(Permission, id=request_id)
    permission_request=PermissionApproval.objects.get(permission_id=permission.id)

    if request.method == 'POST':
        # Handle the approval logic
        approval_status = request.POST.get('approval_status')
       # print(approval_status)
        comments = request.POST.get('comments')
        if (approval_status == "approved"):
            permission_request.approval_status= "Approved"
            permission.status = False
        elif (approval_status == "rejected"):
            permission_request.approval_status= "Rejected"
            permission.status = False

        else:
            permission_request.approval_status== "Pending"



        permission_request.comments = comments
        permission_request.approver= request.user.username

        #permission_request.approver=  approvername

        # Create a notification
        recipient = permission.employee
       # print(recipient)
        message = "Your leave request with {},ID {} has been {}".format(str( permission.employee), permission.id, approval_status)
        notification = Notification(recipient=recipient, message=message, notification_type= "permission")
        notification.save()
        permission.save()
        permission_request.save()
        messages.success(request,"saved data successfully", extra_tags='alert alert-success')

        return redirect('pending_permission')  # Redirect to a leave request list page or success page

    context = {
        'permission': permission,
    }
    return render(request, 'permission/permisson_approval.html', context)




def permission_request_list_view(request):
    if request.user.is_authenticated:
        if request.user.is_employee == True:
            user = CustomUser.objects.get(id=request.user.id)
            # emp = EmployeeMaster.objects.filter(user_id=user.id)
            permission = Permission.objects.all().filter(employee_id=user.id)
          #  print(permission)
            #leave_requests = LeaveRequest.objects.all()
            permissionapprove=PermissionApproval.objects.all().filter(permission_id__in=[k.id  for k in permission])
            context = {
               "list": zip(permission,permissionapprove)
            }
            return render(request, 'permission/permission_request_list.html', context)
        if request.user.is_hr== True:

            user = CustomUser.objects.get(id=request.user.id)
            # emp = EmployeeMaster.objects.filter(user_id=user.id)
            permission = Permission.objects.all().filter(employee_id=user.id)
            #print(permission)
            #leave_requests = LeaveRequest.objects.all()
            permissionapprove=PermissionApproval.objects.all().filter(permission_id__in=[k.id  for k in permission])
            context = {
               "list": zip(permission,permissionapprove)
            }
            return render(request, 'hr/pending_permission_request_list.html', context)

def Pending_permission(request):
    if request.user.is_manager == True:

        permissionspending = PermissionApproval.objects.filter(approval_status="Pending").all()
        permission_data=Permission.objects.filter(id__in=[k.permission_id for k in permissionspending]).all()
       # print(permission_data)
        emp_data=[]
        for k in permission_data:
            if k:
                employeedata= CustomUser.objects.filter(id=k.employee_id).all()
                emp_data.append(employeedata)


        context = {
            'pending_permission_list': zip(permissionspending,permission_data,emp_data)
        }

        return render(request, 'hr/pending_permission_request_list.html', context)
    return render(request, 'hr/pending_permission_request_list.html')


def Total_Pending_permission(request):
    if request.user.is_manager == True:
        today = timezone.now().date()

        permissionspending = PermissionApproval.objects.filter(permission__StartDate__lte=today,permission__EndDate__gte=today,approval_status="Approved").all()
        permission_data=Permission.objects.filter(id__in=[k.permission_id for k in permissionspending]).all()
       # print(permission_data)
        emp_data=[]
        for k in permission_data:
            if k:
                employeedata= CustomUser.objects.filter(id=k.employee_id).all()
                emp_data.append(employeedata)


        context = {
            'pending_permission_list': zip(permissionspending,permission_data,emp_data)
        }

        return render(request, 'permission/total_permissions.html', context)
    return render(request, 'permission/total_permissions.html')

def faq(request):
    return render(request,'hr/faq.html')

def faq_employee(request):
    return render(request,'hr/faqemployee.html')





def Attendance_view(request):
    user = CustomUser.objects.get(id=request.user.id)
    emp_data = EmployeeMaster.objects.get(user_id=user.id)
    return render(request, 'attendence/attendence.html',{'emp_data':emp_data})




@csrf_exempt
def punch(request):
    user = CustomUser.objects.get(id=request.user.id)
    if request.method=='POST':
        punch=request.POST.get('action')
      #  print(punch)







        if punch == "Punch_In":
            obj = Attendance.objects.create(employee_id=user.id)
            obj.entry_type = "PunchIn"
            obj.attendance_date = timezone.now()
            obj.save()
            messages.success(request, 'Logged In successfully',
                             extra_tags='alert alert-success')


        elif punch == "Punch_Out":
            obj = Attendance.objects.create(employee_id=user.id)
            obj.entry_type = "PunchOut"
            obj.attendance_date =timezone.now()
            obj.save()
            messages.success(request, 'Logged Out successfully',
                             extra_tags='alert alert-success')

        elif punch == "Break_In":
            obj = Attendance.objects.create(employee_id=user.id)
            obj.entry_type = "BreakIn"
            obj.attendance_date = timezone.now()
            obj.save()


        elif punch == "Break_Out":
            obj = Attendance.objects.create(employee_id=user.id)
            obj.entry_type = "BreakOut"
            obj.attendance_date = timezone.now()
            obj.save()
        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def Attndence_Records(request):
    if request.method == "GET":
        start_date_str = request.GET.get('from_date')
        end_date_str = request.GET.get('to_date')

        # Check if both start_date and end_date are provided
        if not start_date_str or not end_date_str:
           # messages.error(request,'Please provide both start and end dates.',extra_tags='alert alert-success alert-dismissible fade show')

            return render(request, 'attendence/attendece_records.html',
                          )

        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            return render(request, 'attendence/attendece_records.html',
                          {'error_message': 'Invalid date format. Please use YYYY-MM-DD.'})

        step = timedelta(days=1)
        date_range = []

        current_date = start_date
        while current_date <= end_date:
            date_range.append(current_date)
            current_date += step

        d1 = OrderedDict()

        for d in date_range:
            d_format = d.strftime("%Y-%m-%d")
            d1_format = d.strftime("%Y/%m/%d")
            obj = Attendance.objects.filter(employee_id=request.user.id, attendance_date__contains=d_format).all()
            leave_verify = LeaveRequest.objects.filter(employee_id=request.user.id, status="Approved",
                                                       StartDate__lte=d_format, EndDate__gte=d_format).all()
            if obj:
                d1[(d1_format, "Present")] = obj
            elif d.weekday() in [5, 6]:  # 5 is Saturday, 6 is Sunday
                d1[(d1_format, "WeekOff")] = None
            elif leave_verify:
                    d1[(d1_format, "Leave")] = leave_verify
            else:
                d1[(d1_format, )] = obj


        context = {
            'd1': d1
        }
       # print(d1.values())
        return render(request, 'attendence/attendece_records.html', context)

    return render(request, 'attendence/attendece_records.html')

@csrf_exempt
def leave_verify(request):
    today = timezone.now()
    is_leave_status = LeaveRequest.objects.filter(employee_id=request.user.id,status="Approved", StartDate__lte=today, EndDate__gte=today).exists()

    return JsonResponse({'is_leave_status': is_leave_status,

                         })






def daywise_records(request):
    start_date =datetime.datetime(2023, 7, 7)
    end_date = datetime.datetime(2023,8,7)  # Two months have a maximum of 60 days
    punch_records = Attendance.objects.filter(employee=1, attendance_date__gte=start_date, attendance_date__lt=end_date)
    return render(request,'k5.html',{'punch_records': punch_records})






    #if punch =="Punch_in":
     #    obj.punch_in=timezone.now()
   # elif punch == "Punch_Out":
    #    obj.punch_out= timezone.now()
  #  elif punch == "Break_In":
    #        obj.break_in  = timezone.now()
  #  elif punch == "Break_Out":
       #     obj.break_out  = timezone.now()

       # obj.save()


   # return render(request,'k.html')
                  #'attendence/attendence.html')
                         # 'attendence/attendence.html'
                  #'k.html')








def punch_in_prasent(request):
    if request.method == 'GET':
        date=datetime.date.today()
       # print(date)
        user = CustomUser.objects.get(id=request.user.id)
        is_punch_in_present = Attendance.objects.filter(employee_id=user.id,entry_type = "PunchIn",attendance_date__startswith=date).exists()
        is_punch_out_present = Attendance.objects.filter(employee_id=user.id, entry_type="PunchOut",attendance_date__startswith=date).exists()
        is_break_in_present = Attendance.objects.filter(employee_id=user.id, entry_type="BreakIn",attendance_date__startswith=date).exists()
        is_break_out_present = Attendance.objects.filter(employee_id=user.id, entry_type="BreakOut",attendance_date__startswith=date).exists()
       # print(is_punch_in_present)
        return JsonResponse({'is_punch_in_present': is_punch_in_present,
                             'is_punch_out_present':is_punch_out_present,
                             'is_break_in_present':is_break_in_present,
                             'is_break_out_present': is_break_out_present
                             })







def punch_out(request):
    if request.method == 'POST':
        user = CustomUser.objects.get(id=request.user.id)
        obj = Attendance.objects.create(employee_id=user.id)
        obj.entry_type = "PunchOut"
        obj.attendance_date = timezone.now()
        obj.save()
        return HttpResponse("home")












def Calculation_view(request):
    if request.method == 'POST':

        shift = request.POST.get('shift')
        print(shift)

        # If the selected shift is "custom," retrieve the custom shift value


        # Process the input values as needed
        result = shift
        return HttpResponse(result)
    else:
        # Handle other HTTP methods if needed
        return render(request, 'cal.html')




"""
    def emp_update(request,id):
        if request.method == 'POST':
            username = request.POST['username']
            obj=User.objects.get(username=username)
            print(obj)
            obj1=EmployeeMaster.objects.get(user=obj)
            print(obj1)
            username = request.POST['username']
            firstname = request.POST['firstname']
            lastname = request.POST['lastname']
            password = request.POST['password']
            confirmpassword = request.POST['cpwd']
            Email = request.POST['email']
            dob = request.POST['dob']
            empid = request.POST['empid']
            course = request.POST.get('course')
            department = request.POST.get('department')
            phone = request.POST['phno']
            gender = request.POST.get('gender')
            role = request.POST.get('role')
            shift = request.POST.get('shift')
            category = request.POST.get('category')


        else:
            return render(request,'emp_update.html')

    """


#=============================
#=============================

#HR RELATED VIEWS

def Hr_view(request):
    if request.user.is_authenticated:
        if request.user.is_manager== True:


            today = timezone.now().date()
            announcements = Announcements.objects.filter(start_date__lte=today,end_date__gte=today).order_by('start_date')



            #s = LeaveRequest.objects.filter(status='Pending').count()


            s = CustomUser.objects.filter(is_active=True,is_employee=True,leaverequest__status='Pending').count()



            r = CustomUser.objects.filter(is_active=True,is_employee=True,permission__status=True).count()

            total_permissions = PermissionApproval.objects.filter(permission__StartDate__lte=today,permission__EndDate__gte=today,approval_status="Approved").count()

            #total_permissions = Permission.objects.all().count()

            #r = Permission.objects.filter(status=True).count()
            e = CustomUser.objects.filter(is_active=True,is_employee=True).count()


            total_employee_count = CustomUser.objects.filter(is_employee=True).count()






            active_employee_count = CustomUser.objects.filter(is_active=True,is_employee=True,attendance__entry_type='PunchIn',attendance__attendance_date__contains=today).count()


            # Get today's date
            today = date.today()
            # Define the SQL query with table aliases
            absent_employee_query = CustomUser.objects.exclude(
                Q(attendance__entry_type='PunchIn', attendance__attendance_date__contains=today) |
                (

                        Q(~Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                           Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                          ~Q(leaverequest__StartDate__gt=today, leaverequest__EndDate__gt=today)) &
                        Q(leaverequest__StartDate__lte=today, leaverequest__EndDate__gte=today) &

                        Q(Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                          Q(leaverequest__StartDate__lte=today, leaverequest__EndDate__gte=today)) &
                        ~Q(leaverequest__status="Rejected")
                )
            )
            #print(absent_employee_query)


            # Count the number of employees who are both active and employees and meet the absent criteria
            absent_employees_count = absent_employee_query.filter(
                is_active=True,
                is_employee=True
            ).count()


            interviews = Job_Interview.objects.filter(interview_date__gte=today).count()
            pending_tickets = Tickets.objects.filter(status="Pending").count()
            pending_expenses = Expenses.objects.filter(status="Pending").count()

            employee_leave = CustomUser.objects.filter(is_active=True,is_employee=True,leaverequest__status="Approved",leaverequest__StartDate__lte=today,leaverequest__EndDate__gte=today).count()


           #employee_leave = LeaveRequest.objects.filter(status="Approved", StartDate__lte=today,EndDate__gte=today).count()


           #bdate = EmployeeMaster.objects.values('BirthDate').count()
            birthdaye_emp_mon = CustomUser.objects.filter(is_active=True)
            current_year = today.year



            bdate = EmployeeMaster.objects.annotate(
                birthdate_month=ExtractMonth('BirthDate'),
                 birthdate_day = ExtractDay('BirthDate')

            ).filter(
                Q(user_id__in=[i.id for i in birthdaye_emp_mon]) &

                Q(Q(birthdate_month__gt=today.month) | (
                    Q(birthdate_month=today.month, birthdate_day__gte=today.day))) & (
                        Q(birthdate_month__lte=12) | Q(birthdate_month__gte=today.month, birthdate_month__lte=12)),
            ).order_by('birthdate_month', 'birthdate_day').count()


            # Find employees whose birthdays fall within the current year or the next year from today's date
            department = EmployeeMaster.objects.values('Department').distinct().count()

           #role_count = EmployeeMaster.objects.values('Role').distinct().count()

            today = date.today()

            birthdaye_emp = CustomUser.objects.filter(is_active=True,is_employee=True)

            #Filter EmployeeMaster objects for birthdays based on the month and day, and order them

            birthdates = EmployeeMaster.objects.annotate(
                birthdate_month=ExtractMonth('BirthDate'),
                birthdate_day=ExtractDay('BirthDate')
            ).filter(
                Q(birthdate_month=today.month, birthdate_day__gte=today.day) |
                Q(birthdate_month__gt=today.month)
            ).order_by('birthdate_month', 'birthdate_day')[:3]
            #birthdaye_emp = CustomUser.objects.filter(is_active=True).all()
            #birthdates = EmployeeMaster.objects.filter(user_id__in=[i.id for i in birthdaye_emp]).order_by('BirthDate')[:3]

            #birthdays = EmployeeMaster.objects.filter(BirthDate__gte=today).order_by('BirthDate')[:3]
            holidays=Holidays.objects.filter(date__gte=today).order_by('date')[:5]

            events=Events.objects.filter(event_date__gte=today).order_by('event_date')[:3]

            meetings=Meetings.objects.filter(meeting_date=today).order_by('meeting_date')[:5]


            Total_Tasks = My_Tasks.objects.all().count()
            clients = Client.objects.all().count()
            projects = Projects.objects.all().count()

            return render(request, 'hr/hr_dashboard.html', {'pending_expenses':pending_expenses,'pending_tickets':pending_tickets,'announcements':announcements,'meetings':meetings,'total_employee_count':total_employee_count,'Total_Tasks':Total_Tasks,'absent_employees_count':absent_employees_count,'holidays':holidays,'birthdates': birthdates,'interviews':interviews,'events':events,'employee_leave':employee_leave,'active_employee_count': active_employee_count,'e':e,'s':s,"bdate":bdate,'r':r,'department':department,'total_permissions':total_permissions,'clients':clients,'projects':projects})



def birthdate(request):
    today = date.today()
    current_year = today.year

    # Find employees whose birthdays fall within the current year or the next year from today's date
    birthdays = EmployeeMaster.objects.annotate(
        birthdate_month=ExtractMonth('BirthDate'),
        birthdate_day=ExtractDay('BirthDate')
    ).filter(
        Q(Q(birthdate_month__gt=today.month) | (Q(birthdate_month=today.month, birthdate_day__gte=today.day))) & (Q(birthdate_month__lte=12) | Q(birthdate_month__gte=today.month, birthdate_month__lte=12)),
    ).order_by('birthdate_month', 'birthdate_day')

    return render(request, 'hr/birthdate.html', {'birthdays': birthdays, 'current_year': current_year})






def Employee_leave(request):
    today = timezone.now().date()




    employees_leave = CustomUser.objects.filter(is_active=True, is_employee=True, leaverequest__status="Approved",leaverequest__StartDate__lte=today,leaverequest__EndDate__gte=today).all()

    employees_leave_data = LeaveRequest.objects.all()


    p2 = EmployeeMaster.objects.filter(user_id__in=[i.id for i in employees_leave])







    context ={
        'emp_leave':zip(employees_leave,p2,employees_leave_data)
     }



    return render(request,'hr/employee_leave.html',context)






def present_employee(request):

    today = timezone.now().date()
    present = CustomUser.objects.filter(is_active=True,is_employee=True,attendance__entry_type='PunchIn',attendance__attendance_date__contains=today).all()

    p2 = EmployeeMaster.objects.filter(user_id__in=[i.id for i in present])

    return render(request,'hr/present_employee.html',{'present':present,'p2':p2})







def absent_employee(request):
    today = timezone.now().date()

    absent_employee_query = CustomUser.objects.exclude(
        Q(attendance__entry_type='PunchIn', attendance__attendance_date__contains=today) |
        (

                Q(~Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                  Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                  ~Q(leaverequest__StartDate__gt=today, leaverequest__EndDate__gt=today)) &
                Q(leaverequest__StartDate__lte=today, leaverequest__EndDate__gte=today) &

                Q(Q(leaverequest__StartDate__lt=today, leaverequest__EndDate__lt=today) |
                  Q(leaverequest__StartDate__lte=today, leaverequest__EndDate__gte=today)) &
                ~Q(leaverequest__status="Rejected")
        )
    )
    # print(absent_employee_query)

    # Count the number of employees who are both active and employees and meet the absent criteria
    absent_employees_count = absent_employee_query.filter(
        is_active=True,
        is_employee=True
    ).all()

    p2 = EmployeeMaster.objects.filter(user_id__in=[i.id for i in absent_employees_count])

    return render(request, 'hr/absent_employee.html', {'p2': p2, 'absent_employee_query': absent_employee_query, 'absent_employees_count': absent_employees_count})

def role(request):

    role_counts = EmployeeMaster.objects.values('Role').annotate(count=models.Count('user'))


    labels = [entry['Role'] for entry in role_counts]
    data = [entry['count'] for entry in role_counts]

    return JsonResponse({'labels': labels, 'data': data})




def List_holidays_data(request):
    today = timezone.now().date()

    data45 = Holidays.objects.filter(date__gte=today).order_by('date')

    return render(request, 'hr/holidays/list_holiday.html', {'data45':data45})





def Adding_holidays_import_data(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        imported_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            BANK, IFSC, BRANCH = row
            data_object = Holidays.objects.create(title=BANK, date=IFSC, description=BRANCH)
            imported_data.append(data_object)

        return render(request, 'hr/holidays/add_holidays_import.html', {'imported_data': imported_data})
    return render(request, 'hr/holidays/add_holidays_import.html')






def Adding_holidays_data(request):
    if request.method == 'POST':


        title = request.POST.get('title')
        date = request.POST.get('date')
        description = request.POST.get('description')
        obj = Holidays.objects.create(
                title=title,
                date=date,
                description=description,

            )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_holidays')
    else:
        return render(request, 'hr/holidays/add_holiday.html', )











def update_holidays_data(request,id):
    if request.method == 'POST':


        title_upd = request.POST.get('title')
        date_upd = request.POST.get('date')
        description_upd = request.POST.get('description')

       # print(title_upd,date_upd, description_upd)



        obj = Holidays.objects.get(id=id)


        obj.title = title_upd
        obj.date = date_upd
        obj.description = description_upd


        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_holidays')
    else:

        obj = Holidays.objects.get(id=id)
        return render(request,'hr/holidays/update_holidays.html',{"obj": obj})




def delete_holidays_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Holidays.objects.get(id=id)
        except Holidays.DoesNotExist:
            return JsonResponse({'error': 'Holidays entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_holidays')
    else:
        return render(request, 'hr/holidays/holidays_delete.html', )


def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        if '@' in email:
            try:
                user = CustomUser.objects.get(email=email)
                token_generator = default_token_generator
                token = token_generator.make_token(user)
                PasswordResetView.as_view()(
                    request=request,
                    extra_context={
                       'email': user.email,
                       'token': token,
                     }
                     )

                messages.success(request, 'Password reset email sent. Please check your email.')
                return redirect('password_reset_done')

            except ObjectDoesNotExist:
               messages.error(request, 'Email not found. Please enter a valid email address.')
               return render(request, 'password_reset.html')
        else:
            try:
                user = CustomUser.objects.get(username=email)
                token_generator = default_token_generator
                token = token_generator.make_token(user)
                PasswordResetView.as_view()(
                    request=request,
                    extra_context={
                        'email': user.email,
                        'token': token,
                    }
                )

                messages.success(request, 'Password reset email sent. Please check your email.')
                return redirect('password_reset_done')

            except ObjectDoesNotExist:
                messages.error(request, 'user not found. Please enter a valid username.')
            return render(request, 'password_reset.html')
    else:
        return render(request, 'password_reset.html')




def get_tasks(request):
    if request.method=="GET":
        clients = Client.objects.all()
        employees = CustomUser.objects.all()
        camp = Company.objects.all()
        projects=Projects.objects.all()


        all_tasks_with_employees = Tasks.objects.all()


        context={
            'all_tasks_with_employees': all_tasks_with_employees,
            'clients': clients,
            'employees': employees,
            'camp': camp,
            'projects':projects

        }

        return render(request,'hr/project/tasks.html',context)









label_colors = {}

def get_color(label):
    if label not in label_colors:
        # Generate a unique random color for the label
        color = '#{0:02X}{1:02X}{2:02X}'.format(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        label_colors[label] = color
    return label_colors[label]



def role_counts(request):
    # Count the number of employees in each department
    role_counts = Department.objects.filter(employeemaster__isnull=False,employeemaster__user__is_active=True,employeemaster__user__is_employee=True,).annotate(count=Count('employeemaster'))

    labels = [entry.department for entry in role_counts]
    data = [entry.count for entry in role_counts]
    colors = [get_color(label) for label in labels]  # You should define the 'get_color' function

    return JsonResponse({'labels': labels, 'data': data, 'colors': colors})




def shift_count(request):


    shift_count = EmployeeMaster.objects.filter(user__is_active=True,user__is_employee=True).annotate(
        shift_time_order=Case(
            When(Shift_time__contains='6:00AM-2:00PM', then=Value(1)),
            When(Shift_time__contains='9:30AM-7:00PM', then=Value(2)),
            When(Shift_time__contains='2:00PM-10:00PM', then=Value(3)),
        )
    ).values('Shift_time', 'shift_time_order').annotate(count=Count('user')).order_by('shift_time_order')
    labels1 = [entry['Shift_time'] for entry in shift_count]
    data1 = [entry['count'] for entry in shift_count]

    return JsonResponse({'labels1': labels1, 'data1': data1})





def contacts(request):
    if request.method=="POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        employee_id = request.POST.get('employee_id')

        subject = name,employee_id
        from_email = 'vsriram888@gmail.com'
        recipient_list = ['sharat.mudigeti@gmail.com','sharathmudigeti123@gmail.com']

        send_mail(subject, message, from_email, recipient_list)
        messages.success(request,'Mail send successfully',extra_tags='alert alert-success alert-dismissible fade show')

        return redirect('contacts_data')
    else:
        return render(request, 'hr/contacts.html')



def emp_update_total_view(request,data_id):
    if request.method == 'GET':


        user=CustomUser.objects.get(id=data_id)
        basic=EmployeeMaster.objects.get(user_id=user.id)
        data = {
            'first_name':user.first_name,
            'last_name': user.last_name,
            'username':user.username,
            'email':user.email,
            'phone_no':basic.phone_number,
            'emp_id':basic.EmployeeID,
            'address':basic.Address,
            'city':basic.City,
            'state':basic.state,
            'zip_code':basic.PinCode,
            'country':basic.country ,
        'date_of_birth':basic.BirthDate,
        'gender':basic.Gender,
        'marital_status':basic.MaratialStatus,
        'department':basic.Department_id,
        'designation':basic.Designation,
        'role':basic.Role,
        'shift':basic.Shift_time,
        'joining_date':basic.JoiningDate,

              }
      #  print(data)
        return JsonResponse(data)


def update_emergency_contacts_data(request):


    if request.method == 'POST':
        user = CustomUser.objects.get(id=request.user.id)
        data_id = request.POST.get('data_id')
        contact_id = request.POST.get('contact_id')
        relation_update= request.POST.get('relation_update')
        contact_work_email_update= request.POST.get('contact_work_email_update')
        contact_name_update = request.POST.get('contact_name_update')
        contact_address_update = request.POST.get('contact_address_update')
        mobile_update = request.POST.get('mobile_update')
        contact_city_update = request.POST.get('contact_city_update')
        contact_state_update = request.POST.get('contact_state_update')
        contact_zip_update = request.POST.get('contact_zip')
        country1_update = request.POST.get('country1_update')
       # print(relation_update,contact_id,data_id, contact_work_email_update,country1_update,contact_state_update,contact_zip_update,contact_city_update,contact_name_update, contact_address_update,mobile_update,contact_address_update)
        obj=Emergency_Contacts.objects.get(id=contact_id,user_id=data_id)
        obj.relation = relation_update
        obj.email = contact_work_email_update
        obj.name = contact_name_update
        obj.address = contact_address_update
        obj.mobile_no = mobile_update
        obj.city = contact_city_update
        obj.state = contact_state_update
        obj.zip_code = contact_zip_update
        obj.country = country1_update
        obj.save()

        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def get_emergency_contacts_data(request,id):


    if request.method == 'GET' :

        obj = Emergency_Contacts.objects.get(id=id)

        data = {
        'relation_update':obj.relation,
        'contact_work_email_update':obj.email,
        'contact_name_update':obj.name,
        'contact_address_update':obj.address,
        'mobile_update':obj.mobile_no,
        'contact_city_update':obj.city,
        'contact_state_update':obj.state,
        'contact_zip_update':obj.zip_code,
        'country1_update':obj.country,


        }
       # print(data)
        return JsonResponse(data, safe=False)


def delete_emergency_contacts_data(request):
    if request.method == 'POST':
        data_id = request.POST.get('data_id')
        contact_id = request.POST.get('contact_id')
        obj = Emergency_Contacts.objects.get(id=contact_id, user_id=data_id)
        obj.delete()
        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def Add_profile_data(request):
    if request.method == 'POST':



        fb_id = request.POST.get('fb_id')
        skype_id = request.POST.get('skype_id')
        linkedIn_id = request.POST.get('linkedIn_id')
        twitter_id = request.POST.get('twitter_id')
        whatsapp_id = request.POST.get('whatsapp_id')
        data_id = request.POST.get('data_id')
        user = CustomUser.objects.get(id=data_id)




       # print(fb_id,skype_id,linkedIn_id,twitter_id,whatsapp_id)

        obj=Social_Profile.objects.create(user_id=user.id, facebook_profile=fb_id,linkedin_profile=linkedIn_id,skype_profile=skype_id,twitter_profile=twitter_id, whatsapp_profile=whatsapp_id)


        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)

def update_profile_data(request):
    if request.method == 'POST':

        data_id = request.POST.get('data_id')
        profile_id = request.POST.get('profile_id')

        fb_id = request.POST.get('fb_id_update')
        skype_id = request.POST.get('skype_id_update')
        linkedIn_id = request.POST.get('linkedIn_id_update')
        twitter_id = request.POST.get('twitter_id_update')
        whatsapp_id = request.POST.get('whatsapp_id_update')

     #   print(fb_id, skype_id, linkedIn_id, twitter_id, whatsapp_id,data_id,profile_id)

        obj = Social_Profile.objects.get(user_id=data_id,id=profile_id)
        obj.facebook_profile=fb_id
        obj.linkedin_profile=linkedIn_id
        obj.skype_profile=skype_id
        obj.twitter_profile=twitter_id
        obj.whatsapp_profile=whatsapp_id

        obj.save()
        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def update_documents_data(request):


    if request.method == 'POST' and request.FILES.get('document_document_file_update'):
      #  user = CustomUser.objects.get(id=request.user.id)

        document_document_file = request.FILES['document_document_file_update']
        data_id=request.POST.get('data_id')
        document_id = request.POST.get('document_id')
        document_title = request.POST.get('document_title_update')
        document_expiry_date_update = request.POST.get('document_expiry_date_update')
        document_end_date_update = request.POST.get('document_end_date_update')
        document_description = request.POST.get('document_description_update')

        user = CustomUser.objects.get(id=data_id)




        obj = Documents_All.objects.get(user_id=user.id,id=document_id)
        obj.title=document_title
        obj.expired_date=document_expiry_date_update
        obj.end_date=document_end_date_update
        obj.description=document_description
        obj.document_file=document_document_file

        obj.save()
        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)





def get_all_documents_data(request,id):


    if request.method == 'GET' :

        obj = Documents_All.objects.get(id=id)



        data = {
            'document_title_update':obj.title,
            'document_expiry_date_update':obj.expired_date,
            'document_end_date_update':obj.end_date,
            'document_description_update':obj.description,



        }
       # print(data)
        return JsonResponse(data)


def delete_documents_data(request):


    if request.method == 'POST':
      #  user = CustomUser.objects.get(id=request.user.id)

     #   document_document_file = request.FILES['document_document_file_update']
        data_id=request.POST.get('data_id')
        document_id = request.POST.get('document_id')
      #  print(data_id,document_id)


        obj = Documents_All.objects.get(user_id=data_id, id= document_id)
        obj.delete()


        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def update_qualifications_data(request):


    if request.method == 'POST' and request.FILES.get('qualification_file_update'):
        institution_name = request.POST.get('institution_name_update')
        education_level = request.POST.get('education_level_update')
        qualification_from_date = request.POST.get('qualification_from_date_update')
        qualification_to_date = request.POST.get('qualification_to_date_update')
        language = request.POST.get('language_update')
        professional_skills = request.POST.get('professional_skills_update')
        qualification_file_update = request.FILES['qualification_file_update']

        qualification_description = request.POST.get('qualification_description_update')
        data_id = request.POST.get('data_id')
        qualification_id = request.POST.get('qualification_id')

      #  print(data_id,qualification_id,qualification_file_update,institution_name,education_level,qualification_from_date,qualification_to_date,language,professional_skills,qualification_description)
        user = CustomUser.objects.get(id=data_id)
        obj = Qualifications.objects.get(user_id=data_id,id=qualification_id)
        obj.name_of_university=institution_name
        obj.branch=education_level
        obj.from_date=qualification_from_date
        obj.to_date=qualification_to_date
        obj.languages=language
        obj.skills=professional_skills
        obj.qualification_file=qualification_file_update

        obj.description=qualification_description
        obj.save()
        messages.success(request, 'Qualifications request submitted successfully.',extra_tags='alert alert-success alert-dismissible fade show')


        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)



def delete_qualifications_data(request):


    if request.method == 'POST':

        data_id = request.POST.get('data_id')
        qualification_id = request.POST.get('qualification_id')

       # print(data_id,qualification_id)
        obj = Qualifications.objects.get(user_id=data_id, id=qualification_id)
        obj.delete()
        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)


def get_qualifications_data(request,id):


    if request.method == 'GET' :

        obj = Qualifications.objects.get(id=id)

        data = {
        'institution_name_update':obj.name_of_university,
       'education_level_update':obj.branch,
        'qualification_from_date_update':obj.from_date,
        'qualification_to_date_update':obj.to_date,
        'language_update':obj.languages,
        'professional_skills_update':obj.skills,
        'qualification_description_update':obj.description


        }
        return JsonResponse(data, safe=False)


def update_work_experience_data(request):


    if request.method == 'POST' and request.FILES.get('workex_document_file_update'):
        work_experience_id = request.POST.get('work_experience_id')


        workex_document_file = request.FILES['workex_document_file_update']

        work_company_name = request.POST.get('work_company_name_update')
        work_experience_from_date = request.POST.get('work_experience_from_date_update')
        work_experience_to_date = request.POST.get('work_experience_to_date_update')
        designations = request.POST.get('designations_update')
        work_experience_description = request.POST.get('work_experience_description_update')

      #  print(work_experience_description,work_experience_to_date,work_experience_from_date,workex_document_file,work_company_name,designations)

        obj = Work_Experience.objects.get(id=work_experience_id)
        obj.company=work_company_name
        obj.document_file=workex_document_file
        obj.from_date=work_experience_from_date
        obj.to_date=work_experience_to_date
        obj.description=work_experience_description
        obj.designation=designations
        obj.save()

        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)

def delete_work_experience_data(request):


    if request.method == 'POST':


        work_experience_id = request.POST.get('work_experience_id')
       # print(work_experience_id)

        obj = Work_Experience.objects.get(id=work_experience_id)

        obj.delete()

        response_data = {'first_name': "workingfine"}
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid form data.'}, status=400)

def get_work_experience_data(request,id):


    if request.method == 'GET':
        obj = Work_Experience.objects.get(id=id)
        data={
            'work_company_name_update':obj.company,
            'work_experience_from_date_update':obj.from_date,
            'work_experience_to_date_update':obj.to_date,
            'work_experience_description_update':obj.description,
             'designations_update': obj.designation,


           }
       # print(data)
        return JsonResponse(data, safe=False)


def update_bank_account_data(request):


    if request.method == 'POST':
        if request.method == 'POST':
            user = CustomUser.objects.get(id=request.user.id)

            bank_account_number = request.POST.get('bank_account_number_update')
            bank_bank_name = request.POST.get('bank_bank_name_update')
            bank_bank_code = request.POST.get('bank_bank_code_update')
            bank_bank_branch = request.POST.get('bank_bank_branch_update')
            status = request.POST.get('status_upd')
            data_id = request.POST.get('data_id')
            bank_account_id = request.POST.get('bank_account_id')
            user = CustomUser.objects.get(id=data_id)
            obj = Bank_Account.objects.get(user_id=user.id,id=bank_account_id)
            obj.account_number=bank_account_number
            obj.bankname=bank_bank_name
            obj.ifsc_code=bank_bank_code
            obj.branch=bank_bank_branch
            obj.is_active=status
            obj.save()

            response_data = {'first_name': "workingfine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)




def get_bank_account_data(request,id):


    if request.method == 'GET' :
        obj=Bank_Account.objects.get(id=id)
        data={
        'bank_account_number_update': obj.account_number,
        'bank_bank_name_update':  obj.bankname,
        'bank_bank_code_update': obj.ifsc_code,
        'bank_bank_branch_update': obj.branch,
        'status_upd': obj.is_active,

        }
        return JsonResponse(data, safe=False)

def delete_bank_account_data(request):


    if request.method == 'POST':
        if request.method == 'POST':

            data_id = request.POST.get('data_id')
            bank_account_id = request.POST.get('bank_account_id')

            obj = Bank_Account.objects.get(user_id=data_id,id=bank_account_id)

            obj.delete()

            response_data = {'first_name': "workingfine"}
            return JsonResponse(response_data)
        else:
            return JsonResponse({'error': 'Invalid form data.'}, status=400)




def dashboard(request):
    Total_Tasks = My_Tasks.objects.filter(assigned_employee__id=request.user.id).count()

    my_user = CustomUser.objects.get(id=request.user.id)

    data50 = Status.objects.get(status="Completed")



    completed_count1 = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=data50.id).count()



   # print(completed_count1)
    todo = Status.objects.get(status="To Do")


    Todo_count1 = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=todo.id).count()



    in_progress_count = Status.objects.get(status="In Progress")

    In_progress_count1 = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=in_progress_count.id).count()


    return render(request,'hr/employee_work.html',{'Total_Tasks':Total_Tasks,'completed_count1':completed_count1,'Todo_count1':Todo_count1,'In_progress_count1':In_progress_count1})






def hr_dashboard(request):


    Total_Tasks = My_Tasks.objects.all().count()
    data50 = Status.objects.get(status="Completed")


    completed_count = My_Tasks.objects.filter(status_id=data50.id).count()

    todo = Status.objects.get(status="To Do")



    Todo_count = My_Tasks.objects.filter(status_id=todo.id).count()

    inprogress = Status.objects.get(status="In Progress")

    In_progress_count = My_Tasks.objects.filter(status_id=inprogress.id).count()


    clients_count = Client.objects.count()

    projects_count = Projects.objects.count()

    return render(request,'hr/hr_my_tasks.html',{'completed_count':completed_count,'Todo_count':Todo_count,'In_progress_count':In_progress_count,'Total_Tasks':Total_Tasks,'clients_count':clients_count,'projects_count':projects_count})


def completed_tasks_hr(request):
    if request.method == 'GET':
        status = Status.objects.get(status="Completed")
        task_data=My_Tasks.objects.filter(status_id=status.id)
        context={
            'status':status,
            'task_data':task_data,
        }
        return render(request,'hr/mytasks/list_tasks_hr.html',context)





def my_tasks_completed(request):
    if request.method == "GET":
        status = Status.objects.get(status="Completed")
        employee_tasks = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=status.id)
        print(employee_tasks)
        return render(request,'hr/mytasks/tasks/list_tasks.html',{'status':status,'employee_tasks':employee_tasks})




def my_tasks_todo(request):
    if request.method == "GET":
        status = Status.objects.get(status="To Do")
        employee_tasks = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=status.id)
        print(employee_tasks)
        return render(request,'hr/mytasks/tasks/list_tasks.html',{'status':status,'employee_tasks':employee_tasks})










def in_progress_tasks_hr(request):
    if request.method == 'GET':
        status = Status.objects.get(status="In Progress")
        task_data=My_Tasks.objects.filter(status_id=status.id)
        context={
            'status':status,
            'task_data':task_data,
        }
        return render(request,'hr/mytasks/list_tasks_hr.html',context)





def my_tasks_inprogress(request):
    if request.method == "GET":
        status = Status.objects.get(status="In Progress")
        employee_tasks = My_Tasks.objects.filter(assigned_employee__id=request.user.id,status_id=status.id)
        print(employee_tasks)
        return render(request,'hr/mytasks/tasks/list_tasks.html',{'status':status,'employee_tasks':employee_tasks})


def completed_tasks(request):

    data50 = Status.objects.get(status="Completed")


    data48 = My_Tasks.objects.filter(status_id=data50.id).all()


    return render(request,"hr/Tasks/completed.html",{'data48':data48})







def inprogress_tasks(request):

    data50 = Status.objects.get(status="In Progress")


    data48 = My_Tasks.objects.filter(status_id=data50.id).all()


    return render(request,"hr/Tasks/inprogress.html",{'data48':data48})






def todo_tasks(request):
    data50 = Status.objects.get(status="To Do")


    data48 = My_Tasks.objects.filter(status_id=data50.id).all()


    return render(request,"hr/Tasks/todo.html",{'data48':data48})



def Hr_Attendance_weekwise(request):
    # Fetch the list of companies here

    if request.method == "GET":
        k = []
        d2 = {}
        employees = EmployeeMaster.objects.filter(user__is_employee=True,user__is_active=True).all()
        for em in employees:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=7)

            try:
                start_date = datetime.datetime.strptime(str(start_date), "%Y-%m-%d")
                end_date = datetime.datetime.strptime(str(end_date), "%Y-%m-%d")
            except ValueError:
                return render(request, 'hr/attendance_records.html', {'error_message': 'Invalid date format. Please use YYYY-MM-DD.'})

            step = timedelta(days=1)
            date_range = []

            current_date = start_date
            while current_date <= end_date:
                date_range.append(current_date)
                current_date += step

            d1 = {}

            for d in date_range:
                d_format = d.strftime("%Y-%m-%d")
                d1_format = d.strftime("%Y/%m/%d")
                obj = Attendance.objects.filter(employee_id=em.user_id, attendance_date__contains=d_format).all()
                leave_verify = LeaveRequest.objects.filter(employee_id=em.user_id, status="Approved",
                                                           StartDate__lte=d_format, EndDate__gte=d_format).all()
                if obj:
                    d1[(d1_format, "Present",)] = obj
                elif d.weekday() in [5, 6]:  # 5 is Saturday, 6 is Sunday
                    d1[(d1_format, "WeekOff",)] = None
                elif leave_verify:
                    d1[(d1_format, "Leave",)] = leave_verify
                else:
                    d1[(d1_format, 'absent',)] = obj




            d2[em]=d1

        l4=[]
        l5=[]
        for l,m in d2.items():
            for i,k in m:
                #print(i)
                if i not in l4:
                    l4.append(i)
                else:
                    pass


        return render(request, 'hr/attendance/weekwise.html', {'d2':d2,'l4':l4})

def Hr_Attendance_monthwise(request):
    # Fetch the list of companies here

    if request.method == "GET":
        k = []
        d2 = {}
        employees = EmployeeMaster.objects.filter(user__is_employee=True,user__is_active=True).all()
        for em in employees:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days= 31)

            try:
                start_date = datetime.datetime.strptime(str(start_date), "%Y-%m-%d")
                end_date = datetime.datetime.strptime(str(end_date), "%Y-%m-%d")
            except ValueError:
                return render(request, 'hr/attendance_records.html', {'error_message': 'Invalid date format.Please use YYYY-MM-DD.'})

            step = timedelta(days=1)
            date_range = []

            current_date = start_date
            while current_date <= end_date:
                date_range.append(current_date)
                current_date += step

            d1 = {}

            for d in date_range:
                d_format = d.strftime("%Y-%m-%d")
                d1_format = d.strftime("%Y/%m/%d")
                obj = Attendance.objects.filter(employee_id=em.user_id, attendance_date__contains=d_format).all()
                leave_verify = LeaveRequest.objects.filter(employee_id=em.user_id, status="Approved",
                                                           StartDate__lte=d_format, EndDate__gte=d_format).all()
                if obj:
                    d1[(d1_format, "Present",)] = obj
                elif d.weekday() in [5, 6]:  # 5 is Saturday, 6 is Sunday
                    d1[(d1_format, "WeekOff",)] = None
                elif leave_verify:
                    d1[(d1_format, "Leave",)] = leave_verify
                else:
                    d1[(d1_format, 'absent',)] = obj




            d2[em]=d1

        l4=[]
        l5=[]
        for l,m in d2.items():
            for i,k in m:
                #print(i)
                if i not in l4:
                    l4.append(i)
                else:
                    pass


        return render(request, 'hr/attendance/month_wise.html', {'d2':d2,'l4':l4})
























def Hr_Attendance_Records(request):
    companies = Company.objects.all()

    if request.method == "GET":
        start_date_str = request.GET.get('from_date')
        end_date_str = request.GET.get('to_date')

        company=request.GET.get('company')
        department = request.GET.get('department')
        employee = request.GET.get('employee')


        #print( start_date_str ,end_date_str,company,department,employee)
        # Check if both start_date and end_date are provided
        if not start_date_str or not end_date_str:
            # Consider returning an error message or a different template
            return render(request, 'hr/attendance/attendance_records.html',
                          {'error_message': 'Please provide both start and end dates.', 'companies': companies})

        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            return render(request, 'hr/attendance_records.html',
                          {'error_message': 'Invalid date format. Please use YYYY-MM-DD.', 'companies': companies})

        step = timedelta(days=1)
        date_range = []

        current_date = start_date
        while current_date <= end_date:
            date_range.append(current_date)
            current_date += step

        d1 = OrderedDict()
        d1 = OrderedDict()

        for d in date_range:
            d_format = d.strftime("%Y-%m-%d")
            d1_format = d.strftime("%Y/%m/%d")
            obj = Attendance.objects.filter(employee_id=employee, attendance_date__contains=d_format).all()
            leave_verify = LeaveRequest.objects.filter(employee_id=employee, status="Approved",
                                                       StartDate__lte=d_format, EndDate__gte=d_format).all()
            if obj:
                d1[(d1_format, "Present")] = obj
            elif d.weekday() in [5, 6]:  # 5 is Saturday, 6 is Sunday
                d1[(d1_format, "WeekOff")] = None
            elif leave_verify:
                d1[(d1_format, "Leave")] = leave_verify
            else:
                d1[(d1_format,)] = obj

        return render(request, 'hr/attendance/attendance_records.html', {'d1': d1, 'companies': companies})

def Attndence_Records(request):
    if request.method == "GET":
        start_date_str = request.GET.get('from_date')
        end_date_str = request.GET.get('to_date')

        # Check if both start_date and end_date are provided
        if not start_date_str or not end_date_str:
           # messages.error(request,'Please provide both start and end dates.',extra_tags='alert alert-success alert-dismissible fade show')

            return render(request, 'attendence/attendece_records.html',
                          )

        try:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            return render(request, 'attendence/attendece_records.html',
                          {'error_message': 'Invalid date format. Please use YYYY-MM-DD.'})

        step = timedelta(days=1)
        date_range = []

        current_date = start_date
        while current_date <= end_date:
            date_range.append(current_date)
            current_date += step

        d1 = OrderedDict()

        for d in date_range:
            d_format = d.strftime("%Y-%m-%d")
            d1_format = d.strftime("%Y/%m/%d")
            obj = Attendance.objects.filter(employee_id=request.user.id, attendance_date__contains=d_format).all()
            leave_verify = LeaveRequest.objects.filter(employee_id=request.user.id, status="Approved",
                                                       StartDate__lte=d_format, EndDate__gte=d_format).all()
            if obj:
                d1[(d1_format, "Present")] = obj
            elif d.weekday() in [5, 6]:  # 5 is Saturday, 6 is Sunday
                d1[(d1_format, "WeekOff")] = None
            elif leave_verify:
                    d1[(d1_format, "Leave")] = leave_verify
            else:
                d1[(d1_format, )] = obj


        context = {
            'd1': d1
        }
        #print(d1.values())
        return render(request, 'attendence/attendece_records.html', context)

    return render(request, 'attendence/attendece_records.html')



def get_departments(request, company_id):
    departments = Department.objects.filter(name_id=company_id).values_list('id', 'department')
    return JsonResponse(list(departments), safe=False)

def get_employees(request, department_id):
    employees =EmployeeMaster.objects.filter(Department_id=department_id).values_list('user_id', 'fullname')
    return JsonResponse(list(employees), safe=False)






def List_source_data(request):
    data49 = Source.objects.all()
    return render(request, 'corehr/Task_Customization/Source/list_source.html', {'data49':data49})



def Adding_source_data(request):
    if request.method == 'POST':

        source_task = request.POST.get('source')

        obj = Source.objects.create(
            source=source_task,

        )



        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_source')
    else:

        return render(request, 'corehr/Task_Customization/Source/add_source.html')


def update_source_data(request,id):
    if request.method == 'POST':
        source_task_upd = request.POST.get('source')

        obj = Source.objects.get(id=id)
        obj.source = source_task_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_source')
    else:

        source = Source.objects.get(id=id)

        return render(request, 'corehr/Task_Customization/Source/update_source.html',{'source':source})




def delete_source_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Source.objects.get(id=id)
        except Source.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_source')
    else:
        return render(request, 'corehr/Task_Customization/Source/delete_source.html')




def List_status_data(request):
    data50 = Status.objects.all()
    return render(request, 'corehr/Task_Customization/Status/list_status.html', {'data50':data50})



def Adding_status_data(request):
    if request.method == 'POST':
        status_task = request.POST.get('status')

        obj = Status.objects.create(
            status=status_task,

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_status')
    else:

        return render(request, 'corehr/Task_Customization/Status/add_status.html')


def update_status_data(request,id):
    if request.method == 'POST':
        status_task_upd = request.POST.get('status')

        obj = Status.objects.get(id=id)
        obj.status = status_task_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_status')
    else:

        status = Status.objects.get(id=id)

        return render(request, 'corehr/Task_Customization/Status/update_status.html',{'status':status})




def delete_status_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Status.objects.get(id=id)
        except Status.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_status')
    else:
        return render(request, 'corehr/Task_Customization/Status/delete_status.html')






def List_priority_data(request):
    data51 = Priority.objects.all()
    return render(request, 'corehr/Task_Customization/Priority/list_priority.html', {'data51':data51})



def Adding_priority_data(request):
    if request.method == 'POST':

        priority_task = request.POST.get('priority')

        obj = Priority.objects.create(
            priority=priority_task,

        )



        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_priority')
    else:

        return render(request, 'corehr/Task_Customization/Priority/add_priority.html')


def update_priority_data(request,id):
    if request.method == 'POST':
        priority_task_upd = request.POST.get('priority')

        obj = Priority.objects.get(id=id)
        obj.priority = priority_task_upd

        obj.save()


        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_priority')
    else:

        priority = Priority.objects.get(id=id)

        return render(request, 'corehr/Task_Customization/Priority/update_priority.html',{'priority':priority})




def delete_priority_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Priority.objects.get(id=id)
        except Priority.DoesNotExist:
            return JsonResponse({'error': 'Department entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_priority')
    else:
        return render(request, 'corehr/Task_Customization/Priority/delete_priority.html')



def Add_projects(request):
    if request.method=='POST':
        project_title = request.POST.get('project_title')
        client = request.POST.get('client')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        company = request.POST.get('company')
        priority = request.POST.get('Priority')
        employees = request.POST.getlist('employees[]')
        summary = request.POST.get('summary')
        progress = request.POST.get('progress')
        #print(project_title,client,start_date,end_date,company,priority,employees,summary,progress)
        obj=Projects.objects.create(title=project_title,client_id=client,company_id=company,start_date=start_date,end_date=end_date,priority=priority,summary=summary,Progress=progress)

        for user_id in employees:
            user_instance = CustomUser.objects.get(pk=user_id)
            obj.Assigned_employees.add(user_instance)
        obj.save()
        messages.success(request, 'Project Data Is Saved Successfully.', extra_tags='alert alert-success')
        return redirect('list_project')

    else:
        clients=Client.objects.all()
        employees=CustomUser.objects.all()
        camp = Company.objects.all()
        all_projects_with_employees = Projects.objects.all().prefetch_related('Assigned_employees')


        context={
            'all_projects_with_employees': all_projects_with_employees,
            'clients': clients,
            'employees': employees,
            'camp': camp,
        }

        return render(request,'hr/project/add_project.html',context)

def Employee_project(request, project_id):
    project = get_object_or_404(Projects, pk=project_id)
    assigned_employees = project.Assigned_employees.all()

    # Fetch additional information for each assigned employee
    assigned_employees_with_info = []
    for employee in assigned_employees:
        employee_master = EmployeeMaster.objects.get(user_id=employee.id)
        #department = employee_master.Department  # Assuming there's a ForeignKey 'department' in EmployeeMaster

       # phone_number = employee_master.phone_number  # Assuming 'phone_number' is the field in EmployeeMaster
        assigned_employees_with_info.append(employee_master)

    return render(request, 'hr/project/project_employees.html', {'project': project, 'assigned_employees_with_info': assigned_employees_with_info})



def Update_project(request,id):
    if request.method=='POST':

        project_title = request.POST.get('project_title_upd')
        client = request.POST.get('client_upd')
        start_date = request.POST.get('start_date_upd')
        end_date = request.POST.get('end_date_upd')
        company = request.POST.get('company_upd')
        priority = request.POST.get('Priority_upd')
        employee =request.POST.getlist('employees_upd[]')
        summary = request.POST.get('summary_upd')
        progress = request.POST.get('progress_upd')


       # print(employee,progress,summary,priority,company,end_date,start_date,client,project_title)

        obj=Projects.objects.get(id=id)
        obj.title=project_title
        obj.client_id=client
        obj.company_id=company
        obj.start_date=start_date
        obj.end_date=end_date
        obj.priority=priority
        obj.summary=summary
        obj.Progress=progress
        obj.Assigned_employees.clear()




        for user_id in employee:
            user_instance = CustomUser.objects.get(pk=user_id)
            obj.Assigned_employees.add(user_instance)
            obj.save()
        messages.success(request, 'Project Data Is Updated Succesfully.', extra_tags='alert alert-success')
        return redirect('list_project')

    else:
        projectdata = Projects.objects.get(id=id)
        emps_list=CustomUser.objects.filter(is_employee=True,is_active=True).all()
        company_data=Company.objects.all()
        clients_data=Client.objects.all()


        context={
            'projectdata':projectdata,
            'emps_list':emps_list,
            'clients_data':clients_data,
            'company_data':company_data,
            'assign_emp':[i.id for i in projectdata.Assigned_employees.all()]




        }


        return render(request,'hr/project/update_project.html',context)


def List_project_data(request):
    if request.method == 'GET':
        project_data = Projects.objects.all()
        context={
            'project_data':project_data
        }
        return render(request,'hr/project/list_project.html',context)


def delete_project(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Projects.objects.get(id=id)
        except Projects.DoesNotExist:
            return JsonResponse({'error': 'Project entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_project')
    else:
        return render(request, 'hr/project/delete_project.html')



def get_project_details(request):
    if request.method=="GET":
        clients=Client.objects.all()
        employees=CustomUser.objects.all()
        camp = Company.objects.all()
        all_projects_with_employees = Projects.objects.all().prefetch_related('Assigned_employees')


        context={
            'all_projects_with_employees': all_projects_with_employees,
            'clients': clients,
            'employees': employees,
            'camp': camp,
        }

        return render(request,'hr/project/add_project.html',context)




def tasks(request):
    if request.method == "GET":
        employee_tasks = My_Tasks.objects.filter(assigned_employee__id=request.user.id)
        #print(employee_tasks)
        return render(request,'hr/mytasks/tasks/list_tasks.html',{'employee_tasks':employee_tasks})



def add_my_tasks(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        project = request.POST.get('project')
        assigned_employees = request.POST.get('assign_to')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        source = request.POST.get('source')
        status = request.POST.get('status')
        priority = request.POST.get('priority')
        description = request.POST.get('description')
        get_user_id = request.user.id
        user_id = CustomUser.objects.get(id=get_user_id)
        emp_fullname = EmployeeMaster.objects.get(user_id=user_id.id)
        date = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

        description_msg= f"\n{emp_fullname.fullname} ({date}): {description}"


        #print(title,project,assigned_employees,start_date,end_date,source,status,priority,description)



        #print(type(project))

        obj = My_Tasks.objects.create(

                title=title,
                project_id =project ,
                start_date=start_date,
                end_date=end_date,
                source_id=source,
                status_id=status,
                priority_id=priority,
                comments=description_msg

            )

        user_instance = CustomUser.objects.get(username=assigned_employees)
        obj.assigned_employee.add(user_instance)
        obj.save()
        messages.success(request, 'Task Data Is Saved Successfully.', extra_tags='alert alert-success')
        return redirect('task_list')
    else:
        my_user=CustomUser.objects.get(id=request.user.id)

        source_data = Source.objects.all()
        project_data = Projects.objects.all()
        client_data=Client.objects.all()
        status_data = Status.objects.all()
        priority_data = Priority.objects.all()
        mytask_data = My_Tasks.objects.filter()
        return render(request, "hr/mytasks/tasks/add_tasks.html", {'client_data':client_data,'source_data':source_data,'project_data':project_data,'status_data':status_data,'priority_data':priority_data,'my_user':my_user})




def update_my_tasks_data(request,id):
    if request.method == 'POST':


        status_upd = request.POST.get('status')
        progress_upd = request.POST.get('progress')
        description_upd = request.POST.get('description')


        #print(status_upd,progress_upd,description_upd)
        get_user_id=request.user.id
        user_id=CustomUser.objects.get(id=get_user_id)
        emp_fullname=EmployeeMaster.objects.get(user_id=user_id.id)
        #print(emp_fullname.fullname)


        past_coment = My_Tasks.objects.get(id=id)
        date = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")





        if past_coment:
            #new_message = past_coment.comments  + "\n" +f"\n{emp_fullname.fullname} : {description_upd}"
            #new_message = past_coment.comments + f" {emp_fullname.fullname}: {description_upd}"
            new_message = past_coment.comments + f"\n{emp_fullname.fullname} ({date}) : {description_upd}"


            #new_message = past_coment.comments + "\n" + f"\n\033[1m{emp_fullname.fullname}:\033[0m {description_upd}"

        #print(new_message)
        obj = My_Tasks.objects.get(id=id)
        obj.status_id = status_upd
        obj.comments = new_message
        obj.progress=progress_upd
        obj.save()
        messages.success(request, 'Task Data Is Updated Succesfully.', extra_tags='alert alert-success')
        return redirect('task_list')
    else:
        tasks_data = My_Tasks.objects.get(id=id)
        assigned_employees = tasks_data.assigned_employee.all()
        employees = [employee.username for employee in assigned_employees]

        my_user = CustomUser.objects.get(id=request.user.id)

        source_data = Source.objects.all()
        project_data = Projects.objects.all()
        client_data = Client.objects.all()
        status_data = Status.objects.all()
        priority_data = Priority.objects.all()
        context={
            'tasks_data':tasks_data,
            'my_user':my_user,
            'project_data':project_data,
            'source_data':source_data,
            'client_data':client_data,
            'status_data':status_data,
            'priority_data':priority_data,
            'employees':employees,


        }
        return render(request,'hr/mytasks/tasks/update_tasks.html',context)



def get_my_tasks_data(request, id):
    if request.method == 'GET':
        obj = My_Tasks.objects.get(id=id)

        data = {
            'title_upd': obj.title,
            'project_upd': obj.project,
            'assign_to_upd': obj.assign_to,
            'start_date_upd': obj.start_date,
            'end_date_upd': obj.end_date,
            'source_upd': obj.source_id,
            'status_upd': obj.status_id,
            'priority_upd': obj.priority_id,
            'description_upd': obj.description,
            'comments_upd': obj.comments

        }
        #print(data)
        return JsonResponse(data)

def leave_data(request):
    if request.user.is_authenticated:
        user = EmployeeMaster.objects.all()
        leave_requests = LeaveRequest.objects.all()
        context = {
            'leave_requests': leave_requests,
            'user':user
        }
        return render(request, 'hr/leave_data.html', context)



def employee_details(request):
    if request.user.is_authenticated:
        today = timezone.now().date()

        user = EmployeeMaster.objects.all()
        leave_management = LeaveRequest.objects.all().order_by('StartDate')

        context = {
            'leave_management': leave_management,
            'user':user
        }
        return render(request, 'hr/leave_management.html', context)

from django.db.models import Max

def employee_leaves(request):
    if request.user.is_authenticated:
        #employee_leave = LeaveRequest.objects.all()

        employee_leave =  LeaveRequest.objects.select_related('employee__employeemaster').values('employee_id').annotate(total_leave_requests=Max('id')).all()
        leave_list=[]
        for d in employee_leave:
            leave_total_data=LeaveRequest.objects.filter(id=d['total_leave_requests']).all()
            leave_list.append(leave_total_data)




        context = {
            'employee_leave':leave_list,
            'total_leaves_count_remaining':24.0,
            'casual_leaves_count_remaining':12.0,
            'medical_leaves_count_remaining':12.0,

        }
        return render(request,'hr/employee_leaves.html',context)




def Add_tasks_hr(request):
    if request.method == 'POST':

        title = request.POST.get('title')
        projects = request.POST.get('project')
        assigned_employees = request.POST.getlist('employees[]')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        source = request.POST.get('source')
        status = request.POST.get('status')
        priority = request.POST.get('priority')
        description = request.POST.get('description')
        get_user_id = request.user.id
        user_id = CustomUser.objects.get(id=get_user_id)
        emp_fullname = EmployeeMaster.objects.get(user_id=user_id.id)
        date = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

        description_msg = f"\n{emp_fullname.fullname} ({date}) : {description}"
        #print(title, projects, assigned_employees, start_date, end_date, source, status, priority, description)

        #print(type(projects))

        obj = My_Tasks.objects.create(

            title=title,
            project_id=projects,
            start_date=start_date,
            end_date=end_date,
            source_id=source,
            status_id=status,
            priority_id=priority,
            comments=description_msg

        )
        for user_id in assigned_employees:
            user_instance = CustomUser.objects.get(pk=user_id)
            obj.assigned_employee.add(user_instance)
            obj.save()
        messages.success(request, 'Task Is Saved Succesfully.', extra_tags='alert alert-success')
        return redirect('list_tasks_hr')

    else:
        my_user=CustomUser.objects.get(id=request.user.id)

        source_data = Source.objects.all()
        status_data = Status.objects.all()
        priority_data = Priority.objects.all()
        projects_data = Projects.objects.all()

        users_data = CustomUser.objects.all()
        client_data = Client.objects.all()


        tasks_data = My_Tasks.objects.all()





        return render(request, "hr/mytasks/add_tasks_hr.html", {'priority_data':priority_data,'source_data':source_data,'status_data':status_data,'projects_data':projects_data,'users_data':users_data,'tasks_data':tasks_data,'my_user':my_user,'client_data':client_data})
    return JsonResponse({'error': 'Invalid form data.'}, status=400)





def update_tasks_hr_data(request,id):
    if request.method == 'POST':

        status_upd = request.POST.get('status')
        description_upd = request.POST.get('description')


        #print(status_upd,description_upd)
        get_user_id=request.user.id
        user_id=CustomUser.objects.get(id=get_user_id)
        emp_fullname=EmployeeMaster.objects.get(user_id=user_id.id)
        #print(emp_fullname.fullname)


        past_coment = My_Tasks.objects.get(id=id)
        date = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

        if past_coment:
            new_message = past_coment.comments + f"\n{emp_fullname}({date}): {description_upd}"
            #new_message = past_comment.comments + f"\n{emp_fullname.fullname} : {description_upd}"
        #print(new_message)
        obj = My_Tasks.objects.get(id=id)
        obj.status_id = status_upd
        obj.comments = new_message
        obj.save()
        messages.success(request, 'Task Is updated Succesfully.', extra_tags='alert alert-success')
        return redirect('list_tasks_hr')

    else:
        my_user = CustomUser.objects.get(id=request.user.id)

        source_data = Source.objects.all()
        status_data = Status.objects.all()
        priority_data = Priority.objects.all()
        projects_data = Projects.objects.all()

        users_data = CustomUser.objects.all()
        client_data = Client.objects.all()
        tasks_data=My_Tasks.objects.get(id=id)
        assigned_employees = tasks_data.assigned_employee.all()
        employees = [employee.username for employee in assigned_employees]
        context={
            'my_user':my_user,
            'source_data':source_data,
            'status_data':status_data,
            'priority_data':priority_data,
            'projects_data':projects_data,
            'users_data':users_data,
            'client_data':client_data,
            'tasks_data':tasks_data,
            'employees':employees

        }
        return render(request, 'hr/mytasks/update_tasks_hr.html', context)



def List_tasks_hr_data(request):
    if request.method == 'GET':
        task_data=My_Tasks.objects.all()
        context={
            'task_data':task_data,
        }
        return render(request,'hr/mytasks/list_tasks_hr.html',context)




def todo_tasks_hr(request):
    if request.method == 'GET':
        status = Status.objects.get(status="To Do")
        task_data=My_Tasks.objects.filter(status_id=status.id)
        context={
            'status':status,
            'task_data':task_data,
        }
        return render(request,'hr/mytasks/list_tasks_hr.html',context)






def get_employees_for_project(request):
    if request.method == 'GET':
        project_id = request.GET.get('project')
        #print(project_id)

        if project_id:
            #if project_id:
            try:
                project = Projects.objects.get(id=project_id)
                employees = project.Assigned_employees.all()  # Access the many-to-many relationship
                #print(employees)
                employee_data = [{'id': employee.id, 'username': employee.username} for employee in employees]

                return JsonResponse(employee_data, safe=False)
            except Projects.DoesNotExist:
                # Handle cases where the project doesn't exist
                return JsonResponse([], safe=False)

    return JsonResponse([], safe=False)






def Compose(request):
    if request.method == 'POST':
        rcl = request.POST.getlist('recipients[]')

        # recipients = CustomUser.objects.filter(id__in=request.POST.getlist('recipients'))
        subject = request.POST['subject_title']
        message_text = request.POST['message']
        #print(rcl, subject, message_text)

        # Create a message and save it to the database
        message = Email_inbox.objects.create(subject=subject, message=message_text)

        user_instance = CustomUser.objects.get(pk=request.user.id)
        message.sender.add(user_instance)
        for user_id in rcl:
            user_instance = CustomUser.objects.get(pk=user_id)
            message.recipient.add(user_instance)
        message.save()
        messages.success(request, 'Mail send successfully',
                         extra_tags='alert alert-success alert-dismissible fade show')

        return redirect('sent_emails')





    else:
        user_list=CustomUser.objects.all()
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {

            'user_list': user_list,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }
        return render(request, 'hr/email/compose.html', context)





def Email_details(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies=original_email.replies.all().order_by('-timestamp')[:1:]
    context = {
         'replies':replies,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/email/email_details.html',context)

def Sent_email_details(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies = original_email.replies.all()
    forwards = original_email.forwords.all()
    context = {
         'replies':replies,
        'forwards': forwards,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/email/sent_emil_detalis.html',context)




def Receive_email_details(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies = original_email.replies.all().order_by('-timestamp')
    forwards = original_email.forwords.all()
    context = {
         'replies':replies,
        'forwards':forwards,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/email/receive_email_details.html',context)





def Forword_email(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    if request.method == "POST":

        sender_user = request.user
        recipents = request.POST.getlist('recipients[]')
        new_message = request.POST.get('new_message')
        if new_message:
            # Create a new reply object
            reply = EmailForword.objects.create(
                email=original_email,
                sender=request.user,
                message=new_message,
                timestamp=timezone.now()
            )

            # Update the timestamp of the original email to indicate the latest reply

            user_instance = CustomUser.objects.get(pk=request.user.id)
            original_email.sender.add(user_instance)
            for i in recipents:
                user_instance = CustomUser.objects.get(pk=i)
                original_email.recipient.add(user_instance)

            original_email.timestamp = timezone.now()
            original_email.status = True
            original_email.save()

            return redirect('receive_emails')

    else:
        replies = original_email.replies.all()
        forwards=original_email.forwords.all()
        user_list = CustomUser.objects.all()
        detail_email = Email_inbox.objects.get(id=id)
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {
            'replies': replies,
            'forwards':forwards,
            'user_list': user_list,
            'detail_email': detail_email,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }

        return render(request,'hr/email/forward.html',context)


















def delete_sent_Email(request,id):
    detail_email = Email_inbox.objects.get(id=id)
    if request.method =='POST':
        detail_email.status= False
        detail_email.save()
        return redirect('sent_emails')




    return render(request,'hr/email/delete_sent_email.html',)

def delete_receive_Email(request,id):
    detail_email = Email_inbox.objects.get(id=id)
    if request.method =='POST':
        detail_email.status= False
        detail_email.save()
        return redirect('receive_emails')




    return render(request,'hr/email/delete_receive_email.html',)


def Email_delete(request):
    delete_emails = Email_inbox.objects.filter(status=False)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    context = {

        'delete_emails': delete_emails,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/email/delete_email.html',context)



def Replay_emails(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    if request.method=="POST":

        sender_user=request.user
        new_message = request.POST.get('new_message')
        if new_message:
            # Create a new reply object
            reply = EmailReply.objects.create(
                original_email=original_email,
                sender=request.user,
                message=new_message,
                timestamp=timezone.now()
            )

            # Update the timestamp of the original email to indicate the latest reply

            user_instance = CustomUser.objects.get(pk=request.user.id)
            original_email.sender.add(user_instance)
            for i in original_email.sender.all():
                user_instance = CustomUser.objects.get(pk=i.id)
                original_email.recipient.add(user_instance)



            original_email.timestamp = timezone.now()
            original_email.status=True
            original_email.save()

            return redirect('receive_emails')



    else:
        replies = original_email.replies.all()
        user_list= CustomUser.objects.all()
        detail_email = Email_inbox.objects.get(id=id)
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {
            'replies':replies,
            'user_list' : user_list,
            'detail_email': detail_email,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }

        return render(request, 'hr/email/replay.html', context)



def Sent_Email(request):
    latest_10_emails =  Email_inbox.objects.filter(sender=request.user.id,status=True).all().order_by('-timestamp')
    #print(latest_10_emails)
    # Create a list to store email details
    email_details = []

    for email in latest_10_emails:
        # Retrieve recipient details for each email
        recipient_details = email.recipient.all()

        email_data = {
            'id':email.id,
            'sender':email.sender,
            'subject': email.subject,
            'message': email.message,
            'recipients': [recipient for recipient in recipient_details],
            'timestamp':email.timestamp
        }

        email_details.append(email_data)
    emails_per_page = 3

        # Create a Paginator instance
    paginator = Paginator(email_details, emails_per_page)

        # Get the current page number from the request's GET parameters
    page_number = request.GET.get('page')

        # Get the Page object for the current page
    emails_page = paginator.get_page(page_number)

    k = CustomUser.objects.get(id=request.user.id)
    users = CustomUser.objects.all()
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id,status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count=Email_inbox.objects.filter(sender=request.user.id,recipient=request.user.id, status=False).count()
    return render(request, 'hr/email/email.html', {'users': users, 'k': k, "send_mail": send_mail,'email_details':emails_page,'sent_emails_count':sent_emails_count,'received_emails_count':received_emails_count,'delete_emails_count':delete_emails_count})









def Receive_Email(request):
    received_emails = Email_inbox.objects.filter(recipient=request.user.id,status=True).all().order_by('-timestamp')


    emails_per_page = 3
    paginator = Paginator(received_emails, emails_per_page)
    page_number = request.GET.get('page')

    # Get the Page object for the current page
    emails_page = paginator.get_page(page_number)
    # Create a list to store email details

    k = CustomUser.objects.get(id=request.user.id)
    users = CustomUser.objects.all()
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    context={

        'received_emails':emails_page,
        'sent_emails_count':sent_emails_count,
        'received_emails_count':received_emails_count,
        'delete_emails_count':delete_emails_count,
    }
    return render(request, 'hr/email/receive.html',context)









def Compose_employee(request):
    if request.method == 'POST':
        rcl = request.POST.getlist('recipients[]')
        # recipients = CustomUser.objects.filter(id__in=request.POST.getlist('recipients'))
        subject = request.POST['subject_title']
        message_text = request.POST['message']
        #print(rcl, subject, message_text)

        # Create a message and save it to the database
        message = Email_inbox.objects.create(subject=subject, message=message_text)

        user_instance = CustomUser.objects.get(pk=request.user.id)
        message.sender.add(user_instance)
        for user_id in rcl:
            user_instance = CustomUser.objects.get(pk=user_id)
            message.recipient.add(user_instance)
            message.save()
        messages.success(request, 'Mail send successfully',
                         extra_tags='alert alert-success alert-dismissible fade show')

        return redirect('sent_emails_employee')





    else:
        user_list=CustomUser.objects.all()
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {

            'user_list': user_list,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }
        return render(request, 'hr/emails/compose.html', context)



def Email_details_employee(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies=original_email.replies.all().order_by('-timestamp')[:1:]
    context = {
         'replies':replies,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/emails/email_details.html',context)

def Sent_email_details_employee(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies = original_email.replies.all()
    forwards = original_email.forwords.all()
    context = {
         'replies':replies,
        'forwards': forwards,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/emails/sent_emil_detalis.html',context)




def Receive_email_details_employee(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    replies = original_email.replies.all().order_by('-timestamp')
    forwards = original_email.forwords.all()
    context = {
         'replies':replies,
        'forwards':forwards,
        'detail_email': original_email,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/emails/receive_email_details.html',context)





def Forword_email_employee(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    if request.method == "POST":

        sender_user = request.user
        recipents = request.POST.getlist('recipients[]')
        new_message = request.POST.get('new_message')
        if new_message:
            # Create a new reply object
            reply = EmailForword.objects.create(
                email=original_email,
                sender=request.user,
                message=new_message,
                timestamp=timezone.now()
            )

            # Update the timestamp of the original email to indicate the latest reply

            user_instance = CustomUser.objects.get(pk=request.user.id)
            original_email.sender.add(user_instance)
            for i in recipents:
                user_instance = CustomUser.objects.get(pk=i)
                original_email.recipient.add(user_instance)

            original_email.timestamp = timezone.now()
            original_email.status = True
            original_email.save()

            return redirect('receive_emails_employee')

    else:
        replies = original_email.replies.all()
        forwards=original_email.forwords.all()
        user_list = CustomUser.objects.all()
        detail_email = Email_inbox.objects.get(id=id)
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {
            'replies': replies,
            'forwards':forwards,
            'user_list': user_list,
            'detail_email': detail_email,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }

        return render(request,'hr/emails/forward.html',context)







def delete_sent_Email_employee(request,id):
    detail_email = Email_inbox.objects.get(id=id)
    if request.method =='POST':
        detail_email.status= False
        detail_email.save()
        return redirect('sent_emails_employee')




    return render(request,'hr/emails/delete_sent_email.html',)

def delete_receive_Email_employee(request,id):
    detail_email = Email_inbox.objects.get(id=id)
    if request.method =='POST':
        detail_email.status= False
        detail_email.save()
        return redirect('receive_emails_employee')




    return render(request,'hr/emails/delete_receive_email.html',)


def Email_delete_employee(request):
    delete_emails = Email_inbox.objects.filter(status=False)
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    context = {

        'delete_emails': delete_emails,
        'sent_emails_count': sent_emails_count,
        'received_emails_count': received_emails_count,
        'delete_emails_count': delete_emails_count,
    }

    return render(request,'hr/emails/delete_email.html',context)



def Replay_emails_employee(request,id):
    original_email = get_object_or_404(Email_inbox, id=id)
    if request.method=="POST":

        sender_user=request.user
        new_message = request.POST.get('new_message')
        if new_message:
            # Create a new reply object
            reply = EmailReply.objects.create(
                original_email=original_email,
                sender=request.user,
                message=new_message,
                timestamp=timezone.now()
            )

            # Update the timestamp of the original email to indicate the latest reply

            user_instance = CustomUser.objects.get(pk=request.user.id)
            original_email.sender.add(user_instance)
            for i in original_email.sender.all():
                user_instance = CustomUser.objects.get(pk=i.id)
                original_email.recipient.add(user_instance)



            original_email.timestamp = timezone.now()
            original_email.status=True
            original_email.save()

            return redirect('receive_emails_employee')



    else:
        replies = original_email.replies.all()
        user_list= CustomUser.objects.all()
        detail_email = Email_inbox.objects.get(id=id)
        sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
        received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
        delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                         status=False).count()
        context = {
            'replies':replies,
            'user_list' : user_list,
            'detail_email': detail_email,
            'sent_emails_count': sent_emails_count,
            'received_emails_count': received_emails_count,
            'delete_emails_count': delete_emails_count,
        }

        return render(request, 'hr/emails/replay.html', context)



def Sent_Email_employee(request):
    latest_10_emails =  Email_inbox.objects.filter(sender=request.user.id,status=True).all().order_by('-timestamp')
    #print(latest_10_emails)
    # Create a list to store email details
    email_details = []

    for email in latest_10_emails:
        # Retrieve recipient details for each email
        recipient_details = email.recipient.all()

        email_data = {
            'id':email.id,
            'sender':email.sender,
            'subject': email.subject,
            'message': email.message,
            'recipients': [recipient for recipient in recipient_details],
            'timestamp':email.timestamp
        }

        email_details.append(email_data)
    emails_per_page = 3

        # Create a Paginator instance
    paginator = Paginator(email_details, emails_per_page)

        # Get the current page number from the request's GET parameters
    page_number = request.GET.get('page')

        # Get the Page object for the current page
    emails_page = paginator.get_page(page_number)

    k = CustomUser.objects.get(id=request.user.id)
    users = CustomUser.objects.all()
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id,status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count=Email_inbox.objects.filter(sender=request.user.id,recipient=request.user.id, status=False).count()
    return render(request, 'hr/emails/email.html', {'users': users, 'k': k, "send_mail": send_mail,'email_details':emails_page,'sent_emails_count':sent_emails_count,'received_emails_count':received_emails_count,'delete_emails_count':delete_emails_count})









def Receive_Email_employee(request):
    received_emails = Email_inbox.objects.filter(recipient=request.user.id,status=True).all().order_by('-timestamp')


    emails_per_page = 3
    paginator = Paginator(received_emails, emails_per_page)
    page_number = request.GET.get('page')

    # Get the Page object for the current page
    emails_page = paginator.get_page(page_number)
    # Create a list to store email details

    k = CustomUser.objects.get(id=request.user.id)
    users = CustomUser.objects.all()
    sent_emails_count = Email_inbox.objects.filter(sender=request.user.id, status=True).count()
    received_emails_count = Email_inbox.objects.filter(recipient=request.user.id, status=True).count()
    delete_emails_count = Email_inbox.objects.filter(sender=request.user.id, recipient=request.user.id,
                                                     status=False).count()
    context={

        'received_emails':emails_page,
        'sent_emails_count':sent_emails_count,
        'received_emails_count':received_emails_count,
        'delete_emails_count':delete_emails_count,
    }
    return render(request, 'hr/emails/receive.html',context)




def create_shift(request):

    today = timezone.now().date()
    shift_data = CustomUser.objects.filter(is_active=True,is_employee=True).all()

    p2 = EmployeeMaster.objects.filter(user_id__in=[i.id for i in shift_data])

    shift_time = Shift.objects.all()

    return render(request,'hr/shift.html',{'shift_data':shift_data,'p2':p2,'shift_time':shift_time})






def List_Tickets_data(request):
    today = timezone.now().date()

    data60 = Tickets.objects.filter(employee_id=request.user.id).all()

    return render(request, 'hr/Tickets/list_tickets.html', {'data60':data60})



def Adding_Tickets_data(request):
    if request.method == 'POST':




        title = request.POST.get('title')
        date = request.POST.get('date')
        time = request.POST.get('time')
        priority = request.POST.get('priority')
        description = request.POST.get('description')
        obj = Tickets.objects.create(
                title=title,
                date=date,
                time=time,
                priority=priority,
                description=description,
                employee_id=request.user.id,

            )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_tickets')
    else:
        return render(request, 'hr/Tickets/add_tickets.html', )








def update_Tickets_data(request,id):
    if request.method == 'POST':


        title_upd = request.POST.get('title')
        date_upd = request.POST.get('date')
        time_upd = request.POST.get('time')
        priority = request.POST.get('priority')
        description_upd = request.POST.get('description')

       # print(title_upd,date_upd, description_upd)



        obj = Tickets.objects.get(id=id)


        obj.title = title_upd
        obj.date = date_upd
        obj.time = time_upd
        obj.priority = priority
        obj.description = description_upd


        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_tickets')
    else:

        obj = Tickets.objects.get(id=id)
        return render(request,'hr/Tickets/update_tickets.html',{"obj": obj})




def delete_Tickets_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Tickets.objects.get(id=id)
        except Tickets.DoesNotExist:
            return JsonResponse({'error': 'Tickets entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_tickets')
    else:
        return render(request, 'hr/Tickets/delete_tickets.html', )


def Ticket_reports(request):
    if request.user.is_authenticated:
        ticket_requests = Tickets.objects.filter(employee_id=request.user.id)
        context = {
        'ticket_requests': ticket_requests,
     }
    return render(request, 'hr/Tickets/tickets_reports.html', context)




def Ticket_reports_admin(request):
    if request.user.is_authenticated:
        ticket_requests = Tickets.objects.filter(status="Pending")
        context = {
        'ticket_requests': ticket_requests,
     }
    return render(request, 'hr/Tickets/tickets_reports_admin.html', context)




def Ticket_reports_admin_data(request):
    if request.user.is_authenticated:
        ticket_requests = Tickets.objects.all()
        context = {
        'ticket_requests': ticket_requests,
     }
    return render(request, 'hr/Tickets/tickets_reports_admin_data.html', context)



def Tickets_approval(request,id):
    tickets_approval = Tickets.objects.get(id=id)
    if request.method=="POST":
        approval_status = request.POST.get('approval_status')
        comments = request.POST.get('comments')
        tickets_approval.status=approval_status
        tickets_approval.comments=comments
        tickets_approval.save()
        messages.success(request, "Data Is Updated Successfully", extra_tags='alert alert-success')
        return redirect('tickets_reports_admin')
    else:
       return render(request,'hr/Tickets/tickets_approval.html',{'tickets_approval':tickets_approval})





def List_Expenses_data(request):
    today = timezone.now().date()

    data61 = Expenses.objects.filter(employee_id=request.user.id).all()

    return render(request, 'hr/Expenses/list_expenses.html', {'data61':data61})




def Adding_Expenses_data(request):
    if request.method == 'POST':
        date = datetime.datetime.now().strftime("%H:%M")

        title = request.POST.get('title')
        time = request.POST.get('time', date)

        date = request.POST.get('date')
        purpose = request.POST.get('purpose')
        amount = request.POST.get('amount')
        #time = request.POST.get('time')
        document = request.POST.get('document')
        reason = request.POST.get('reason')

        obj = Expenses.objects.create(
                title=title,
                date=date,
                purpose=purpose,
                amount=amount,
                document=document,
                description=reason,
                time=time,
                employee_id=request.user.id,

            )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_expenses')
    else:
        return render(request, 'hr/Expenses/add_expenses.html', )







def update_Expenses_data(request,id):
    obj = Expenses.objects.get(id=id)

    if request.method == 'POST':
        date = datetime.datetime.now().strftime("%H:%M")
        time = request.POST.get('time', date)



        title = request.POST.get('title')
        date = request.POST.get('date')
        purpose = request.POST.get('purpose')
        amount = request.POST.get('amount')
        #time = request.POST.get('time')
        reason = request.POST.get('reason')


        document = request.FILES.get('document')
        if document:
            obj.document=document


       # print(title_upd,date_upd, description_upd)

        obj.title = title
        obj.date = date
        obj.purpose = purpose
        obj.amount = amount
        obj.time = time
        obj.description = reason


        obj.save()
        messages.success(request, "Data Updated Successfully", extra_tags='alert alert-success')
        return redirect('list_expenses')
    else:

        return render(request,'hr/Expenses/update_expenses.html',{"obj": obj})




def delete_Expenses_data(request,id):

    if request.method == 'POST':


      #  holidays_id = request.POST.get('holidays_id')

        try:
            obj = Expenses.objects.get(id=id)
        except Expenses.DoesNotExist:
            return JsonResponse({'error': 'Expenses entry not found.'}, status=400)

        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_expenses')
    else:
        return render(request, 'hr/Expenses/delete_expenses.html', )





def Expenses_reports(request):
    if request.user.is_authenticated:
        expenses_requests = Expenses.objects.filter(employee_id=request.user.id)
        context = {
        'expenses_requests': expenses_requests,
     }
    return render(request, 'hr/Expenses/expenses_reports.html', context)




def Expenses_reports_admin(request):
    if request.user.is_authenticated:
        expenses_requests = Expenses.objects.filter(status="Pending")
        context = {
        'expenses_requests': expenses_requests,
     }
    return render(request, 'hr/Expenses/expenses_reports_admin.html', context)




def Expenses_reports_admin_data(request):
    if request.user.is_authenticated:
        expenses_requests = Expenses.objects.all()
        context = {
        'expenses_requests': expenses_requests,
     }
    return render(request, 'hr/Expenses/tickets_reports_admin_data.html', context)



def Expenses_approval(request,id):
    expenses_approval = Expenses.objects.get(id=id)
    if request.method=="POST":
        approval_status = request.POST.get('approval_status')
        comments = request.POST.get('comments')
        expenses_approval.status=approval_status
        expenses_approval.comments=comments
        expenses_approval.approved_by=request.user.username

        expenses_approval.save()
        messages.success(request, "Data Is Updated Successfully", extra_tags='alert alert-success')
        return redirect('expenses_reports_admin')
    else:
       return render(request,'hr/Expenses/expenses_approval.html',{'expenses_approval':expenses_approval})



def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        imported_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            BANK, IFSC, BRANCH, ADDRESS, CITY1, CITY2, STATE, STD_CODE, PHONE = row
            data_object = Rbi_Data_Import.objects.create(bank=BANK, ifsc=IFSC, branch=BRANCH, address=ADDRESS, city1=CITY1, city2=CITY2, state=STATE, std_code=STD_CODE, phone=PHONE)
            imported_data.append(data_object)

        return render(request, 'import_form.html', {'imported_data': imported_data})

    return render(request, 'import_form.html')





def List_financial_data(request):
    today = timezone.now().date()

    data62 = Financial_year.objects.all()
    return render(request, 'hr/Financial_year/list_financial.html', {'data62':data62})



def Adding_finanacial_data(request):
    if request.method == 'POST':
        financial_year = request.POST.get('financial_year')
        year = request.POST.get('year')


        obj = Financial_year.objects.create(
            financial_year=financial_year,
            year=year

        )

        obj.save()
        messages.success(request, "Data Is Added Successfully", extra_tags='alert alert-success')
        return redirect('list_financial_year')
    else:
        return render(request, 'hr/Financial_year/add_financial.html')



def delete_financial_data(request,id):
    if request.method == 'POST':
      #  holidays_id = request.POST.get('holidays_id')
        try:
            obj = Financial_year.objects.get(id=id)
        except Financial_year.DoesNotExist:
            return JsonResponse({'error': 'Financial Year entry not found.'}, status=400)
        obj.delete()
        messages.warning(request,"Data Deleted Successfully",extra_tags='alert alert-success alert-dismissible fade show')
        return redirect('list_financial_year')
    else:
        return render(request, 'hr/Financial_year/delete_financial.html')


